
# -*- coding: utf-8 -*-

import os
import xlrd
import xlwt
import openpyxl
import xlsxwriter
from xlsxwriter.workbook import Workbook
import datetime
import re
import time
import shutil
import logging
from shutil import copyfile
# from openpyxl.styles import colors

snow = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
logging.basicConfig(format='%(asctime)s %(levelname)s:%(message)s',
                    filename=os.getcwd().replace("\Protocol", "") + '\log\\' + snow + 'ooptest.log', level=logging.DEBUG)

#能源稳定性代理成功率
ATAGENTSULIST = ['name','successrate']
#公变所用测试项
PUBLIKTENTRY = ['冻结','事件上报','以太网和4G登录','代理成功率']
#专变所有测试项
MAJENTRY = ['冻结','事件上报','以太网和4G登录']
#打开一个excel文件
def open_xls(file):
    f = xlrd.open_workbook(file)
    return f

#获取excel中所有的sheet表
def getsheet(f):
    return f.sheets()

#获取sheet表的行数
def get_Allrows(f,sheet):
    table=f.sheets()[sheet]
    return table.nrows

#读取文件内容并返回行内容
def getFileori(file,shnum,datavalue):
    # 存储所有读取的结果
    f=open_xls(file)
    table=f.sheets()[shnum]
    num=table.nrows
    for row in range(num):
        rdata=table.row_values(row)
        datavalue.append(rdata)
    return datavalue

#读取表格第一行数据内容并返回行内容
def getFile0(file,shnum,datavalue):
    # 存储所有读取的结果
    f=open_xls(file)
    table=f.sheets()[shnum]
    num=1
    for row in range(num):
        rdata=table.row_values(row)
        datavalue.append(rdata)
    return datavalue
#获取sheet表的个数
def getshnum(f):
    x=0
    sh=getsheet(f)
    for sheet in sh:
        x+=1
    return x
#带标点符号的文字，筛选为只剩数字。例如：'0.94：怎么办；；出现异常；；出现异常。'-----0.94
def find_unchinese(file):
    pattern = re.compile(r'[\u4e00-\u9fa5]')
    punctuation = re.compile(r'[\u3002\uff1b\uff0c\uff1a\u201c\u201d\uff08\uff09\u3001\uff1f\u300a\u300b]')
    chinese = re.sub(pattern, '', file)
    num = re.sub(punctuation, '', chinese)
    return num
def file_name(file_dir):
    for root, dirs, files in os.walk(file_dir):
        # print(root)  # 当前目录路径
        # print(dirs)  # 当前路径下所有子目录
        # print(files)  # 当前路径下所有非目录子文件
        return root,files
#一个或多个表格中的第一行数据存储在一个列表中（为即将生成的表格做准备,这里只取了第一行，用来作为列的字典的键值）
def oneormoreexceltolist0(allxls):
    # 存储所有读取的结果
    datavalue = []
    rvalue = []
    for fl in allxls:
        f = open_xls(fl)
        x = getshnum(f)
        for shnum in range(x):
            # print("正在读取文件：" + str(fl) + "的第" + str(shnum) + "个sheet表的内容...")
            rvalue = getFile0(fl, shnum,datavalue)
    return rvalue
#一个或多个表格中的第一行数据存储在一个列表中（为即将生成的表格做准备,这里只取了第一行，用来作为列的字典的键值）
def oneormoreexceltolistori(allxls):
    # 存储所有读取的结果
    datavalue = []
    rvalue = []
    for fl in allxls:
        f = open_xls(fl)
        x = getshnum(f)
        for shnum in range(x):
            # print("正在读取文件：" + str(fl) + "的第" + str(shnum) + "个sheet表的内容...")
            rvalue = getFileori(fl, shnum,datavalue)
    return rvalue


#xlrd 为只读库，不用担心（不同线程之间）资源占用的问题  20211117
def readexcel(lfile, sname, etype,DAYDATAROWLIST):
    workbook = xlrd.open_workbook(lfile)
    print(workbook)
    rtem = {}
    for ns in workbook.sheet_names():   #excel文件
        sheet2 = workbook.sheet_by_name(ns)   #sheet页
        if ns == sname:
            for ir in range(0, sheet2.nrows):  # excel行数循环
                rowvalue = {}
                y = 0
                if etype == "dayData":
                    rowlist = DAYDATAROWLIST #列名称
                    seq = 0
                    for y in rowlist:
                        rowvalue[y] = sheet2.cell(ir, seq).value
                        seq += 1
                    rtem[ir] = rowvalue
                elif etype == "hplcData":
                    rowlist = DAYDATAROWLIST #列名称
                    seq = 0
                    for y in rowlist:
                        rowvalue[y] = sheet2.cell(ir, seq).value
                        seq += 1
                    rtem[ir] = rowvalue
                else:
                    print('增加表格列类型')
    return rtem

# 判断交采日冻结结果是否合格datafiles = {'DayResult': u'交采日冻结.xlsx'}

# 检查所需配置文件是否在路径目录下存在
def checkfl(originalpath,sheetname,DAYDATAROWLIST):
    daydata = readexcel(originalpath, sheetname, 'dayData',DAYDATAROWLIST)
    if daydata != '':
        return daydata
    else:
        return None

#如果没有记录过不合格，就记录一下。并且将不合格标志位置1
def unqualifiedrecord(unqualified,sheetkeyjudeg,key):
    if unqualified == 0:
        sheetkeyjudeg[0].append(key + '不合格')
        unqualified = 1
    return sheetkeyjudeg

#分析完成后，不合格标志位没有被置一，说明该项全部合格，给出合格结论。
def qualified(unqualified,sheetkeyjudeg,key):
    if unqualified == 0:
        sheetkeyjudeg[1].append(key + '合格')
    return sheetkeyjudeg

def checkEnergyConsum(key, avoid,daydata,sheetkeyjudeg):#检查用电量走字，形参为键值和允许偏差
    # daydata = checkfl()
    powerConsumBase = 0
    powerConsum = 0
    row = len(daydata) #多少行
    unqualified = 0
    for item in range(1,row):
        try:
            if item < row -1:
                powerBefore = float(find_unchinese(daydata[item][key]))
                powerAfter = float(find_unchinese(daydata[item + 1][key]))
                powerConsum = powerAfter - powerBefore #数据偏差
                if(powerConsum < 0):
                    daydata[item+1][key] =  daydata[item+1][key] + '；出现异常：电量倒走'
                    print(item+1,key,'电量倒走，需要标注')
                    sheetkeyjudeg = unqualifiedrecord(unqualified, sheetkeyjudeg, key)
            if(item == 1):
                powerConsumBase = powerConsum
            diff = abs(powerConsum - powerConsumBase)
            if(round(diff,2) > avoid) and  (item < row -1):
                daydata[item + 1][key] =  daydata[item + 1][key] +'；出现异常：电量变化量不符合预期'
                print(item+1,key,'电量变化量不符合预期，需要标注')
                sheetkeyjudeg = unqualifiedrecord(unqualified, sheetkeyjudeg, key)
        except:
            daydata[item + 1][key] = daydata[item + 1][key] + '；出现异常：电量数据项未召测到'
            sheetkeyjudeg = unqualifiedrecord(unqualified, sheetkeyjudeg, key)
    qualified(unqualified, sheetkeyjudeg, key)

    print(key,'电量变化量比较结束')

#检查总和分费率的关系,或者总和分相，形参为键值(表格每一行的键值)和允许偏差，num=4代表总和分相；num=5代表总和分费率
def checkPowerTotalPart(key, key1, key2, key3, key4, avoid, num,daydata,sheetkeyjudeg):
    row = len(daydata)  # 多少行
    if (num == 4):
        keyTemp = [key1, key2, key3, key]
    elif (num == 5):
        keyTemp = [key1, key2, key3, key4, key]
    power = []
    colTemp = [0,0,0,0,0]
    powerAdd = 0
    unqualifiedflag = 0
    for item in range(1, row-1):
        try:
            #分相或分费率相加
            for i in range(num-1):
                keyTail = keyTemp[i][7:8]
                if (int(keyTail) == 0):  # 组合电量
                    lng = find_unchinese(daydata[item][keyTemp[i]])
                    lng = lng.strip('[]').split("|")
                    power = [float(x) for x in lng] #总和四个费率
                    powerAdd += power[0] #计算总电能和
                else:
                    lng = float( find_unchinese(daydata[item][keyTemp[i]]))
                    powerAdd += lng
            #计算总
            keyTail = keyTemp[num-1][7:8]
            if (int(keyTail) == 0):  # 组合电量
                lng2 =  find_unchinese(daydata[item][keyTemp[num-1]])
                lng2 = lng.strip('[]').split("|")
                power = [float(x) for x in lng2]  # 总和四个费率
                powerTotal = power[0]  # 计算分相总电能和
            else:
                powerTotal = float(find_unchinese(daydata[item][keyTemp[num-1]]))
            diff = abs(powerTotal - powerAdd)#总和分相，或者总和分费率比较
            if (round(diff,2) > avoid):
                for m in range(0, num):
                    daydata[item][keyTemp[m]] = daydata[item][keyTemp[m]] + '；出现异常：总和分相或分费率不相等'
                    if(num == 4):
                        if unqualifiedflag == 0:
                            sheetkeyjudeg[0].append(str(keyTemp) + '不合格')
                        print(item, keyTemp[m], '总和分相和不相等')
                        unqualifiedflag = 1

                    else:
                        if unqualifiedflag == 0:
                            sheetkeyjudeg[0].append(str(keyTemp) + '不合格')
                        print(item, keyTemp[m], '总和分费率和不相等')
                        unqualifiedflag = 1
            powerAdd = 0
        except:
            for inum in range(num):
                daydata[item][keyTemp[inum]] = daydata[item][keyTemp[inum]] + '；出现异常：总和分相或分费率数据项未召测到'
            if unqualifiedflag == 0:
                sheetkeyjudeg[0].append(str(keyTemp) + '不合格')
            unqualifiedflag = 1
    if unqualifiedflag == 0:
        sheetkeyjudeg[1].append(str(keyTemp) + '合格')
    if(num == 4):
        print('总和分相电量比较结束')
    elif(num == 5):
        print('总和分费率比较结束')



def checkEnergyTotalrate(key, avoid,daydata,sheetkeyjudeg):#检查电能组合的正确性，包括总和分费率的关系，以及总的走字
    # daydata = checkfl()
    row = len(daydata)  # 多少行
    power = []
    powerRow = []
    powerAdd = 0
    powerConsum = 0
    powerConsumBase = 0
    unqualified = 0
    for item in range(1,row ):
        try:
            lng = find_unchinese(daydata[item][key])
            lng = lng.strip('[]').split("|")
            power = [float(x) for x in lng] #电能集合的总和四个费率
            diff = abs(power[0] - (power[1] + power[2] + power[3] + power[4]))#分相总和费率和比较
            powerRow.append(power[0])
            if (round(diff,2) > avoid):
                daydata[item][key] =  daydata[item][key] + '；出现异常：总与分费率之和不相等'
                print('总与分费率之和不相等',item,key)
                if unqualified == 0:
                    sheetkeyjudeg[0].append(key + '不合格')
                    unqualified = 1
            if (item < row) and (item > 0):
                # 数据偏差(看2行之间是否有倒走)
                if len(powerRow) >=2:
                    powerConsum = powerRow[len(powerRow)- 1] - powerRow[len(powerRow)- 2]
                if (powerConsum < 0):
                    daydata[item][key] = daydata[item][key] + '；出现异常：电量倒走'
                    print(item, key, '电量倒走')
                    if unqualified == 0:
                        sheetkeyjudeg[0].append(key + '不合格')
                        unqualified = 1
                # 取一个误差基准(如果误差基准错误，恰好说明走字有问题)
                if (item == 2):
                    powerConsumBase = powerConsum
                #2行相减的差都与最开始取到的基准（最开始2个数据的差）比较，不能大于avoid(允许误差)
                diff = abs(powerConsum - powerConsumBase)
                if (round(diff,2) > round(avoid,2)):
                    daydata[item][key] = daydata[item][key] + '；出现异常：电量变化量不符合预期'
                    print(item , key, '电量变化量不符合预期')
                    if unqualified == 0:
                        sheetkeyjudeg[0].append(key + '不合格')
                        unqualified = 1
        except:
            daydata[item][key] = daydata[item][key] + '；出现异常：电量数据项；出现异常'
            if unqualified == 0:
                sheetkeyjudeg[0].append(key + '不合格')
                unqualified = 1
    if unqualified == 0:
        sheetkeyjudeg[1].append(key + '合格')
    print(key, '电量集合变化量比较结束')


#不同点之间保持恒定值 比较函数（总召、分召均适用）
def checkvaTotalrate(key, avoid,name,daydata,sheetkeyjudeg):#检查电能组合的正确性，包括总和分费率的关系，以及总的走字
    row = len(daydata)  # 多少行
    power = []
    powerConsum = 0
    unqualified =0
    for item in range(1,row):
        try:
            lng = find_unchinese(daydata[item][key])
            lng = lng.strip('[]').split("|")
            power = [float(x) for x in lng] #变量类和功率类  总和分相
            #如果是功率类，并且是集合的情况下，要进行总和分相之间的比较
            if (name == '功率类') and  (len(power) >1):
                diff = abs(power[0] - (power[1] + power[2] + power[3]))  # 分相总和费率和比较
                if (round(diff, 2) > avoid):
                    daydata[item][key] = daydata[item][key] + '；出现异常：'+ name + '总与分相之和不相等'
                    print(name + '总与分相之和不相等', item, key)
                    unqualifiedrecord(unqualified,sheetkeyjudeg,key)
            if (item < row) and (item > 0):
                # 数据偏差(看2行之间是否有偏差过大的情况)
                if item == 1 :
                    Coneach = power
                if item >= 2:
                    if len(power) == 1:
                        diff = abs(power[0] - Coneach[0])
                        if (round(diff,2) > round(avoid,2)):
                            daydata[item][key] = daydata[item][key] + '；出现异常：'+ name + '不符合预期'
                            print(item , key,  name + '不符合预期')
                            unqualifiedrecord(unqualified, sheetkeyjudeg, key)
                    elif len(power) >= 1:
                        for Coneachnum in range(len(Coneach)):
                            diff = abs(power[Coneachnum] - Coneach[Coneachnum])
                            if (round(diff, 2) > round(avoid, 2)):
                                daydata[item][key] = daydata[item][key] + '；出现异常：'+ name + '不符合预期'
                                print(item, key, '变量类不符合预期')
                                unqualifiedrecord(unqualified, sheetkeyjudeg, key)
                                break
        except:
            daydata[item][key] = daydata[item][key] + '；出现异常：' + name + '数据项未召测到'
            unqualifiedrecord(unqualified, sheetkeyjudeg, key)

    print(key, name +'比较结束')
    qualified(unqualified, sheetkeyjudeg, key)
    return None
#比较变量类和功率因数是否合格；根据不同的数据类型，给误差范围。
def eachphase(key, avoid,name1,name2,daydata,sheetkeyjudeg):
    if key.find('200A') >= 0:
        checkvaTotalrate(key, 0.01,name2,daydata,sheetkeyjudeg)
    elif key[0:4] in ['2000','2002']:
        checkvaTotalrate(key, 5,name1,daydata,sheetkeyjudeg)
    elif key[0:4] in ['2001']:
        checkvaTotalrate(key, 1,name1,daydata,sheetkeyjudeg)

def psum(key, avoid,name,daydata,sheetkeyjudeg):
    #不同点间的值保持恒定，集合召测的形式，总和分相之间的比较
    checkvaTotalrate(key, 0.01,name,daydata,sheetkeyjudeg)
    #总和分相之和相等（01,02,03,04分开召测的形式）
    keyHead = key[0:3]
    key4 = int(key[3:4], 16)
    keyTail = key[7:8]
    if ((keyHead == '200') and (keyTail == '1') and (key4 >= 4) and (key4 <= 9)):
        powerTotalPhase(daydata,sheetkeyjudeg)







#电压、电流、功率、需量等测试过程中保持不变的量 比较
def constantdata(daydata,sheetkeyjudeg):
    for key in daydata[0]:
        try:
            keyHead = key[0:3]
            keyTail = int(key[3:4],16)
            if (keyHead == '200') and ((keyTail <= 3) or (keyTail == 10)): #分相变量类和功率因数 每一个点之间基本保持恒定
                print('分相变量类和功率因数',key)
                eachphase(key, 0.01,'分相变量类','功率因数',daydata,sheetkeyjudeg)
            elif ((keyHead == '200')and ((keyTail >= 4)and(keyTail <= 9))): #功率类,总和分之间有关系
                print('功率类',key)
                psum(key, 0.01,'功率类',daydata,sheetkeyjudeg)
        except:
            logging.warning('表格第一行出现异常')

#键值为电能量的比较，包括单个电量的走字判断和集合电量的总和分费率的判断及走字判断
def EnergyConsum(daydata,sheetkeyjudeg):
    print('daydata:',daydata)
    for key in daydata[0]:
        try:
            keyHead = int(key[0:1])
            keyTail = int(key[7:8])
            if (keyHead == 0)and (keyTail == 0): #电能量集合
                print('集合',key)
                checkEnergyTotalrate(key, 0.01,daydata,sheetkeyjudeg)#比较电能集合中总和分费率的关系及走字判断
            elif ((keyHead == 0)and ((keyTail >= 1)and(keyTail <= 4))): #电能量
                print('非集合',key)
                checkEnergyConsum(key, 0.01,daydata,sheetkeyjudeg)
        except:
            logging.warning('表格第一行出现异常')

#电量总和分相比较
def EnergyTotalPhase(total1,total2,a1,a2,b1,b2,c1,c2,daydata,sheetkeyjudeg):
    keyEnergy = [0,0,0,0]
    for key in daydata[0]:
        keyHead = key[0:8]
        if((keyHead == total1) or (keyHead == total2)):
            keyEnergy[0] = key
        elif ((keyHead == a1) or (keyHead == a2)):
            keyEnergy[1] = key
        elif ((keyHead == b1) or (keyHead == b2)):
            keyEnergy[2] = key
        elif ((keyHead == c1) or (keyHead == c2)):
            keyEnergy[3] = key
    if((keyEnergy[0] != 0) and (keyEnergy[1] != 0) and (keyEnergy[2] != 0) and (keyEnergy[3] != 0)):
        checkPowerTotalPart(keyEnergy[0], keyEnergy[1], keyEnergy[2], keyEnergy[3], 0, 0.01, 4,daydata,sheetkeyjudeg)

#在指定的位置创建文件夹
def mkdir(path):
    #  去除首位空格
    path=path.strip()
    #  去除尾部 \ 符号
    path=path.rstrip("\\")
    # 判断路径是否存在 # 存在 True # 不存在 False
    isExists=os.path.exists(path) # 判断结果
    # 如果不存在则创建目录 　# 创建目录操作函数
    if not isExists:
        os.makedirs(path)
        print(path+' 创建成功')
        return True
    else: # 如果目录存在则不创建，并提示目录已存在
        print(path+' 目录已存在')
        return False



#电量总和分相的关系判断
def EnergyTotalPhasecom(daydata,sheetkeyjudeg):
    EnergyTotalPhase('00100201', '00100200', '00110201', '00110200', '00120201', '00120200', '00130201', '00130200',daydata,sheetkeyjudeg)
    EnergyTotalPhase('00200201', '00200200', '00210201', '00210200', '00220201', '00220200', '00230201', '00230200',daydata,sheetkeyjudeg)
    EnergyTotalPhase('00300201', '00300200', '00310201', '00310200', '00320201', '00320200', '00330201', '00330200',daydata,sheetkeyjudeg)
    EnergyTotalPhase('00400201', '00400200', '00410201', '00410200', '00420201', '00420200', '00430201', '00430200',daydata,sheetkeyjudeg)
    #如有需要，可扩展其它比较项

#功率类总和分相之间比较（分别召测的总和分相功率1、2、3、4，在已经找到总的情况下，找3个分相功率）
def powerTotalPhase(daydata,sheetkeyjudeg):
    for key in daydata[0]:
        keyHead = key[0:3]
        key4 = int(key[3:4],16)
        keyTail = key[7:8]
        if((keyHead == '200') and (keyTail == '1') and(key4 >= 4)and(key4 <= 9)):
            keyEnergy = [0, 0, 0, 0]
            str2 = ['1', '2', '3', '4']
            str1 = key[0:7]
            for key2 in daydata[0]:
                keyID = key2[0:8]
                for i in range(len(str2)):
                    if(keyID == str1 + str2[i]):
                        keyEnergy[i] = key2
            if((keyEnergy[0] != 0) and (keyEnergy[1] != 0) and (keyEnergy[2] != 0) and (keyEnergy[3] != 0)):
                checkPowerTotalPart(keyEnergy[0], keyEnergy[1], keyEnergy[2], keyEnergy[3],0,  0.01, 4,daydata,sheetkeyjudeg)



#总和分费率的关系判断（分别召测的总和费率1、2、3、4，在已经找到总的情况下，找4个费率）
def JudgeTotalpart(daydata,sheetkeyjudeg):
    for key in daydata[0]:
        keyHead = key[0:1]
        keyTail = key[7:8]
        if((keyHead == '0') and (keyTail == '1')):
            keyEnergy = [0, 0, 0, 0, 0]
            str2 = ['1', '2', '3', '4', '5']
            str1 = key[0:7]
            for key2 in daydata[0]:
                keyID = key2[0:8]
                for i in range(len(str2)):
                    if(keyID == str1 + str2[i]):
                        keyEnergy[i] = key2
            if((keyEnergy[0] != 0) and (keyEnergy[1] != 0) and (keyEnergy[2] != 0) and (keyEnergy[3] != 0)and (keyEnergy[4] != 0)):
                checkPowerTotalPart(keyEnergy[0], keyEnergy[1], keyEnergy[2], keyEnergy[3], keyEnergy[4], 0.01, 5,daydata,sheetkeyjudeg)

#分析后的报告存储  20211110  WM
def analysisexcel(temfile,daydatalist,sheetnamelist,columnlistlist):
    wb = Workbook(temfile)
    for daydatalistnum in range(len(daydatalist)):
        ws0 = wb.add_worksheet(sheetnamelist[daydatalistnum])
        rowlist = columnlistlist[daydatalistnum]
        if '' in rowlist:
            rowlist.remove('')
        tlist = daydatalist[daydatalistnum]
        for key in tlist:
            counter = 0
            # y = 0
            for y in rowlist:
                try:
                    if (tlist[key][y].find("；出现异常") >= 0):
                        bold = wb.add_format({'color': 'red', 'bold': 1, 'bg_color': '#FFFF00'})
                        ws0.write(key, counter, tlist[key][y], bold)
                    else:
                        ws0.write(key, counter, tlist[key][y])
                except:
                    # logging.info('硬件自动化测试表格追加时出现错误')
                    if isinstance(tlist[key][y], list):
                        strcuv = ''
                        for i in range(len(tlist[key][y])):
                            strcuv += str(tlist[key][y][i]) + '\n'
                        ws0.write(key, counter, strcuv)
                counter += 1
    wb.close()


def demansinglecom(lng,item,key,timeold,valueold,daydata,timekey):
    #当需量值为0时的发生时间
    timenull = '20000101000000'
    if lng.find(',') >= 0:
        lng = lng.strip('[]').split(",")
    elif lng.find('|') >= 0:
        lng = lng.strip('[]').split("|")
    else:
        print('增加其它需量处理形式')
    devalue = lng[0]
    detime = lng[1].strip('"').strip("'")
    if float(devalue) == 0:
        if detime != timenull:
            daydata[item][key] = daydata[item][key] + '；出现异常：需量值为0时，需量时间错误'

    else:
        # 第一次给有值的需量和正常的需量发生时间赋值
        if float(valueold) == 0:
            timeold = int(detime)
            valueold = float(devalue)
            if detime == timenull:
                daydata[item][key] = daydata[item][key] + '；出现异常：需量发生时间记录错误；'
        else:
            if (float(devalue) < valueold) and (daydata[item][timekey][5:7] == daydata[item-1][timekey][5:7]):
                daydata[item][key] = daydata[item][key] + '；出现异常：需量变化量错误；'
            else:
                # 比较完成后，重新赋值给值的需量和正常的需量发生时间。用于下次比较
                valueold = float(devalue)
            if int(detime) < timeold or detime == timenull:
                daydata[item][key] = daydata[item][key] + '；出现异常：需量发生时间记录错误；'
            else:
                timeold = int(detime)
    return timeold,valueold

#需量，非集合召测时，时间和值的比较
def checkdemandsingle(key,daydata,timekey,sheetkeyjudeg):
    row = len(daydata)  # 多少行
    timeold = 0
    valueold = 0
    for item in range(1, row):
        try:
            lng = find_unchinese(daydata[item][key])
            timeold,valueold = demansinglecom(lng,item,key,timeold,valueold,daydata,timekey)
        except:
            daydata[item][key] = daydata[item][key] + '；出现异常：数据项未召测到'

    print(key, '需量和时间比较完成')

def checkdemandall(key,daydata,timekey,sheetkeyjudeg):
    row = len(daydata)  # 多少行
    # 用来存储第一次有值时的需量和时间
    timeold = [0,0,0,0,0]
    valueold = [0,0,0,0,0]
    timenull = '20000101000000'
    unqualified = 0
    for item in range(1, row):
        try:
            lng = find_unchinese(daydata[item][key])
        #需量集合召测上来的方式有2种
        #[[0.0,'20000101000000']|[0.0,'20000101000000']|[0.0,'20000101000000']|[0.0,'20000101000000']|[0.0,'20000101000000']]
        # 0.0|20000101000000#0.0|20000101000000#0.0|20000101000000#0.0|20000101000000#0.0|20000101000000
            if lng.find('#') >= 0:
                lng = lng.split("#")
            else:
                lng = lng.split("|")
            for itemnum in range(len(lng)):
                print('lng:',lng)
                if lng[itemnum].find(',') >= 0:
                    lng[itemnum] = lng[itemnum].strip('[]').split(",")
                elif lng[itemnum].find('|') >= 0:
                    lng[itemnum] = lng[itemnum].strip('[]').split("|")
                else:
                    print('增加其它需量处理形式')
                devalue = lng[itemnum][0]
                detime = lng[itemnum][1].strip('"').strip("'")
                if float(devalue) == 0:
                    if detime != timenull:
                        daydata[item][key] = daydata[item][key] + '；出现异常：需量值为0时，需量时间错误'
                        unqualifiedrecord(unqualified,sheetkeyjudeg,key)
                        break
                else:
                    # 第一次给有值的需量和正常的需量发生时间赋值
                    if float(valueold[itemnum]) == 0:
                        timeold[itemnum] = int(detime)
                        valueold[itemnum] = float(devalue)
                        #一旦出现错误就跳出循环，不再查找其它错误
                        if detime[itemnum] == timenull:
                            daydata[item][key] = daydata[item][key] + '；出现异常：需量发生时间记录错误；'
                            unqualifiedrecord(unqualified, sheetkeyjudeg, key)
                        break
                    else:
                        if (float(devalue) < valueold[itemnum]) and (daydata[item][timekey][5:7] == daydata[item-1][timekey][5:7] ) and (daydata[item][timekey][7:9] != '02') :
                            daydata[item][key] = daydata[item][key] + '；出现异常：需量变化量错误；'
                            unqualifiedrecord(unqualified, sheetkeyjudeg, key)
                            break
                        else:
                            # 比较完成后，重新赋值给值的需量和正常的需量发生时间。用于下次比较
                            valueold[itemnum] = float(devalue)
                        if int(detime) < timeold[itemnum] or detime == timenull:
                            daydata[item][key] = daydata[item][key] + '；出现异常：需量发生时间记录错误；'
                            unqualifiedrecord(unqualified, sheetkeyjudeg, key)
                            break
                        else:
                            timeold[itemnum] = int(detime)
        except:
            daydata[item][key] = daydata[item][key] + '；出现异常：需量数据项未召测到'
            unqualifiedrecord(unqualified, sheetkeyjudeg, key)
    qualified(unqualified, sheetkeyjudeg, key)



def demandtimeandvalue(daydata,sheetkeyjudeg):
    for key in daydata[0]:
        # try:
            keyHead = int(key[0:1])
            keyTail = int(key[7:8])
            if (keyHead == 1)and (keyTail == 0): #需量集合
                checkdemandall(key,daydata,'20210200数据冻结时间',sheetkeyjudeg)
                print('集合',key)
            elif ((keyHead == 1)and ((keyTail >= 1)and(keyTail <= 4))): #需量
                print('非集合',key)
                checkdemandsingle(key,daydata,'20210200数据冻结时间',sheetkeyjudeg)
        # except:
        #     logging.warning('表格第一行出现异常')

# 20191126 00 01 42 0000->   datetime
def bcdtodatetime(dt):
    dt.zfill(14)
    if len(dt) == 14:
        dt += '0000'
    tt = datetime.datetime(int(dt[0:4]), int(dt[4:6]), int(dt[6:8]), int(dt[8:10]),
                           int(dt[10:12]), int(dt[12:14]), int(dt[14:])*1000%1000000)
    return tt
#timestr1,timestr2:两个要比较的时间字符串形式；返回值为两者相差的秒数。（只能比到差30天23小时59分的时间差，无法比月）
def timecompare(timestr1,timestr2):
  if bcdtodatetime(timestr1) > bcdtodatetime(timestr2):
    timediffer = bcdtodatetime(timestr1) - bcdtodatetime(timestr2)
  else:
    timediffer = bcdtodatetime(timestr2) - bcdtodatetime(timestr1)
  print('timediffer:' + str(timediffer))
  return timediffer
#返回时间差的具体秒数
def differseconds(timestyle):
    timediffers= timestyle.days * 86400 + timestyle.seconds
    return timediffers

#2点：1、检查20210200是否按照时间间隔正常递增；2、检查不同sheet页直接，是否存在漏点的情况。
def pointjudgment(daydata,key,pointlist,sheetnamenum,sheetkeyjudeg):
    row = len(daydata)  # 多少行
    pointcomlist = []
    timeintervalbs = 0
    unqualified = 0
    #只有1行或2行时，进不到下面的循环，单独处理。只有1行，说明未召测到时数据；只有2行，要保证所有sheet的2021保持一致。
    try:
        if 1<=row <= 2 :
            if  row == 1 :
                daydata[0][key] = daydata[0][key] + "；出现异常：未召测到冻结数据"
                unqualifiedrecord(unqualified,sheetkeyjudeg,key)
            elif row == 2:
                pointcomlist.append(find_unchinese(daydata[1][key]).strip('"').strip("'").strip(" "))
                if pointlist != []:
                    if pointlist != pointcomlist:
                        daydata[1][key] = daydata[1][key] + "；出现异常：数据冻结时间异常或出现漏点"
                        unqualifiedrecord(unqualified, sheetkeyjudeg, key)
                else:
                    pointlist = pointcomlist
    except:
        daydata[0][key] = daydata[0][key] + "；出现异常：冻结数据未召测到，分析出现异常"
        unqualifiedrecord(unqualified, sheetkeyjudeg, key)


    else:
        for item in range(1, row -1):
            try:
                pointcomlist.append(find_unchinese(daydata[item][key]).strip('"').strip("'").strip(" "))
                point1time = find_unchinese(daydata[item][key]).strip('"').strip("'").strip(" ")
                point2time = find_unchinese(daydata[item + 1][key]).strip('"').strip("'").strip(" ")
                #先取到时间间隔，作为后面比较的基准
                if item == 1:
                    timeintervalbs = differseconds(timecompare(point1time,point2time))
                #2行之差要与基准时间间隔保持一致
                if item > 2:
                    #时间间隔大于10天，小于40天，说明是月冻结或结算日冻结  如果2个月的差值不为1或11，说明冻结时标异常
                    if  864000 < differseconds(timecompare(point1time,point2time)) < 3456000:
                        if  (abs(int(point1time[4:6]) - int(point2time[4:6])) != 1) and (abs(int(point1time[4:6]) - int(point2time[4:6])) != 11):
                            daydata[item + 1][key] = daydata[item + 1][key] + "；出现异常：出现漏点或数据冻结时间异常，请检查"
                            unqualifiedrecord(unqualified, sheetkeyjudeg, key)
                    else:
                        if differseconds(timecompare(point1time,point2time)) !=  timeintervalbs:
                            daydata[ item+1][key] = daydata[ item+1 ][key] + "；出现异常：出现漏点或数据冻结时间异常，请检查"
                            unqualifiedrecord(unqualified, sheetkeyjudeg, key)

            except:
                daydata[item + 1][key] = daydata[item + 1][key] + "；出现异常冻结数据未召测到，分析出现异常"
                unqualifiedrecord(unqualified, sheetkeyjudeg, key)
        try:
            #取点数多的为比较时的依据。第一个表格时，取好参考比较值。后面有更多的行时，才考虑更新参考值
            #总点数判断：如果后面的sheet页比前面的行数多，说明前面的sheet页有漏点.要更新比较值
            if sheetnamenum == 0 :
                pointlist = pointcomlist
            if (len(pointcomlist) > len(pointlist)) and (sheetnamenum != 0):
                if len(pointcomlist) == (len(pointlist) + 1) and (pointlist[0] == pointcomlist[0]):
                    daydata[1][key] = daydata[1][key] + "；出现异常：上一个sheet页出现漏点"
                    unqualifiedrecord(unqualified, sheetkeyjudeg, key)
                    pointlist = pointcomlist
                else:pass
            #如果2个sheet页的行数差大于1行，说明有漏点
            elif len(pointcomlist) < (len(pointlist) -1) :
                daydata[1][key] = daydata[1][key] + "；出现异常：这个sheet页中的冻结出现漏点"
                unqualifiedrecord(unqualified, sheetkeyjudeg, key)
            #2个sheet页的行数差为1时，判一下是不是因为第一次召测时任务还未执行导致，是就判合格，不是说明漏点
            elif len(pointcomlist) == (len(pointlist) -1)  and pointlist[0] == pointcomlist[0] :
                daydata[row -1][key] = daydata[row - 1][key] + "；出现异常：这个sheet页中的冻结出现漏点"
                unqualifiedrecord(unqualified, sheetkeyjudeg, key)
        except:
            daydata[1][key] = daydata[1][key] + "；出现异常：冻结数据分析出现异常"
            unqualifiedrecord(unqualified, sheetkeyjudeg, key)
    qualified(unqualified, sheetkeyjudeg, key)
    return pointlist





#判断是否漏点  最后分析时才调用
def losspoint(daydata,pointlist,sheetnamenum,sheetkeyjudeg):
    for key in daydata[0]:
        if key[:8] == '20210200':
            pointlist = pointjudgment(daydata,key,pointlist,sheetnamenum,sheetkeyjudeg)
    return pointlist

#追加生成报告
def write_excel_xlsx_jccomrow(filepath,allexcellist,sheetname):
    sheetnum = 0
    try:
        workbook = openpyxl.load_workbook(filepath)
        print('打开指定路径下的表格')
    except:
        workbook = openpyxl.Workbook()
        print('创建了一个空白的Excel')
    # 不使用第一个sheet,所以先删除。
    if "Sheet" not in workbook.sheetnames:
        pass
    else:
        ws = workbook["Sheet"]
        workbook.remove(ws)
    if sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]
    else:
        worksheet = workbook.create_sheet(sheetname)
    maxRows = worksheet.max_row  # 获取表格中已存在的数据的行数
    #处理表格的模块本身有问题，不应该在没有行的情况下默认为1行，所以程序做了特殊处理，强制处理为初始行为0行。
    if maxRows == 1 and worksheet['A1'].value == None:
        maxRows = 0
    for i in range(len(allexcellist)):
        value = allexcellist[i]
        index = len(value)  # 获取需要写入数据的行数
        # 追加写入数据，注意是从i+rows_old行开始写入
        for j in range(0, index):
            try:
                worksheet.cell(row=maxRows + i + 1, column= j + 1, value=str(value[j]))
            except:
                worksheet.cell(row=maxRows + i + 1, column=j + 1, value=str(value[j]))
    #设置A列的列宽为100
    worksheet.column_dimensions['A'].width = 70
    #设置sheet（工作表）标签颜色
    if sheetname.find('不合格') >= 0:
        worksheet.sheet_properties.tabColor = 'FF0000'
    workbook.save(filepath)
    try:
        #如果想在执行过程中查看报告，需要在temp_report路径中拷贝出来查看，否则程序会因为文件被占用而报错，导致数据漏存。
        welookpath = (filepath.split('展示')[0] + '展示可拷贝查看' + ".xlsx").replace('report','temp_report')
        workbook.save(welookpath)
    except:
        print('不是展示类的报告，不走这个流程。忽略')
    return None

#每个表格都进行数据分析
def stablereport(eportfloderpath,sheetname,daydata,pointlist,sheetnamenum,sheetkeyjudeg):
    #如果是电能量,且存在总值再尝试去比较总和分费率的关系
    # daydata = checkfl()
    JudgeTotalpart(daydata,sheetkeyjudeg)
    #总和分费率的关系判断(单独分开召测的)
    EnergyTotalPhasecom(daydata,sheetkeyjudeg)
    #总和分相的关系判断(正向有功电能量)
    EnergyConsum(daydata,sheetkeyjudeg)
    #包括单个电量的走字判断和集合电量的总和分费率的判断及走字判断
    constantdata(daydata,sheetkeyjudeg)
    #电压、电流、功率、需量、相角、功率因数 等测试过程中保持不变的量 比较
    demandtimeandvalue(daydata,sheetkeyjudeg)
    # 需量时间比较、值比较、总和分费率（集合形式的比较）
    pointlist = losspoint(daydata,pointlist,sheetnamenum,sheetkeyjudeg)
    print('sheetkeyjudeg:',sheetkeyjudeg)
    return pointlist,sheetkeyjudeg

def eventtimetotal(workbook,nums,name):
    rows = workbook.sheet_by_name(name)
    nrow = rows.nrows
    nums = nums + (nrow -1)
    return nums

def eventtimesrecord(pathlist,staecandpceventnum,eventresult,resultflag):
    eventtimes = [0, 0, 0, 0, 0, 0]
    for eachpathnum in range(len(pathlist)):
        workbook = xlrd.open_workbook(pathlist[eachpathnum])
        sheetnamelist = workbook.sheet_names()
        print(sheetnamelist)
        for item in sheetnamelist:
            if item == '32020200购电参数设置记录':
                eventtimes[0] = eventtimetotal(workbook, eventtimes[0], '32020200购电参数设置记录')
            elif item == '32030200电控告警记录':
                eventtimes[1] = eventtimetotal(workbook, eventtimes[1], '32030200电控告警记录')
            elif item == '32010200电控跳闸记录':
                eventtimes[2] = eventtimetotal(workbook, eventtimes[2], '32010200电控跳闸记录')
            elif item == '32000200功控跳闸记录':
                eventtimes[3] = eventtimetotal(workbook, eventtimes[3], '32000200功控跳闸记录')
            elif item == '31040200遥信变位记录':
                eventtimes[4] = eventtimetotal(workbook, eventtimes[3], '31040200遥信变位记录')
            elif item == '31200200回路巡检变位记录':
                eventtimes[5] = eventtimetotal(workbook, eventtimes[3], '31200200回路巡检变位记录')
    print('eventtimes:', eventtimes)
    if staecandpceventnum[0] == 0:
        pass
    else:
        #购电参数
        if (eventtimes[0] == staecandpceventnum[0]+1):
            eventresult[0] = ['购电参数设置事件上报次数为'+ f'{eventtimes[0]}'+ '【合格】']
        else:
            resultflag = 1
            eventresult[0] = ['购电参数设置事件预期次数为：' + f'{staecandpceventnum[0]+1}'+ ',实际次数为：' + f'{eventtimes[0]}' + '【不合格！！！！！！！！！！！！！！！！！！！！】']
        #电控告警
        if eventtimes[1] == staecandpceventnum[0]:
            eventresult[1] = ['电控告警事件上报次数为'+ f'{eventtimes[1]}'+'【合格】']
        else:
            eventresult[1] = ['电控告警事件预期次数为：' + f'{staecandpceventnum[0]}' + ',实际次数为：' + f'{eventtimes[1]}'+'【不合格！！！！！！！！！！！！！！！！！！！！】']
            resultflag = 1
        #电控跳闸
        if eventtimes[2] == staecandpceventnum[0]:
            eventresult[2] = ['电控跳闸事件上报次数为'+ f'{eventtimes[2]}'+'【合格】']
        else:
            eventresult[2] = ['电控跳闸事件预期次数为：' + f'{staecandpceventnum[0]}' + ',实际次数为：' + f'{eventtimes[2]}'+'【不合格！！！！！！！！！！！！！！！！！！！！】']
            resultflag = 1
        #功控跳闸
        if eventtimes[3] == staecandpceventnum[1]:
            eventresult[3] = ['功控跳闸事件上报次数为'+ f'{eventtimes[3]}'+'【合格】']
        else:
            eventresult[3] = ['功控跳闸事件预期次数为：' + f'{staecandpceventnum[1]}' + ',实际次数为：' + f'{eventtimes[3]}'+'【不合格！！！！！！！！！！！！！！！！！！！！】']
            resultflag = 1
    #遥信
    if eventtimes[4] == 0 :
        eventresult[4] = ['遥信变位事件上报次数为'+ f'{eventtimes[4]}'+'【合格】']
    else:
        eventresult[4] = ['遥信变位事件预期次数为：0 ,实际次数为：' + f'{eventtimes[4]}'+'【不合格！！！！！！！！！！！！！！！！！！！！】']
        resultflag = 1
    #回路巡检
    if eventtimes[5] == 0:
        eventresult[5] = ['回路巡检变位事件上报次数为' + f'{eventtimes[5]}'+'【合格】']
    else:
        eventresult[5] = ['回路巡检变位事件预期次数为：0 ,实际次数为：' + f'{eventtimes[5]}'+'【不合格！！！！！！！！！！！！！！！！！！！！】']
        resultflag = 1
    print('eventresult:',eventresult)
    return eventtimes,eventresult,resultflag

def eventdeal(imeventpathlist,imeacheventnamelist,teeventpathlist,teeacheventnamelist,staecandpceventnum,stasavepath):
    # 32020200购电参数设置记录  32030200电控告警记录  32010200电控跳闸记录  32000200功控跳闸记录 ,
    # 31040200遥信变位记录，31200200回路巡检变位记录
    imeventtimes = [0, 0, 0, 0, 0, 0]
    teimeventtimes = [0, 0, 0, 0, 0, 0]
    imeventresult = [[],[],[],[],[],[]]
    teeventresult = [[],[],[],[],[],[]]
    imflag = 0
    teflag = 0
    imeventtimes,imeventresult,imflag = eventtimesrecord(imeventpathlist,staecandpceventnum,imeventresult,imflag)
    teeventtimes,teeventresult,teflag = eventtimesrecord(teeventpathlist,staecandpceventnum,teeventresult,teflag)
    #事件上报最终结果处理流程
    if imflag == 0:
        imeventresult.insert(0,['主通道事件上报【合格】：'])
    else:
        imeventresult.insert(0, ['主通道事件上报【不合格！！！！！！！！！！！！！！！！！！！！】：'])
    #插入空行和备注描述，方便报告查看
    teeventresult.insert(0, [''])
    teeventresult.insert(1, [''])
    if teflag == 0:
        teeventresult.insert(2, ['临时通道事件上报【合格】：'])
    else:
        teeventresult.insert(2, ['临时通道事件上报【不合格！！！！！！！！！！！！！！！！！！！！】：'])
    #如果imeventresult事件统计的结果里面没有进行控制类的统计，说明是公变，不统计这些，出报告时删掉即可
    if  imeventresult[1:5] == [[],[],[],[]]:
        del imeventresult[1]
        del imeventresult[1]
        del imeventresult[1]
        del imeventresult[1]
        del teeventresult[3]
        del teeventresult[3]
        del teeventresult[3]
        del teeventresult[3]
    if imflag == 0 and teflag == 0:
        write_excel_xlsx_jccomrow(stasavepath, imeventresult, '事件上报合格')
        write_excel_xlsx_jccomrow(stasavepath, teeventresult, '事件上报合格')
    else:
        write_excel_xlsx_jccomrow(stasavepath, imeventresult, '事件上报不合格')
        write_excel_xlsx_jccomrow(stasavepath, teeventresult, '事件上报不合格')
    return None

def rateget(expTime):
    a = r'率(.*?)%'
    slotList = re.findall(a, expTime)
    offset = float(slotList[0])
    return offset


def ratejudge(rate,rateexpect,resultlist,name,item,resultflag):
    if rate <= rateexpect:
        resultlist[ item-1] = [name + '代理成功率为：'+'{:.2%}'.format(rate/100)+'【不合格！！！！！！！！！！】']
        resultflag = 1
    else:
        resultlist[item-1] = [name + '代理成功率为：'+'{:.2%}'.format(rate/100)+'【合格】']
    return resultlist,resultflag

def newreportget(hplcratepathlist):
    datalist = []
    num = 0
    maxdata = 0
    for itemnum in range(len(hplcratepathlist)):
        datalist.append(int(dataagentget(hplcratepathlist[itemnum])))
    maxdata = datalist[0]
    for datanum in  range(len(datalist)):
        if datalist[datanum] > maxdata:
            maxdata = datalist[datanum]
            num = datanum
        else:
            pass
    return hplcratepathlist[num]




def dataagentget(expTime):
    a = r'率(.*?).xls'
    slotList = re.findall(a, expTime)
    offset = float(slotList[0])
    return offset

def hplcreadrate(hplcratepathlist,stasavepath):
    resultlist = [[],[],[],[],[],[]]
    ratepath = newreportget(hplcratepathlist)
    print('ratepath:',ratepath)
    data = readexcel(ratepath, 'sheet', 'hplcData',ATAGENTSULIST)
    resultflag = 0
    print('data:',data)
    for item in data:
        if item >0:
            rate = rateget(data[item]['successrate'])
            #485代理成功率要求为99，内控为99.5；；电力线载波要求为97，内控为97.5
            if (item ==1) or (item ==2):
                resultlist,resultflag = ratejudge(rate, 99.5, resultlist,data[item]['name'],item,resultflag)
            elif item >2  :
                resultlist,resultflag = ratejudge(rate, 97.5, resultlist,data[item]['name'],item,resultflag)
    if resultflag == 0:
        write_excel_xlsx_jccomrow(stasavepath, resultlist, '代理成功率合格')
    else:
        write_excel_xlsx_jccomrow(stasavepath, resultlist, '代理成功率不合格')


def excwrite(rowvalue,ews):
    for a in range(len(rowvalue)):
        for b in range(len(rowvalue[a])):
            c=rowvalue[a][b]
            ews.cell(row=a+1, column=b+1, value=c)
    return ews

#登录信息统计
def regreportdeal(regpathlist,stasavepath):
    datalist = oneormoreexceltolistori(regpathlist)
    print('datalist:',datalist)
    netlist =[['以太网登录情况统计：']]
    wirlist = [[],[],['4G登录情况统计：']]
    for item in datalist:
        if str(item).find('以太网') >= 0:
            netlist.append(item)
        else:
            wirlist.append(item)
    savedatalist = netlist + wirlist
    write_excel_xlsx_jccomrow(stasavepath, [], '以太网和4G登录不合格')
    wb = openpyxl.load_workbook(stasavepath)
    ws = wb['以太网和4G登录不合格']
    excwrite(savedatalist, ws)
    wb.save(stasavepath)

#处理 功率和正向有功，存储到列表中，方便取用
def pecreport(staecandpceventnum,stasavepath,overflag):
    pandecresultlist = ['', '']
    if stasavepath.find('专变') >= 0:
        if staecandpceventnum[2] != 0:
            pandecresultlist[0] = '总加组和脉冲功率【不合格！！！！！！！！！！】'
        else:
            pandecresultlist[0] = '总加组和脉冲功率【合格】'
        #脉冲计量（正向有功的召测），要等测试到最后才有，否则没有
        if overflag == 1:
            if staecandpceventnum[3] != 0:
                pandecresultlist[1] = '脉冲计数【不合格！！！！！！！！！！】'
            else:
                pandecresultlist[1] = '脉冲计数【合格】'
        else:
            del pandecresultlist[1]
        print('pandecresultlist:',pandecresultlist)
    else:pandecresultlist = []
    return pandecresultlist


#公变所用测试项
PUBLIKTENTRY = ['冻结','事件上报','以太网和4G登录','代理成功率']
#稳定性最终结论sheet页处理
def finalreport(stasavepath,staecandpceventnum,overflag):
    resultlist = []
    resulttruelist = [['测试结论：'],['【合格项】']]
    resultfalselist =[[],[],['【不合格项！！！！！】']]
    workbook = xlrd.open_workbook(stasavepath)
    allsheetname = workbook.sheet_names()
    freezefalseflag = 0
    print('allfilename',allsheetname)
    for item in allsheetname:
        if item.find('冻结不合格') >= 0:
            if '冻结合格报告' in allsheetname:
                allsheetname.remove('冻结合格')
    if '以太网和4G登录不合格' not in allsheetname :
        allsheetname.append('以太网和4G登录合格')
    allsheetname.remove('稳定性测试结论')
    resultlist = allsheetname
    #将所有的最终结论放在一个列表中，输出最后的结果
    resultlist.extend(pecreport(staecandpceventnum, stasavepath, overflag))
    for item in resultlist:
        if item.find('不合格') >= 0:
            resultfalselist.append([item])
        else:
            resulttruelist.append([item])
    write_excel_xlsx_jccomrow(stasavepath, resulttruelist, '稳定性测试结论')
    write_excel_xlsx_jccomrow(stasavepath, resultfalselist, '稳定性测试结论')








#冻结、稳定性结果处理流程（里面有分支来分析稳定性和计量）
#路径、预期值及结果带入
def anlaysispathreport(originalpath,staecandpceventnum,overflag):
    DAYDATAROWLIST = []
    daydata = []
    #每个表格里面应该有几个sheet页的判断：
    if originalpath.find('专变') >= 0:
        excelsheetnum = 9
    elif originalpath.find('公变') >= 0:
        excelsheetnum = 21
    else:
        excelsheetnum = 1
    #新建一个文件夹来保存分析后的报告
    #每次分析前，先清空之前的分析结果(先判断一下，如果路径存在，再进行删除操作)
    if os.path.exists(originalpath + '分析后报告\\'):
        shutil.rmtree(originalpath + '分析后报告\\')
    newpath = originalpath + '分析后报告\\'
    mkdir(newpath)
    #最终报告，路径和名称处理
    if originalpath.find('稳定') >= 0:
        stasavepath = newpath + "\\" + '稳定性最终报告.xlsx'
    else:
        stasavepath = newpath + "\\" + '计量冻结最终报告.xlsx'
    expath, allfilename = file_name(originalpath)
    #每个表格路径保存
    excelpathlist = []
    eachexcelnamelist = []
    #主通道事件报告路径
    imeventpathlist = []
    imeacheventnamelist = []
    #临时通道事件报告路径
    teeventpathlist = []
    teeacheventnamelist = []
    #HPLC代理成功率报告路径
    hplcratepathlist = []
    #存储是否有漏抄电表的情况
    problemlist = [[],[]]
    #登录异常信息报告路径
    regpathlist = []
    #稳定性结果处理比较复杂。计量只需要处理冻结数据即可
    if originalpath.find('稳定') >= 0:
        #先存储一个空的'稳定性测试结论'sheet页，方便结果查看
        write_excel_xlsx_jccomrow(stasavepath, [], '稳定性测试结论')
    for eachexcelname in allfilename:
        #根据报告名称进行区分。计量冻结类报告，原报告也带冻结，带也带合格/不合格，所以要区分。如果有合格/不合格，说吗是原报告，不去解析
        if ((eachexcelname.find('冻结') >= 0) or (eachexcelname.find('负荷记录') >= 0)) and (eachexcelname.find('合格') < 0):
            excelpathlist.append(expath + '/' + eachexcelname)
            eachexcelnamelist.append(eachexcelname)
        elif eachexcelname.find('稳定性事件记录上报展示主通道') >= 0:
            imeventpathlist.append(expath + '/' + eachexcelname)
            imeacheventnamelist.append(eachexcelname)
        elif eachexcelname.find('稳定性事件记录上报展示临时通道') >= 0:
            teeventpathlist.append(expath + '/' + eachexcelname)
            teeacheventnamelist.append(eachexcelname)
        elif eachexcelname.find('能源稳定性代理成功率') >= 0:
            hplcratepathlist.append(expath + '/' + eachexcelname)
        elif eachexcelname.find('登录') >= 0:
            regpathlist.append(expath + '/' + eachexcelname)
        elif eachexcelname.find('模组及组件版本信息') >= 0 or eachexcelname.find('出厂检查') >= 0 or eachexcelname.find('能源稳定性测试') >= 0 :
            try:
                copyfile(expath + '/' + eachexcelname, newpath + "\\" + eachexcelname)
            except:
                print('copy file faild')

    print('excelpathlist:', excelpathlist)
    print('imeacheventnamelist:',imeacheventnamelist)
    print('teeacheventnamelist:', teeacheventnamelist)
    sheetnamelist = []
    columnlist = []
    # 用来存放每个表格中的不同sheet页的不合格项（每个电表一个sheet页；1号电表、2号电表）;;列表中的2个元素，一个存合格项，一个存不合格项.。数据太多，暂时不启用。也不删掉，以后需要可以用
    allexceljudge = [[],[]]
    #存储最终合格和不合格的报告名称（0：不合格；1：合格）
    allexcelresult = [[],[]]
    #分析冻结类报告
    for eachpathnum in range(len(excelpathlist)):
        daydatalist = []
        columnlistlist = []
        pointlist = []
        # 用来存放每个表格中的不同sheet页的不合格项（每个电表一个sheet页；1号电表、2号电表）
        excelsheetjudge = [[],[]]
        # 不合格测量点记录（每个表格的不合格是sheet页名称记录）
        pointunq = []
        workbook = xlrd.open_workbook(excelpathlist[eachpathnum])
        sheetnamelist = workbook.sheet_names()
        if len(sheetnamelist) < excelsheetnum:
            problemlist[0].append(excelpathlist[eachpathnum].split('/')[1]+ '不合格；漏抄电表，请检查接线或电表')
        else:
            problemlist[1].append(excelpathlist[eachpathnum].split('/')[1] + '抄到所有电表')
        print(' problemlist:', problemlist)
        for sheetnamenum in range(len(sheetnamelist)):
            # try:
            # 用来存放每个sheet页不同数据是否合格（电压是否合格、电流、需量等）
            sheetkeyjudeg = [[],[]]
            columnlistlist = oneormoreexceltolist0([excelpathlist[eachpathnum]])
            print('sheetnamelist[sheetnamenum]:', sheetnamelist[sheetnamenum])
            print('columnlistlist:', columnlistlist)
            DAYDATAROWLIST = columnlistlist[sheetnamenum]
            if '' in DAYDATAROWLIST:
                DAYDATAROWLIST.remove('')
            print('oneormoreexceltolist0([excelpathlist[eachpathnum]]):',
                  oneormoreexceltolist0([excelpathlist[eachpathnum]]))
            daydata = checkfl(excelpathlist[eachpathnum], sheetnamelist[sheetnamenum],DAYDATAROWLIST)
            pointlist,sheetkeyjudeg = stablereport(originalpath, sheetnamelist[sheetnamenum],daydata,pointlist,sheetnamenum,sheetkeyjudeg)
            daydatalist.append(daydata)
            if sheetkeyjudeg[0] != []:
                excelsheetjudge[0].append(sheetnamelist[sheetnamenum]+":"+ str(sheetkeyjudeg[0]))
                pointunq.append([sheetnamelist[sheetnamenum]])
            if sheetkeyjudeg[1] != []:
                excelsheetjudge[1].append(sheetnamelist[sheetnamenum]+":"+ str(sheetkeyjudeg[1]))
        #一个表格中的所有sheet页都合格，并且没有漏抄电表的情况，说明这个表格合格
        if  (excelsheetjudge[0] == []) and (len(sheetnamelist) == excelsheetnum):
            allexcelresult[1].append([eachexcelnamelist[eachpathnum] + ":【合格】"])
        else:
            allexcelresult[0].append([eachexcelnamelist[eachpathnum] + ":【不合格】,不合格测量点如下："])
            allexcelresult[0].extend(pointunq)
        allexceljudge[0].append(excelsheetjudge[0])
        allexceljudge[1].append(excelsheetjudge[1])
        print('allexceljudge:', allexceljudge)
        logging.info("不合格项："+ str(allexceljudge[0]))
        logging.info("合格项：" + str(allexceljudge[1]))
        #最终输出到报告中的结果
        print('allexcelresult:', allexcelresult)
        logging.info("冻结数据处理后最终结果记录：" + str(allexcelresult))

        # 将分析后的报告再重新合成表格
        path1 = newpath + "\\" + eachexcelnamelist[eachpathnum] + '分析后.xlsx'
        analysisexcel(path1, daydatalist, sheetnamelist, columnlistlist)
    #冻结最终报告存储
    #将哪个表格是否有漏点，也进行报告输出
    if  problemlist[0] != []:
        for itemp in problemlist[0]:
            allexcelresult[0].append([itemp])
    if allexcelresult[0] != []:
        write_excel_xlsx_jccomrow(stasavepath, allexcelresult[0], '冻结不合格')
    if allexcelresult[1] != []:
        write_excel_xlsx_jccomrow(stasavepath, allexcelresult[1], '冻结合格')
    # except:pass
        # 稳定性结果处理比较复杂。计量只需要处理冻结数据即可
    if originalpath.find('稳定') >= 0:
        #分析事件报告
        eventdeal(imeventpathlist,imeacheventnamelist,teeventpathlist,teeacheventnamelist,staecandpceventnum,stasavepath)
        #公变时才考察代理成功率
        if originalpath.find('公变') >= 0:
            try:
                hplcreadrate(hplcratepathlist,stasavepath)
            except:
                logging.warning('公变代理成功率处理时出现异常')
        #如果没有登录信息的文件，说明4G登录没有出现异常，不用处理
        if regpathlist != []:
            regreportdeal(regpathlist,stasavepath)
        finalreport(stasavepath,staecandpceventnum,overflag)
    else:
        print('计量用例只需要分析冻结类报告')





if __name__ == '__main__':
    originalpath = 'D:\python_oop_20211028\\report\\000000000001计量自动化测试20220706140031\\'
    #事件记录次数和功率、正向有功是否合格预期值代入
    staecandpceventnum = [0,0,0,1]
    #标记稳定性测试结束时
    overflag = 1
    anlaysispathreport(originalpath,staecandpceventnum,overflag)





    # DAYDATAROWLIST = oneormoreexceltolist0([excelpathlist[0]])[0]
    # daydata = checkfl(excelpathlist[0],sheetnamelist[0][0])
    # print('DAYDATAROWLIST:',DAYDATAROWLIST)







