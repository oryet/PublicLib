# -*- coding: utf-8 -*-
import datetime
import operator
import arrow
import time
import logging
import os
import xlwt
#追加excel使用
import xlutils
from xlutils.copy import copy
from shutil import copyfile
import ctypes
import xlrd
import openpyxl
import re
import win32api, win32con
import xlsxwriter
from xlsxwriter.workbook import Workbook
import sys
import cmath
from Protocol import Exceldeal
from Protocol import dlreaddata
import shutil
import subprocess
import threading
from Protocol import stableanaly
from Protocol import Placefunction
from Protocol import excelchange



snow = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
logging.basicConfig(format='%(asctime)s %(levelname)s:%(message)s',
                    filename=os.getcwd().replace("\Protocol", "") + '\log\\' + snow + 'ooptest.log', level=logging.DEBUG)
#需量和时间里面的时间格式与虚拟表读上来的格式不一致，少了一个字节的年，故要做特殊处理。
demandtimelist=['10100200', '10100201', '10100202', '10100203', '10100204','10100205', '10110200', '10120200', '10130200',
                '10200200', '10200201', '10200202', '10200203', '10200204', '10200205','10210200', '10220200', '10230200',
                '10300200', '10300201', '10300202', '10300203', '10300204',  '10300205','10310200', '10320200', '10330200',
                '10400200', '10400201', '10400202', '10400203', '10400204', '10400205','10410200', '10420200', '10430200', '10500200',
                '10500201', '10500202', '10500203', '10500204', '10500205','10510200', '10520200', '10530200', '10600200', '10600201',
                '10600202', '10600203', '10600204','10600205', '10610200', '10620200', '10630200', '10700200', '10700201', '10700202',
                '10700203', '10700204','10700205', '10710200', '10720200', '10730200', '10800200', '10800201', '10800202', '10800203',
                '10800204','10800205', '10810200', '10820200', '10830200']
ext_compare ='EXTREME_REALDATA'
ext_param = 'EXTREME_M'
#EXTREME，识别到EXTREME时，说明要存储极值
ext_min_data = 'EXTREME_MIN_DATA'
#EXTREME，识别到EXTREME时，说明要存储极值
ext_max_data = 'EXTREME_MAX_DATA'
#EXTREME，识别到EXTREME时，说明要存储极值
ext_min_time = 'EXTREME_MIN_TIME'
#EXTREME，识别到EXTREME时，说明要存储极值
ext_max_time = 'EXTREME_MAX_TIME'
#EXTREME，识别到RECKON_时，说明要存储均值
ave_param = 'RECKON_DATA'
APPIDROWLIST = ['num','appid','detail']
RTUREADLIST = ['data']
VERSIONLIST = ['num','name','version','date']
TASKROWLIST = ['nu', 'rtutype','cut', 'name', 'secure', 'vadd', 'addrtype', 'caddr','op', 'oad_omd','param','save', 'delay', 'expect', 'real', 'result']
HARDWAREROWLIST = ['rated','rtudata','meterdata','error','variation','time']
# 定义常量autoparamset
paramfilesappid = {'Usedappid': 'APPID已用.xlsx',
             'newappid':'APPID本次送检.xlsx'}
#计量相关数据标识对应含义
OADTOMEAN = {'20210200': '数据冻结时间','202A0200': '目标服务器地址', '60400200': '采集启动时标', '60410200': '采集成功时标', '60420200': '采集存储时标','00000200': '组合有功电能','00100200': '正向有功电能总及分',  '00100201': '正向有功电能总', '00100202': '正向有功电能1', '00100203': '正向有功电能2', '00100204': '正向有功电能总3', '00100205': '正向有功电能4', '00200200': '反向有功电能总及分','00200201': '反向有功电能总', '00200202': '反向有功电能1', '00200203': '反向有功电能2', '00200204': '反向有功电能3', '00200205': '反向有功电能4', '00110200': 'A相正向有功电能', '00120200': 'B相正向有功电能', '00130200': 'C相正向有功电能', '00210200': 'A相反向有功电能', '00220200': 'B相反向有功电能', '00230200': 'B相反向有功电能', '00500200': '第1象限无功', '00600200': '第2象限无功', '00700200': '第3象限无功', '00800200': '第4象限无功', '00300200': '组合1无功', '00400200': '组合2无功', '10100200': '正向有功最大需量', '20030200': '电压电流相角',  '20040200': '有功功率', '20040201': '总有功', '20040202': '有功A相', '20040203': '有功B相', '20040204': '有功C相', '00300201': '组合无功1', '00400201': '组合无功2', '00500201': '第一象限无功电能总', '00600201': '第二象限无功电能总', '00700201': '第三象限无功电能总', '00800201': '第四象限无功电能总', '10200200': '反向有功最大需量', '10300200': '组合无功1', '10400200': '组合无功2', '20000200': '电压','20000201': 'A相电压', '20000202': 'B相电压',  '20000203': 'C相电压','20010200': '电流', '20010201': 'A相电流', '20010202': 'B相电流', '20010203': 'C相电流', '20010400': '零线电流', '20010600': '零序电流', '20050200': '无功功率', '200A0200': '功率因数','00000201':'组合有功电能总','00000202':'组合有功电能1','00000203':'组合有功电能2','00000204':'组合有功电能3','00000205':'组合有功电能4','10100201':'正向有功最大需量总','10100202':'正向有功最大需量1','10100203':'正向有功最大需量2','10100204':'正向有功最大需量3','10100205':'正向有功最大需量4','10200201':'反向有功最大需量总','10200202':'反向有功最大需量1','10200203':'反向有功最大需量2','10200204':'反向有功最大需量3','10200205':'反向有功最大需量4','40000200':'当前时间'}

EVENTMENA = {'20220200':'事件记录序号','201E0200':'事件发生时间','20200200':'事件结束时间','20240200':'事件发生源','33000200':'事件上报状态','33050206':'事件发生后2分钟功率','33050207':'控制对象','33050208':'跳闸轮次','33050209':'功控定值','3305020A':'','23012300':'事件发生前总加组有功功率','810C2201':'','33060206':'控制对象','33060207':'跳闸轮次','33060208':'电控定值','33060209':'跳闸发生时总加组电能量','23014900':'','33070206':'控制对象','33070207':'','201E4200':'事件发生时间','F2034201':'遥信状态（1,2）','F2034202':'遥信状态（3，4）','F2034203':'遥信状态（5,6）','F2034204':'遥信状态（7,8）','33180206':'异常标志','33180207':'发生时工况信息','33180208':'发生前工况信息'}

# 获取位状态   istart:Dn, iend:Dn-1, lbytes 字节数
def getbit(data, istart, iend, lbytes):
    if len(data) == 0: return ''
    sbin = bin(int(data, 16)).replace('0b', '').zfill(lbytes*8)
    # print(sbin)
    ilen = len(sbin)
    if istart>ilen or iend > ilen: return ''
    stmp = sbin[lbytes*8-iend:lbytes*8-istart]
    return stmp

# A-XDR格式处理
# '8201AA'->426, 6
def getXDR(s):
    ipos = 0
    ilen = 0
    bfh = getbit(s[:2], 7, 8, 1)
    # print('getXDR', s[:2], bfh)
    if bfh == '0':
        ipos += 2
        ilen = int(s[:2], 16)
    elif bfh == '1':
        ihl = int(getbit(s[:2], 0, 7, 1), 2)
        # print('ih1',ihl)
        ipos = 2 + ihl*2
        ilen = int(s[2:2+ihl*2], 16)
    return ilen, ipos

def inttosHex(iNumber, ibcount):
    stemp = ''
    if isinstance(iNumber, int) == False  or isinstance(ibcount,int) == False:
        return stemp
    stemp = hex(iNumber).upper().replace('0X', '').zfill(ibcount)
    return stemp[-ibcount:]

# 084D2D534353417070 ->M-SCSApp
def hextovisiblestring(s):
    stmp = ''
    ipos = 0
    if len(s) > 2:
        ill, ippp = getXDR(s)
        ipos += ippp
        for i in range(0, ill*2, 2):
            if int(s[ipos:ipos + 2], 16) != 0:
                stmp += chr(int(s[ipos:ipos + 2], 16))
            ipos += 2
    # stmp = str(stmp)
    return stmp, ipos

# 给需要单独生成文件夹的报告单独存储一个文件夹
def reportpath(testexcelname,rtuinilist,reportfloderpath,rtutypelist):
    now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    #为稳定性测试建立独立存储报告的文件夹，并返回存储路径。所有本次测试的稳定性报告，都放入一个文件夹中
    if testexcelname == '能源稳定性测试':
        #如果路径没有，说明第一层测试这个表格里面的测试用例；如果已经有了，就不需要再去生成了。
        if reportfloderpath == '':
            for item in rtutypelist:
                if item.find('专变') >= 0:
                    reportfloderpath = rtuinilist['ADDR'] + '能源稳定性测试专变' + now
                    break
                elif item.find('公变') >= 0:
                    reportfloderpath = rtuinilist['ADDR'] + '能源稳定性测试公变' + now
                    break
                else:pass
            mkdir(os.getcwd() + r"\report\\" + f'{reportfloderpath}\\')
            reportfloderpath = os.getcwd() + r"\report\\" + f'{reportfloderpath}\\'
        else:
            pass
    elif testexcelname == '交采秘钥自动化测试用例集':
        # 如果路径没有，说明第一次测试这个表格里面的测试用例；如果已经有了，就不需要再去生成了。
        if reportfloderpath == '':
            reportfloderpath = rtuinilist['ADDR'] + '计量自动化测试' + now
            mkdir(os.getcwd() + r"\report\\" + f'{reportfloderpath}\\')
            reportfloderpath = os.getcwd() + r"\report\\" + f'{reportfloderpath}\\'
        else:
            pass
    elif testexcelname == '地方抄表测试':
        # 如果路径没有，说明第一次测试这个表格里面的测试用例；如果已经有了，就不需要再去生成了。
        if reportfloderpath == '':
            reportfloderpath = rtuinilist['ADDR'] + '地方抄表测试' + now
            mkdir(os.getcwd() + r"\report\\" + f'{reportfloderpath}\\')
            reportfloderpath = os.getcwd() + r"\report\\" + f'{reportfloderpath}\\'
        else:
            pass
    return reportfloderpath






#二分法查找算法:分标识走这个流程
def dichotomysingle(list1, target):
    length = len(list1)
    index_low = 0
    index_high = length - 1

    while index_low <= index_high:
        index_midle = int((index_low + index_high) / 2)
        guess = list1[index_midle]
        if guess == target:
            return index_midle
        elif guess > target:
            index_high = index_midle - 1
        elif guess < target:
            index_low = index_midle + 1
    return None

#二分法查找算法:总标识走这个流程
def dichotomy(list1, target):
    length = len(list1)
    index_low = 0
    index_high = length - 1

    while index_low <= index_high:
        index_midle = int((index_low + index_high) / 2)
        guess = list1[index_midle][:6]
        if guess == target:
            return index_midle
        elif guess > target:
            index_high = index_midle - 1
        elif guess < target:
            index_low = index_midle + 1
    return None
#16进制转换（float64类型的）为10进制显示,s为整型，可将字符格式的16进制转换为int后，传入即可
def hex2float64(s):
    cp = ctypes.pointer(ctypes.c_longlong(s))
    fp = ctypes.cast(cp, ctypes.POINTER(ctypes.c_double))
    return fp.contents.value

#16进制转换（float32类型的）为10进制显示,s为整型，可将字符格式的16进制转换为int后，传入即可
def hex2float32(s):
    i = int(s, 16)  # convert from hex to a Python int
    cp = ctypes.pointer(ctypes.c_int(i))  # make this into a c integer
    fp = ctypes.cast(cp, ctypes.POINTER(ctypes.c_float))  # cast the int pointer to a float pointer
    return round(fp.contents.value,5)

# bytes->hex
def hexShow(argv):
    result = ''
    hLen = len(argv)
    for i in range(0, hLen, 1):
        # print(argv[i])
        # hvol = ord(argv[i])
        hhex = '%02x' % argv[i]
        result += hhex.upper()
    # print('hexShow:' + result)
    now = datetime.datetime.now()
    stt = '[' + now.strftime('%Y-%m-%d-%H:%M:%S:%f')[:23] + '] '
    result = stt + result
    return result

# bytes->hex  不加时间
def hexoriShow(argv):
    result = ''
    hLen = len(argv)
    for i in range(0, hLen, 1):
        hhex = '%02x' % argv[i]
        result += hhex.upper()
    return result


# 20191126 00 01 42 0000->   datetime
def bcdtodatetime(dt):
    dt.zfill(14)
    if len(dt) == 14:
        dt += '0000'
    tt = datetime.datetime(int(dt[0:4]), int(dt[4:6]), int(dt[6:8]), int(dt[8:10]),
                           int(dt[10:12]), int(dt[12:14]), int(dt[14:])*1000%1000000)
    return tt

# hex->bytes
def hexascii(shex):
    result = []
    schar = shex.replace(' ', '')
    result = bytes.fromhex(schar)
    return result

#setparamlist->[{'oad/omd':'','value':''},.....]
#Idatarow：当前行；rlist:readparamlist
def listappeddict(Idatarow,rlist):
    paramappendoaddict={}
    if (Idatarow['save'].find("SUM") >= 0) or  (Idatarow['save'].find("APP") >= 0 )or (Idatarow['save'].find("rate") >= 0) or (Idatarow['save'].find("V_METER") >= 0) or (Idatarow['save'].find("VOLPASSRATE") >= 0) or (Idatarow['save'].find('REGION_PARAM') >= 0) or (Idatarow['save'].find('EXTREME') >= 0) or (Idatarow['save'].find('RECKON') >= 0) or (Idatarow['save'].find('_RIGHTVALUE_') >= 0):
        paramappendoaddict['oad/omd'] = Idatarow['save']
    else:
        paramappendoaddict['oad/omd']=Idatarow['oad_omd']
    paramappendoaddict['value']=rlist
    print(paramappendoaddict)
    return paramappendoaddict

def getcurrenttime():
    sytime = time.localtime(time.time())
    sytime = time.strftime('%Y%m%d%H%M%S', sytime)
    return sytime

#用正则表达式来获取时分秒
def getcurrenttimenew(param):
    paramneed =   re.findall(r'\d+', param)[0]
    sytime = time.localtime(time.time())
    sytime = time.strftime('%Y%m%d', sytime) + paramneed
    return sytime


#对时至当前日的后一天
def getcurrenttimeafter(param):
    paramneed =   re.findall(r'\d+', param)[0]
    sytime = time.localtime(time.time())
    dayget = str(int(time.strftime('%Y%m%d', sytime)[6:8]) + 1).zfill(2)
    sytime = time.strftime('%Y%m', sytime) + dayget + paramneed
    return sytime
#获取当前月的最后一天
def last_day_of_month(any_day):
    """
    获取获得一个月中的最后一天
    :param any_day: 任意日期
    :return: string
    """
    print(any_day.replace(day=28))
    next_month = any_day.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
    print('next_month:',next_month)
    print(datetime.timedelta(days=next_month.day))
    return str(next_month - datetime.timedelta(days=next_month.day)).replace('-','')

#获取当前月末（最后一天）
def getcurmonlastday(param):
    paramneed =   re.findall(r'\d+', param)[0]
    sytime = time.localtime(time.time())
    ymd = time.strftime('%Y%m%d', sytime)
    # 注意: 年月日，这些变量必须是数字，否则报错！
    year = int(ymd[:4])  # 年
    month = int(ymd[4:6])  # 月
    day = int(ymd[6:8])  # 日
    ymdget = last_day_of_month(datetime.date(year, month, day))
    sytime = ymdget + paramneed
    return sytime

#对时至当前年末（最后一天）
def getcuryearlastday(param):
    paramneed =   re.findall(r'\d+', param)[0]
    sytime = time.localtime(time.time())
    ymd = time.strftime('%Y%m%d', sytime)
    # 注意: 年月日，这些变量必须是数字，否则报错！
    year = ymd[:4]  # 年
    ymdget = year + '1231'
    sytime = ymdget + paramneed
    return sytime

#15分钟曲线召测的召测区间处理
def freezeminread(recordparam,data6012record):
    paramresult = ''
    fifmin = datetime.timedelta(minutes=15)
    readparam = recordparam.split(':')[1]
    #能源稳定性测试时，（曲线和实时数据）第一次召测时，记录下首次召测时间
    if recordparam.split(':')[0].find('TIME_FIRST_GET') >=0:
        testtime = datetime.datetime.now().strftime('%Y%m%d%H%M') + '00'
        #15分钟的整数倍
        savedatatime = str(int(str(int(testtime[-4:-2]) // 15 * 15))).zfill(2)
        resulttime = testtime[:-4] + savedatatime + '00'
        # 前
        readdatatime = str(bcdtodatetime(resulttime) - fifmin).replace("-", '').replace(':', '').replace(' ', '')
        #后
        endopenttime = str(bcdtodatetime(resulttime)).replace("-", '').replace(':', '').replace(' ', '')
        # #后
        # endopenttime = str(bcdtodatetime(resulttime) + fifmin).replace("-", '').replace(':', '').replace(' ', '') + '00'
        data6012record[0] = readdatatime
        data6012record[1] = endopenttime
    #能源稳定性测试时，（曲线和实时数据）下一次召测时，时间需要在原有的基础上累加15分钟。
    elif recordparam.split(':')[0].find('TIME_ADD') >=0 :
        readdatatime = str(bcdtodatetime(data6012record[0]) + fifmin).replace("-", '').replace(':', '').replace(' ', '')
        endopenttime = str(bcdtodatetime(data6012record[1]) + fifmin).replace("-", '').replace(':', '').replace(' ', '')
        data6012record[0] = readdatatime
        data6012record[1] = endopenttime
    # 能源稳定性测试时，（曲线和实时数据）一次召测6行，只有首行的时候改变召测时间。然后再开始下一次召测
    else:
        readdatatime = data6012record[0]
        endopenttime = data6012record[1]
    paramresult = strtolist(readparam)
    paramresult[0][1][0] = readdatatime
    paramresult[0][1][1] = endopenttime
    print(f'param:{paramresult}')
    return paramresult,data6012record

#1h曲线召测的召测区间处理
def freezehourread(recordparam,data6012record):
    paramresult = ''
    hour_one = datetime.timedelta(hours=1)
    readparam = recordparam.split(':')[1]
    #能源稳定性测试时，（曲线和实时数据）第一次召测时，记录下首次召测时间
    if recordparam.split(':')[0].find('TIME_FIRST_GET') >=0:
        testtime = datetime.datetime.now().strftime('%Y%m%d%H%M') + '00'
        #15分钟的整数倍
        savedatatime = str(int(str(int(testtime[-6:-4])))).zfill(2)
        resulttime = testtime[:-6] + savedatatime + '0000'
        #前
        readdatatime = str(bcdtodatetime(resulttime)).replace("-", '').replace(':', '').replace(' ', '') + '00'
        #后
        endopenttime = str(bcdtodatetime(resulttime) + hour_one).replace("-", '').replace(':', '').replace(' ', '') + '00'
        data6012record[2] = readdatatime
        data6012record[3] = endopenttime
    #能源稳定性测试时，（曲线和实时数据）下一次召测时，时间需要在原有的基础上累加15分钟。
    elif recordparam.split(':')[0].find('TIME_ADD') >=0 :
        readdatatime = str(bcdtodatetime(data6012record[2]) + hour_one).replace("-", '').replace(':', '').replace(' ', '')
        endopenttime = str(bcdtodatetime(data6012record[3]) + hour_one).replace("-", '').replace(':', '').replace(' ', '')
        data6012record[2] = readdatatime
        data6012record[3] = endopenttime
    # 能源稳定性测试时，（曲线和实时数据）一次召测6行，只有首行的时候改变召测时间。然后再开始下一次召测
    else:
        readdatatime = data6012record[2]
        endopenttime = data6012record[3]
    paramresult = strtolist(readparam)
    paramresult[0][1][0] = readdatatime
    paramresult[0][1][1] = endopenttime
    print(f'param:{paramresult}')
    return paramresult,data6012record

#终端数据筛选:曲线召测区间处理（定期读取脉冲和总加组功率曲线）
def rtuselect(readparam):
    paramresult = ''
    fifmin = datetime.timedelta(minutes=5)
    testtime = datetime.datetime.now().strftime('%Y%m%d%H%M') + '00'
    #15分钟的整数倍
    savedatatime = str(int(str(int(testtime[-4:-2]) // 5 * 5))).zfill(2)
    resulttime = testtime[:-4] + savedatatime + '00'
    #前
    readdatatime = str(bcdtodatetime(resulttime)).replace("-", '').replace(':', '').replace(' ', '')
    #后
    endopenttime = str(bcdtodatetime(resulttime) + fifmin).replace("-", '').replace(':', '').replace(' ', '')
    paramresult = strtolist(readparam)
    paramresult[0][2] = readdatatime
    paramresult[0][3]= endopenttime
    print(f'param:{paramresult}')
    return paramresult


#datarelat：数据关系表,startlist列表开始追加位置,stoplist列表结束追加位置
def freezetype(datarelat,startlist,stoplist):
    findoad = 0
    rtuoadindex=0
    Rtuoadlist=[]
    for jtem in range(1, len(datarelat), 1):
        if (datarelat[jtem]['rtumajor'].decode()) == startlist:
            findoad = 1
            rtuoadindex = jtem + 1
        if findoad == 1:
            if datarelat[jtem]['rtumajor'].decode() == stoplist:
                break
            else:
                pass
            if len(datarelat[jtem]['rtu']) > 0:
                if len(datarelat[jtem]['rtu']) > 0:
                    Rtuoadlist.append(datarelat[jtem]['rtu'].decode())
                else:
                    pass
    return Rtuoadlist,rtuoadindex


#根据读取参数，来判断要在哪些oad中查找读取的标识
#datarelation, needparam:数据关系表，需要读取的参数
def metertype(datarelation, needparam,rtuoadindex):
    #Rtuoadlist 获得的查找范围
    Rtuoadlist=[]

    if needparam.find('FREEZE') <0:
        liststart = '当前电能'
        liststop = '老曲线'
        Rtuoadlist, rtuoadindex = freezetype(datarelation, liststart, liststop)

    elif needparam.find('FREEZEDAY_1') > 0:
        liststart='上1次日冻结'
        liststop='上2次日冻结'
        Rtuoadlist,rtuoadindex=freezetype(datarelation,liststart,liststop)

    elif needparam.find('FREEZEDAY_2') > 0:
        liststart = '上2次日冻结'
        liststop = '上3次日冻结'
        Rtuoadlist, rtuoadindex = freezetype(datarelation, liststart, liststop)

    elif needparam.find('FREEZEDAY_3') > 0:
        liststart = '上3次日冻结'
        liststop = '上1结算日'
        Rtuoadlist, rtuoadindex = freezetype(datarelation, liststart, liststop)

    elif needparam.find('FREEZEMON_1') > 0 :
        liststart = '上1结算日'
        liststop = '上2结算日'
        Rtuoadlist, rtuoadindex = freezetype(datarelation, liststart, liststop)

    elif needparam.find('FREEZEMON_2') > 0 :
        liststart = '上2结算日'
        liststop = '上3结算日'
        Rtuoadlist, rtuoadindex = freezetype(datarelation, liststart, liststop)
    elif needparam.find('FREEZEMON_3') > 0 :
        liststart = '上3结算日'
        liststop = '结束'
        Rtuoadlist, rtuoadindex = freezetype(datarelation, liststart, liststop)
    else:print('增加冻结标识查找列表流程')


    return Rtuoadlist,rtuoadindex

#处理从数据关系表中获取到的数据：有特殊的，比如，总-多个分（1对5），或者需量和时间的逗号（这个逗号要保留）等处理
#exdata：表格中需要处理的数据
def specialdata(exdata):
    c = ''
    i = 0
    ml=[]
    oadtimes=0
    moerdatalist = exdata.split(',')
    if len(moerdatalist) > 2 and len(moerdatalist) < 6:
        ml.extend(moerdatalist)
        oadtimes=len(ml)
    #数据项大于5个，说明是需量和时间，它们之间的逗号要保留
    elif len(moerdatalist) > 5:
        for item in moerdatalist:
            i += 1
            c += item + ","
            if i == 2:
                i = 0
                ml.append(c[:-1])
                c = ''
                continue
        oadtimes = len(ml)
    else:
        ml.append(exdata)
        oadtimes = len(ml)
    return ml,oadtimes
# 带[]字符串转换成list '[c,d,e]'->[c,d,e]
def strtolist(slist):
    lrtn = eval("%s" % slist)
    return lrtn


#配置交采档案时，有的地区要求特殊，档案地址会配置失效，具体以终端中获取到的为准，此时就要考虑不比较交采地址。（进行了剔除） 例如，比较时，[[1, ['05000000000000', 6, 3, 'F2080201', '000000', 4, 0, 3, 220, 1], ['0000', '00', 10, 100], []]；；；；；'05000000000000'会被剔除，无论地址是什么，都判合格。
def changejcaddr(item,item1,realrow):
    realdatadeladdr = ''
    setdatadeladdr = ''
    realrow['real'] = str(item1['value'])
    realdatadeladdr = item1['value']
    del realdatadeladdr[0][1][0]
    setdatadeladdr = item['value']
    del setdatadeladdr[0][1][0]
    if realdatadeladdr == setdatadeladdr:
        return True
    else:
        return False

#6000参数比较流程setP:设置参数,getP：读取到的参数,realrow：当前行
def compareparam(setP,getP,realrow):
    #comfind:标记是否找到需要比较的标识
    comfind = 0
    for item in  setP:
        #找到对应标识，再进行比较。
        if item['oad/omd'][:4] ==realrow['oad_omd'][:4]:
            for item1 in getP:
                if item['oad/omd'][:4] == item1['oad/omd'][:4]:
                    # 单独直接设置的参数，item['value']进行特殊处理
                    if realrow['param'] == 'DIRECT_SET':
                        if operator.eq(list(strtolist(item['value'][0])), item1['value']):
                            realrow['real'] = str(item1['value'])
                            realrow['result'] = u'合格'
                        else:
                            realrow['real'] = str(item1['value'])
                            realrow['result'] = u'不合格'
                    else:
                        #交采档案地址不比，进行特殊处理 20211104
                        if realrow['save'].find('CHANGE_AS') >= 0:
                            if changejcaddr(item,item1,realrow):
                                realrow['result'] = u'合格'
                            else:
                                realrow['result'] = u'不合格'
                        else:
                            if operator.eq(item['value'], item1['value']):
                                realrow['real'] = str(item1['value'])
                                realrow['result'] = u'合格'
                            else:
                                realrow['real'] = str(item1['value'])
                                realrow['result'] = u'不合格'

                    comfind = 1
                    #找到了，就跳出第二个循环
                    break
                #找到了就跳出第一个循环（每次只比一个标识）
                else:
                    pass
                if comfind == 1:
                  break
                else:
                    pass

        else:pass
        #如果没有找到比较项
    if comfind ==0:
        print("比较标识有误")
        realrow['real'] = '比较标识有误'
        realrow['result'] = u'不合格'
    else:pass





    return comfind




##定义一个如果标识相同，就覆盖原来值，标识不同，则继续追加值的函数。只要相同流程，需要覆盖原来值的，都可以调用。
# ldataoad,valuelist,newvalue:当前行，需要刷新或追加值的列表，新值
def covervalue(ldataoad, valuelist, newvalue):
    findsameoad = 0
    addfindflag = 0
    # 如果在列表中找到了要存储的变量名，说明就是ADD了
    for eachvaluelist in valuelist:
        if (ldataoad['save'].find("ADD") >= 0) and (eachvaluelist['oad/omd'].find(ldataoad['save']) >= 0):
            addfindflag = 1
        else:
            pass
    if len(valuelist) >= 1:
        for valuelistnum in range(len(valuelist)):
            # 存放若干个对象属性的值时，用来区分的是存储列的内容
            if (ldataoad['save'].find("SUM") >= 0) or (ldataoad['save'].find("APP") >= 0) or (ldataoad['save'].find(
                    "rate") >= 0 or ldataoad['save'].find("V_METER") >= 0 or ldataoad['save'].find(
                    "VOLPASSRATE") >= 0) or (ldataoad['save'].find("EXTREME") >= 0) or (ldataoad['save'].find(
                    "RECKON") >= 0) or (ldataoad['save'].find("_RIGHTVALUE_") >= 0) :
                if valuelist[valuelistnum]['oad/omd'] == ldataoad['save'] or (ldataoad['save'].find("ADD") >= 0 and (
                        (addfindflag == 0 and valuelist[valuelistnum]['oad/omd'] == ldataoad['save'][:-4])  or valuelist[valuelistnum][
                    'oad/omd'] == ldataoad['save'])):
                    # 处理电压合格率的特殊情况(有 ADD 表示新值要和老值累加，并将'oad/omd'刷新为带ADD的'oad/omd'),只追加与MON_1_ADD对应的值。
                    if ldataoad['save'].find("ADD") >= 0:
                        # 有ADD的流程需要特殊处理：如果在已经存在的列表中找到X_ADD，说明，不要再去累加 X 的的值，同样，也不要保留上次ADD的值，直接覆盖掉即可。我们的流程，只保留 X 和 X_ADD.
                        if (addfindflag == 0 and valuelist[valuelistnum]['oad/omd'] == ldataoad['save'][:-4]) or \
                                valuelist[valuelistnum]['oad/omd'] == ldataoad['save']:
                            for newvalueitem in range(1, len(valuelist[valuelistnum]['value']), 1):
                                newvalue[newvalueitem] = valuelist[valuelistnum]['value'][newvalueitem] + newvalue[
                                    newvalueitem]
                            newvalue[1] = 10000 - newvalue[2]
                        else:
                            pass
                    else:
                        pass
                    # 如果有ADD,除了要构建 ADD的新值，原本的 day_1的值也要保留，比较的时候会取用。
                    if ldataoad['save'].find("ADD") >= 0 and addfindflag == 0:
                        valuelist.append(listappeddict(ldataoad, newvalue))
                        findsameoad = 1
                    else:
                        valuelist[valuelistnum] = listappeddict(ldataoad, newvalue)
                        findsameoad = 1
                else:
                    pass
            else:
                if valuelist[valuelistnum]['oad/omd'] == ldataoad['oad_omd']:
                    valuelist[valuelistnum] = listappeddict(ldataoad, newvalue)
                    findsameoad = 1
                else:
                    pass
        if findsameoad == 0:
            valuelist.append(listappeddict(ldataoad, newvalue))
    else:
        valuelist.append(listappeddict(ldataoad, newvalue))
    return valuelist
#当前行，raterow，c,d为数据格式处理时的临时变量。
def getcpu_mem_disk_rate(raterow):
    ratelist = []
    c = raterow['real'].split("%")
    for item in c:
        d = item.split("-usage:")
        if len(d) > 1:
            ratelist.append(d[1])
    return ratelist



#cpu\memory\storage占用率存储
def usagerate(realrow,ratelist,rate):
    ratelist = covervalue(realrow,ratelist,rate)
    return ratelist


#taskratereal,当前行cpu_rate,memory_rate,storage_rate：三种定义好的列表，存储三种占有率
def allusagerate(taskratereal,cpu_rate,memory_rate,storage_rate):
    try:
        ratelist = []
        ratesave=[]
        rowsaveold=taskratereal['save']
        #现在save中含有3个变量名，存储在列表中，然后在获取各个变量值的时候分别取用
        ratesave=rowsaveold.split(",")
        ratelist=getcpu_mem_disk_rate(taskratereal)
        taskratereal['save'] = ratesave[0]
        cpu_rate = usagerate(taskratereal, cpu_rate,int(ratelist[0]))
        taskratereal['save'] = ratesave[1]
        memory_rate = usagerate(taskratereal, memory_rate,int(ratelist[1]))
        taskratereal['save'] = ratesave[2]
        storage_rate = usagerate(taskratereal, storage_rate,int(ratelist[2]))
        #获取到需要的值以后，将save中的值恢复，供下次使用。直接改写，会导致下次执行到该函数时，因save值异常而报错。
        taskratereal['save']= rowsaveold
    except IndexError:
        print('IndexError:list index out of range')

    return cpu_rate,memory_rate,storage_rate

 #获取最大值
def max_value(rate_dictinlist):
    print('rate_dictinlist:',rate_dictinlist)
    comparelist=[]
    for item in rate_dictinlist:
        comparelist.append(item['value'])
    maxvalue=max(comparelist)
    return maxvalue

def expect_meter_v(vmsetrow,v_meter):
    v_meter =covervalue(vmsetrow,v_meter,float(vmsetrow['param']))
    return v_meter

#vraterow:电压合格率设置行
def v_passrateparamget(vraterow,v_passrate_param):
    v_settletime = []
    # 列表中最终存储float类型的数据供处理使用。
    vpassrateparamfloat = v_passrate_param
    if vraterow['save'].find('SETTLEDAY_HOUR') >= 0:
        v_settletime = strtolist(vraterow['param'])
    else:
        vpassrateparamfloat = []
        vpassrateparam=vraterow['param'].split(",")
        for item in vpassrateparam:
            vpassrateparamfloat.append(float(item))
    print('vpassrateparamfloat,v_settletime:',vpassrateparamfloat,v_settletime)
    return vpassrateparamfloat,v_settletime










#定义表读取后，处理为多条记录格式函数。mllist为表读取最初格式。
def vmmeterlist(mllist):
    #存储临时变量
    b=[]
    ml=[]
    for item in mllist:
        if item =='!':
            ml.append(b)
            b=[]
            continue
        b.append(item)
    ml.append(b)
    return ml
#定义一个60120300读取记录型对象的属性读取数据的处理函数，处理为一个可以与表读取进行比较的列表。
def read6012list(readlist):
    if readlist != []:
        readlist = readlist.replace("[", '').replace(']', '').replace('|', ';').split(";")
    else:
        pass
    for readlistnum in range(len(readlist)):
        if readlist[readlistnum].find('"') >= 0 or readlist[readlistnum].find("'") >= 0:
            readlist[readlistnum] = readlist[readlistnum].replace('"', '').replace("'", '')
        else:
            pass
    return readlist

#构建arrow对象：strtime：'20200101000000'
def bulidarrowobject(strtime):
    arrowobject=arrow.Arrow(int(strtime[:4]), int(strtime[4:6]),
                int(strtime[6:8]), int(strtime[8:10]),
                int(strtime[10:12]), int(strtime[12:14]))
    return arrowobject

#months:几个月;oldstrtime:需要构建为arrow对象的时间字符串
def lastmonths(months,oldstrtime,plantimelist):
    monthlist=[]
    nowmonthtime = bulidarrowobject(oldstrtime)
    #现在是依据方法7来召测，根据召测时标来确定  存储时间 6042，存在被召测月的最后一天23点59分
    #例：召测时间：20191201000000-20200201000000----那么，存储时间就是  201912235900 和 20200131235900
    #如果是 '235900'：说明是 存储时标7，要处理在  235900，流程特殊。
    if plantimelist[0][-6:] == '235900':
        for j in range(0, (months), 1):
            lastmonth = nowmonthtime.shift(months=-j).format("YYYYMM")
            print('lastmonth:', lastmonth)
            courrentmonth = bcdtodatetime(lastmonth + '01235900')
            oneday = datetime.timedelta(days=1)
            savemonth = str(courrentmonth - oneday).replace("-",'').replace(':','').replace(' ','')
            monthlist.append(savemonth)
    else:
        for j in range(1, (months + 1), 1):
            lastmonth = nowmonthtime.shift(months=-j).format("YYYYMM")
            print('lastmonth:',lastmonth)
            monthlist.append(lastmonth + '01000000')
    monthlist.reverse()
    return monthlist


# bytes->hex
def hexShowNoTime(argv):
    result = ''
    hLen = len(argv)
    for i in range(0, hLen, 1):
        # print(argv[i])
        # hvol = ord(argv[i])
        hhex = '%02x' % argv[i]
        result += hhex.upper()
    # print('hexShow:' + result)
    # now = datetime.datetime.now()
    # stt = '[' + now.strftime('%Y-%m-%d-%H:%M:%S:%f')[:23] + '] '
    # result = stt + result
    return result

# 按字节倒置
def reversebit(data):
    string = ''
    for i in range(len(data) - 1, -1, -2):
        string += data[i - 1]
        string += data[i]
    return string

#处理补抄的tsa和任务存储时间。plantimelist,listtsa,dparam:原任务存储时间列表、地址列表和60120300传入的参数dparam
def freezetsaandplansave( plantimelist,listtsa,dparam):
    if dparam[0][0][0] == '选择方法6' or dparam[0][0][0] == '选择方法7' or dparam[0][0][0] == '选择方法8':
        freezetimelist = []
            # 任务存储时间，直接往前推machdays天或者monthdiffer月。
        if dparam[0][1][4] == "50060200":
            # 月冻结时，计算差几个月
            timestartregmon = int(dparam[0][0][1][0][4:6])
            year1=int(dparam[0][0][1][0][0:4])
            timeendregmon = int(dparam[0][0][1][1][4:6])
            year2 = int(dparam[0][0][1][1][0:4])
            monthdiffer = timeendregmon - timestartregmon
            # 得到负数说明是跨年了，要计算出正确的月的差值。
            if monthdiffer <= 0 and year1 < year2:
                monthdiffer = monthdiffer + 12
            else:
                monthdiffer = monthdiffer
            print('monthdiffer:',monthdiffer)
            if plantimelist == ['日冻结数据冻结时标']:
                pass
            #构建arrow对象
            elif monthdiffer > 1:
                plantimelist = lastmonths(monthdiffer,plantimelist[0],plantimelist)
            else:pass
            # tsa直接复制几次
            for months in range(monthdiffer - 1):
                listtsa.append(listtsa[0])
        elif dparam[0][1][4] == "50040200" or dparam[0][1][4] == "50050200" :
            # 日冻结和结算日冻结时，计算差几天
            timestartreg = (dparam[0][0][1][0])
            timeendreg = (dparam[0][0][1][1])
            timeitem = bcdtodatetime(timeendreg) - bcdtodatetime(timestartreg)
            machdays = timeitem.days
            if plantimelist == ['日冻结数据冻结时标']:
                pass
            elif machdays>1:
                for i in range(1, (machdays + 1), 1):
                    dt = bcdtodatetime(plantimelist[0])
                    delta = datetime.timedelta(days=i)
                    freezetimelist.append((str(dt - delta).replace('-', '').replace(" ", '').replace(":", '')))
                freezetimelist.reverse()
                plantimelist = freezetimelist
                        # tsa直接复制几
            else:pass
            for days in range(machdays - 1):
                listtsa.append(listtsa[0])
        else:pass
    else:
       pass
    print('listtsa,plantimelist:',listtsa,plantimelist)
    return plantimelist,listtsa

#预期值为NULL的数据正确性判断流程。data6012为报文中的data。
def nulldata(data6012):
    data = True
    b = read6012list(data6012)
    for item in range(len(b)):
        if item == 0:
            if ((len(b[item])/2) - 2 )== int(b[item][:2]):

                tsastyle = True
            else:
                data = False
        elif item == 3:
            if len(b[item]) == 14:
                timestyle = True
            else:
                data = False
        else:
            if b[item] == 'NULL':
                pass
            else:
                data = False
    return data


#读取若干个对象属性后，对数据进行格式处理，方便  通过终端获取的值和60120300报文中的值进行比较时使用。
##需量和时间格式处理需要完善！！！0716
def Rtu698datalistdeal(data,exceloadlist):
    #如果是需量和时间，格式要特殊处理，所以要标记一下。
    demandtimeflag=0
    exceloadlist = exceloadlist.split(',')
    if  exceloadlist[0] in demandtimelist:
        demandtimeflag = 1
    else:pass

    if len(data)>0:
        if demandtimeflag == 1:
            data = data.replace("]][[", '],[').strip("[[").strip("]]").split("],[")
        else:
            data = data.strip('[]').replace("][", ',').replace('[',",").replace(']',",").replace(',,,',",").replace(',,',",").split(",")
        # 数据为''，直接忽略的流程，后面需要调试
        if '' in data:
            data.remove('')
        else:
            pass
        for readlistnum in range(len(data)):
            if data[readlistnum].find('"') >= 0 or data[readlistnum].find("'") >= 0:
                data[readlistnum] = data[readlistnum].replace('"', '').replace("'", '')
            else:
                pass

    else:
        pass
    return data


#通过当前行的oad_omd，获得真正的读取若干个对象属性中的每一个数据对应的数据标识。:借鉴表读取流程
#readoad,mdatalist：当前行的oad_omd,数据关系表
def Rtu698oadlistdeal(readoad,mlist):
    oad698list=[]
    oadindex = 0
    if isinstance(readoad,list):
        pass
    else:
        readoad = readoad.split(',')
    Rtudtlist, oadfindindex = metertype(mlist, '', oadindex)
    try:
        for item in readoad:
            # print('item:',type(item),item)
            itemindex = dichotomy(Rtudtlist, item[:6])
            if itemindex == None:
                #如果是2021，直接忽略。交采的2021不在终端读取流程中获取，而是获取6042的值，在comparetime流程中进行处理。
                if item[:6] == "202102":
                    oad698list.append("20210200")
                    pass
                else:
                    print('请检查读取标识是否在数据关系表中！', item[:6])
                continue
            itemindex += oadfindindex
            # 如果查找不到标识，就打印提示信息，并不再向下执行。
            # print('itemindex:',itemindex)
            # print(mlist[itemindex]['rtu'][:6].decode())
            # 如果找到的标识不是第一个，需要处理后取到第一个
            if mlist[itemindex - 1]['rtu'][:6].decode() == item[:6]:
                # 第一个while循环找到序号最小的。
                # print('mlist[itemindex][:6].decode():',mlist[itemindex]['rtu'][:6].decode(),item[:6])
                while mlist[itemindex]['rtu'][:6].decode() == item[:6]:
                    itemindex -= 1
                startindex = itemindex + 1
                # 第二个while循环从第一个值开始往下累加。
                while mlist[startindex]['rtu'][:6].decode() == item[:6]:
                    # 电表没有对应值，表读取时，忽略。但是如果表的位置对应的NULL，说明后面的值无论是null还是有其它值，表读取都要读到。
                    oad698list.append(mlist[startindex]['rtu'].decode())
                    startindex += 1
            # 如果找到的标识是第一个，直接处理。
            else:
                while mlist[itemindex]['rtu'][:6].decode() == item[:6]:
                    oad698list.append(mlist[itemindex]['rtu'].decode())
                    itemindex += 1
    except TypeError:
        print("TypeError: 标识不在数据关系表或标识排序错误")


    return oad698list

#dt_m_list,Rtu698oadlist,Rtu698datalist 60120300读取的标识和值；mrtulist:数据关系表;paramrow参数所在行
# 读取若干个对象属性获取的标识和值（相当于原来的虚拟表-在程序中以列表的形式存储）
def getrtudata(dt_m_list1,Rtu698oadlist1,Rtu698datalist1,mrtulist,paramrow):
    dt_m0_listdeal=[]
    dt_m0_list_datadeal =[]
    Rtu698datalistsum=[]
    tsa = '电表地址'
    celetstarttime = '采集启动时标'
    celetsuccesstime = '采集成功时标'
    celetsavetime = '采集存储时标'
    dt_m0_listdeal.extend(['202A0200', '60400200', '60410200', '60420200'])
    print('Rtu698oadlist1:',Rtu698oadlist1)
    if len(dt_m_list1)>0  and len(Rtu698oadlist1)>0 and len(Rtu698datalist1)>0:
        dt_m_list1.remove(dt_m_list1[0])
        dt_m_list1=(Rtu698oadlistdeal(dt_m_list1[4:], mrtulist))
        print('dt_m_list1:',dt_m_list1)
        #找到终端读取对应的SUM（SUM1或SUM2或……）
        paramsumlist=paramrow['param'].split(",")
        for aramsumlistitem in paramsumlist:
            for Rtu698datalist1item in Rtu698datalist1:
                if Rtu698datalist1item['oad/omd'] == aramsumlistitem:
                    Rtu698datalistsum.append(Rtu698datalist1item['value'])
                else:
                    pass
        logging.info('Rtu698datalistsum'+str(Rtu698datalistsum))
        print(Rtu698datalistsum)
        #标识只需要获取一次（2次数据，标识是一样的），故单独处理。
        for item in dt_m_list1:
            dt_m0_listdeal.append(item)
        #获取数据
        for Rtu698datalistsumnum in range(len(Rtu698datalistsum)):
            dt_m0_list_datadeal.extend([tsa, celetstarttime, celetsuccesstime, celetsavetime])
            for item in dt_m_list1:
                if item == '20210200':
                    pass
                elif item in Rtu698oadlist1:
                    dt_m0_list_datadeal.append(Rtu698datalistsum[Rtu698datalistsumnum][Rtu698oadlist1.index(item)])
                else:
                    print("请检查60120300召测标识是否包含在终端获取的总标识中", item)
            if Rtu698datalistsumnum < (len(Rtu698datalistsum) -1):
                dt_m0_list_datadeal.append('!')
        dt_m0_list_datadeal = vmmeterlist(dt_m0_list_datadeal)


    else:
        print("请检查用例执行顺序是否正确")
    return dt_m0_listdeal,dt_m0_list_datadeal


#定义一个处理60120300召测报文中如果有空值时的处理函数，如果是总的标识，则一个空值其实应该对应5个空值，
#如果是分相标识，则一个空值对应1个空值，不需要做处理。
#data_6012_read:格式处理后的，60120300读取到的值，dt_all698m_list：通过数据关系表获取到的标识列表。
def empty6012data(data_6012_read,dt_all698m_list):
    # print('empty  data_6012_read:',data_6012_read)
    totalflag=0
    try:
        for i in range(len(dt_all698m_list)-1):
            if data_6012_read[i] == '' :
                del data_6012_read[i]
                # print('i:',i)
                if dt_all698m_list[i][:6] == dt_all698m_list[i+1][:6] :
                    print("empty：i:",i)
                    # print('dt_all698m_list[i][:6]:', dt_all698m_list[i][:6])
                    totalflag=1
                    # print('dt_all698m_list[i+1]：',dt_all698m_list[i])
                #如果是总标识，就填充5个空值（这个if条件会成立多次）
                    for j in range(5):
                      if dt_all698m_list[i] in demandtimelist:
                        data_6012_read.insert(i+j, ',')
                        # print('data_6012_read每次改变后：',data_6012_read)
                      else:
                        data_6012_read.insert(i+j, '')
                else:
                  totalflag=0
               #如果是分相标识，不需要做处理
                if dt_all698m_list[i+1][-2:] == '05' and totalflag==0 :
                    if dt_all698m_list[i] in demandtimelist:
                        print('dt_all698m_list[i]：', dt_all698m_list[i])
                        data_6012_read.insert(i, ',')
                    else:
                        pass
    except IndexError:
        print('data_6012_read[i] == '':list index out of range')
    return data_6012_read

#timestr1,timestr2:两个要比较的时间字符串形式；返回值为两者相差的秒数。（只能比到差30天23小时59分的时间差，无法比月）
def timecompare(timestr1,timestr2):
  if bcdtodatetime(timestr1) > bcdtodatetime(timestr2):
    timediffer = bcdtodatetime(timestr1) - bcdtodatetime(timestr2)
  else:
    timediffer = bcdtodatetime(timestr2) - bcdtodatetime(timestr1)
  print('timediffer:' + str(timediffer))
  return timediffer
#返回时间差的具体秒数
def differseconds(timestyle):
    timediffers= timestyle.days * 86400 + timestyle.seconds
    return timediffers

#
# def  builtnewcompare(dealrow,timeorvaluelist,average_time):
#     app_start_average_time=averagevalue(timeorvaluelist)
#     print('app_start_average_time',app_start_average_time)
#     average_time=covervalue(dealrow, average_time, app_start_average_time)
#     dealrow['real'] = average_time



#listanddicttype传入参数是这种格式的[{},{}]
def averagevalue(listanddicttype):
    valueaverage=0
    valueerrflag=0
    sum=0
    for item in listanddicttype:
        if item['value'] == '':
            print('值获取错误')
            valueerrflag = 1

        else:
            sum+=item['value']
    if valueerrflag == 0:
        valueaverage=sum/len(listanddicttype)
        #如果类型错误，就返回一个非常大的时间，使报告判断结果为不合格。
    else:
        valueaverage=999999999
    return valueaverage

#app_name_row,all3timesaveragetime:当前行和追加好的所有的3次平均时间/////后发现该函数在很多地方通用，只需传入当前行和所有追加好的数据即可。
def everyitemsaveget(app_name_row,allsavevalue):
    item_get=[]
    app_name=app_name_row['param'].split(":")[0]
    for item_get_item in allsavevalue:
        if (item_get_item['oad/omd']).find(app_name) >= 0 :
            item_get.append(item_get_item)
        else:
            pass
    print('item_get:',item_get)
    return item_get


def everageandmaxrate(everageormaxreal,old_cpu_usage_rate,old_memory_usage_rate,old_storage_usage_rate):
    Container_cpu_usage_rate=[]
    Container_memory_usage_rate=[]
    Container_storage_usage_rate=[]
    if everageormaxreal['save'].find('AVERAGECPU') >= 0:
        print('old_cpu_usage_rate:',old_cpu_usage_rate)
        Container_cpu_usage_rate = everyitemsaveget(everageormaxreal, old_cpu_usage_rate)
        everageormaxreal['real'] = averagevalue(Container_cpu_usage_rate)
    elif everageormaxreal['save'].find('AVERAGEM') >= 0:
        Container_memory_usage_rate = everyitemsaveget(everageormaxreal, old_memory_usage_rate)
        everageormaxreal['real'] = averagevalue(Container_memory_usage_rate)
    elif everageormaxreal['save'].find('AVERAGEI') >= 0:
        Container_storage_usage_rate = everyitemsaveget(everageormaxreal, old_storage_usage_rate)
        everageormaxreal['real'] = averagevalue(Container_storage_usage_rate)
    elif everageormaxreal['save'].find('MAXCPU') >= 0:
        Container_cpu_usage_rate = everyitemsaveget(everageormaxreal, old_cpu_usage_rate)
        everageormaxreal['real'] = max_value(Container_cpu_usage_rate)
    elif everageormaxreal['save'].find('MAXM') >= 0:
        Container_memory_usage_rate = everyitemsaveget(everageormaxreal, old_memory_usage_rate)
        everageormaxreal['real'] = max_value(Container_memory_usage_rate)
    elif everageormaxreal['save'].find('MAXI') >= 0:
        Container_storage_usage_rate = everyitemsaveget(everageormaxreal,old_storage_usage_rate)
        everageormaxreal['real'] = max_value(Container_storage_usage_rate)
    if int(everageormaxreal['real']) < int(everageormaxreal['expect']):
        everageormaxreal['result'] = '合格'
    else:
        everageormaxreal['result'] = '不合格'
    return None

#rtucurrenttime:这种格式的字符串20200229232522
#返回某个月有多少分钟数
def monthhavemuchminutes(strtime):
    monthtominutes=0
    monthhaveday=0
    year = int(strtime[:4])
    month = int(strtime[4:6])
    if (month == 1 or month == 3 or month == 5 or month == 7 or month == 8 or month == 10 or month == 12):
        monthhaveday=31
    elif (month == 4 or month == 6 or month == 9 or month == 11):
        monthhaveday = 30
    elif (month == 2 and ((year % 4 == 0 and year % 100 != 0) or year % 400 == 0)):
        monthhaveday = 29
    else:
        monthhaveday = 28
    monthtominutes = monthhaveday * 1440
    return monthtominutes
#计算电压合格率的函数（日月均可计算）
def v_2131_calculation(rtucurrentime,lower_time,upper_time):
    v_2131_day=[]
    v_2131_month = []
    monitor_time_day =1440
    monitor_time_month=monthhavemuchminutes(rtucurrentime)
    passlimitrate_day = float('%.4f' % ((lower_time+upper_time)/monitor_time_day))
    passlimitrate_month = float('%.4f' % ((lower_time+upper_time)/monitor_time_month))
    passrate_day = 1-passlimitrate_day
    passrate_month = 1 - passlimitrate_month
    if upper_time>0:
        over_upperlimit_time_day=upper_time
        over_lowerlimit_time_day = 0
        over_upperlimit_time_month = upper_time
        over_lowerlimit_time_month = 0
    else:
        over_upperlimit_time_day = 0
        over_lowerlimit_time_day = lower_time
        over_upperlimit_time_month = 0
        over_lowerlimit_time_month = lower_time
    v_2131_day=[monitor_time_day,passrate_day*10000,passlimitrate_day*10000,over_upperlimit_time_day,over_lowerlimit_time_day]
    v_2131_month=[monitor_time_month,passrate_month*10000,passlimitrate_month*10000,over_upperlimit_time_month,over_lowerlimit_time_month]
    return v_2131_day,v_2131_month
#计算电压越限时间
def volpasstime(delayrow,meter_v,v_passrate_param_set):
    upper_time = 0
    lower_time = 0
    # 20201223新增 安徽或山东越上限和越下限计算流程：山东和安徽是没有上上限和下下限限制的。
    if delayrow['rtutype'] == '安徽' or delayrow['rtutype'] == '山东':
        if  meter_v < v_passrate_param_set[3]:
            lower_time = float(delayrow['delay']) / 60
        elif v_passrate_param_set[2] < meter_v :
            upper_time = float(delayrow['delay']) / 60
        else:
            pass
    else:
        if v_passrate_param_set[1] < meter_v < v_passrate_param_set[3]:
            lower_time = float(delayrow['delay']) / 60
        elif v_passrate_param_set[2] < meter_v < v_passrate_param_set[0]:
            upper_time = float(delayrow['delay']) / 60
        else:
            pass

    return upper_time,lower_time
#计算每一相的电压越限时间（日、月），并存为 [{'oad/omd': 'A_VOLPASSRATE_DAY', 'value': v_2131_day}]
#delaylimitraterow：当前行 v_meter_all_set：电压设置值,v_passrate_param：电压合格率参数,a_volpassrate_day,b_volpassrate_day,c_volpassrate_day,a_volpassrate_month,b_volpassrate_month,c_volpassrate_month（A\B\C三相的日月电压合格率统计数据）
# rtucurrentime:终端当前时间

def v_threephase_dayandmonth(delaylimitraterow,v_meter_all_set,v_passrate_param,a_volpassrate_day,b_volpassrate_day,c_volpassrate_day,a_volpassrate_month,b_volpassrate_month,c_volpassrate_month,rtucurrentime):
    delaylimitraterowsave=delaylimitraterow['save']
    if delaylimitraterow['save'].find('A_EXPECT')>=0:
        print('电压合格率-实时 用例，会直接在save中存6个变量名')
    elif delaylimitraterow['save'].find('VOLPASSRATE')>=0:
        delaylimitraterowsavesix = delaylimitraterowsave.replace(":", ',').split(',')
        daysavechoice = delaylimitraterowsavesix[0] + '_' + delaylimitraterowsavesix[1]
        monthsavechoice = delaylimitraterowsavesix[0] + '_' + delaylimitraterowsavesix[2]
        delaylimitraterowsavesix = [daysavechoice, monthsavechoice]
        delaylimitraterowsavesix1 = ''
        for itemsave in delaylimitraterowsavesix:
            for ABCitem in 'ABC':
                delaylimitraterowsavesix1 += ABCitem + '_' + itemsave + ','
        delaylimitraterowsavesix1 = delaylimitraterowsavesix1[:-1].split(",")
        print(delaylimitraterowsavesix1)
    for item in v_meter_all_set:
        if item['oad/omd'] == 'V_METER_A' or item['oad/omd'] == 'V_M0_A':
            upper_time, lower_time= volpasstime(delaylimitraterow,item['value'],v_passrate_param)
            #a_volpassrate_daynew每次计算获得的新的合格率参数。
            a_volpassrate_daynew, a_volpassrate_monthnew=v_2131_calculation(rtucurrentime,lower_time,upper_time)
            delaylimitraterow['save']=delaylimitraterowsavesix1[0]
            a_volpassrate_day = covervalue(delaylimitraterow,a_volpassrate_day, a_volpassrate_daynew)
            delaylimitraterow['save'] = delaylimitraterowsavesix1[3]
            a_volpassrate_month = covervalue(delaylimitraterow, a_volpassrate_month, a_volpassrate_monthnew)
        elif item['oad/omd'] == 'V_METER_B'or item['oad/omd'] == 'V_M0_B':
            upper_time, lower_time= volpasstime(delaylimitraterow,item['value'],v_passrate_param)
            #a_volpassrate_daynew每次计算获得的新的合格率参数。
            b_volpassrate_daynew, b_volpassrate_monthnew=v_2131_calculation(rtucurrentime,lower_time,upper_time)
            delaylimitraterow['save'] = delaylimitraterowsavesix1[1]
            b_volpassrate_day = covervalue(delaylimitraterow,b_volpassrate_day, b_volpassrate_daynew)
            delaylimitraterow['save'] = delaylimitraterowsavesix1[4]
            b_volpassrate_month = covervalue(delaylimitraterow, b_volpassrate_month, b_volpassrate_monthnew)
        elif item['oad/omd'] == 'V_METER_C' or item['oad/omd'] == 'V_M0_C':
            upper_time, lower_time= volpasstime(delaylimitraterow,item['value'],v_passrate_param)
            #a_volpassrate_daynew每次计算获得的新的合格率参数。
            c_volpassrate_daynew, c_volpassrate_monthnew=v_2131_calculation(rtucurrentime,lower_time,upper_time)
            delaylimitraterow['save'] = delaylimitraterowsavesix1[2]
            c_volpassrate_day = covervalue(delaylimitraterow,c_volpassrate_day, c_volpassrate_daynew)
            delaylimitraterow['save'] = delaylimitraterowsavesix1[5]
            c_volpassrate_month = covervalue(delaylimitraterow, c_volpassrate_month, c_volpassrate_monthnew)
        else:print('v_meter_all_set获取错误')
        #获取ABC电压合格率结束，回复该行save值，供下次测试该用例使用。
    delaylimitraterow['save']=delaylimitraterowsave
    logging.info('a_volpassrate_day：' + str(a_volpassrate_day))
    logging.info('b_volpassrate_day：' + str(b_volpassrate_day))
    logging.info('c_volpassrate_day：' + str(c_volpassrate_day))
    logging.info('a_volpassrate_month：' + str(a_volpassrate_month))
    logging.info('b_volpassrate_month：' + str(b_volpassrate_month))
    logging.info('c_volpassrate_month：' + str(c_volpassrate_month))
    return a_volpassrate_day,b_volpassrate_day,c_volpassrate_day,a_volpassrate_month,b_volpassrate_month,c_volpassrate_month

#结算日冻结的冻结时间获取
def setdayhour(tru_time):
    saveday = '01'
    savehour = '01'
    v_settletime = [[29, 0], [31, 1]]
    for item in v_settletime:
        if int(tru_time[6:8]) == item[0]:
            saveday = str(item[0]).zfill(2)
            savehour = str(item[1]).zfill(2)
    V_2021_freeze = tru_time[:6] + saveday + savehour + '0000'
    return V_2021_freeze

 # 获取电压合格率：2021200  其中，5004 5005的2021是当前日0点    5006月冻结是当前月1日0点,分钟取15整数倍，小时取当前小时、00分
def v_20210200_get(dt_m_list,tru_time):
    V_2021_freeze=''
    if dt_m_list[1] == '50020200':
        rtu_current_time_min = str(int(tru_time[-4:-2]) // 15 * 15).zfill(2)
        V_2021_freeze = tru_time[:-4] + rtu_current_time_min + '00'
    elif dt_m_list[1] == '50030200':
        V_2021_freeze = tru_time[:-4] + '0000'
    elif dt_m_list[1] == '50040200' :
        V_2021_freeze = tru_time[:-6] + '000000'
    elif dt_m_list[1] == '50050200':
        V_2021_freeze = setdayhour(tru_time)
    elif dt_m_list[1] == '50060200':
        V_2021_freeze = tru_time[:-8] + '01000000'
    else:
        print('请检查用例，是否获取过对应冻结值')
    return V_2021_freeze
#21310200,21320200,21330200直接召测,saveitem,phase,dayc,monc:存储列表，哪一相，日选择，月选择
def V_danyandmon(vol_day,vol_month,phase,dayc,monc, dt_m_list_dataitem):
    dt_m_list_dataitem=[]
    for saveitemday in vol_day:
        if saveitemday['oad/omd'] == phase + dayc:
            dt_m_list_dataitem.append(saveitemday['value'])
        else:pass
    if dayc.find('NULL') >= 0:
        dt_m_list_dataitem.append([0,0,0,0,0])
    else:pass
    for saveitemmon in vol_month:
        if saveitemmon['oad/omd'] == phase + monc:
            dt_m_list_dataitem.append(saveitemmon['value'])
        else:
            pass
    if monc.find('NULL') >= 0:
        dt_m_list_dataitem.append([0,0,0,0,0])
    else:pass


    return dt_m_list_dataitem


#dayandmonchoice,dt_m_list, a_volpassrate_day, b_volpassrate_day, c_volpassrate_day, a_volpassrate_month, #b_volpassrate_month, c_volpassrate_month:根据 save列来确定需要读取的总内容、6012召测的标识，预存的日月电压合格率。
# rtuormeter:用来确定，是统计对象是终端还是电表，终端与电表要比较的 OAD获取方式不一样，而且终端的话，需要获取20210200，故在流程中要做区分。
#rtu_current_time:通过终端当前时间来获取20210200的时间。
def getvrate(getexceptvpassrow,dt_m_list, a_volpassrate_day, b_volpassrate_day, c_volpassrate_day, a_volpassrate_month, b_volpassrate_month, c_volpassrate_month,rtuormeter,rtu_current_time):
    if len(dt_m_list) == 1:
        dt_m_list_passrate = dt_m_list
    else:
        if rtuormeter == 1:
            dt_m_list_passrate=dt_m_list[5:]
        else:
            dt_m_list_passrate = dt_m_list[2:]
    #最多的是日和月一起召测的，是6个。
    dt_m_list_data=['','','','','','']
    dayandmonthchoice=getexceptvpassrow['param']
    dayandmonthchoice=dayandmonthchoice.replace(':',',').split(',')
    dayexpectchoice=dayandmonthchoice[1]
    monexpectchoice=dayandmonthchoice[2]
    for item in range(len(dt_m_list_passrate)):
        if dt_m_list_passrate[item] == '20210200':
            dt_m_list_data[item]=v_20210200_get(dt_m_list,rtu_current_time)
        if dt_m_list_passrate[item] == '21310201':
            for item1 in a_volpassrate_day:
                #这里与参数有关，是变量
                if item1['oad/omd']== 'A_VOLPASSRATE_' + dayexpectchoice:
                    dt_m_list_data[item] = item1['value']
        elif dt_m_list_passrate[item] == '21320201':
            for item1 in b_volpassrate_day:
                #这里与参数有关，是变量
                if item1['oad/omd']== 'B_VOLPASSRATE_'+dayexpectchoice:
                    dt_m_list_data[item] = item1['value']
        elif dt_m_list_passrate[item] == '21330201':
            for item1 in c_volpassrate_day:
                #这里与参数有关，是变量
                if item1['oad/omd']== 'C_VOLPASSRATE_'+dayexpectchoice:
                    dt_m_list_data[item] = item1['value']
        elif dt_m_list_passrate[item] == '21310202':
            for item1 in a_volpassrate_month:
                #这里与参数有关，是变量
                if item1['oad/omd']== 'A_VOLPASSRATE_'+monexpectchoice:
                    dt_m_list_data[item] = item1['value']
        elif dt_m_list_passrate[item] == '21320202':
            for item1 in b_volpassrate_month:
                #这里与参数有关，是变量
                if item1['oad/omd']== 'B_VOLPASSRATE_'+monexpectchoice:
                    dt_m_list_data[item] = item1['value']
        elif dt_m_list_passrate[item] == '21330202':
            for item1 in c_volpassrate_month:
                #这里与参数有关，是变量
                if item1['oad/omd']== 'C_VOLPASSRATE_'+monexpectchoice:
                    dt_m_list_data[item] = item1['value']
        elif dt_m_list_passrate[item] == '21310200':
            dt_m_list_data = V_danyandmon(a_volpassrate_day,a_volpassrate_month, 'A_VOLPASSRATE_', dayexpectchoice, monexpectchoice, dt_m_list_data[item])
        elif dt_m_list_passrate[item] == '21320200':
            dt_m_list_data = V_danyandmon(b_volpassrate_day,b_volpassrate_month, 'B_VOLPASSRATE_', dayexpectchoice, monexpectchoice,
                                                dt_m_list_data[item])
        elif dt_m_list_passrate[item] == '21330200':
            dt_m_list_data = V_danyandmon(c_volpassrate_day,c_volpassrate_month, 'C_VOLPASSRATE_', dayexpectchoice, monexpectchoice,
                                                dt_m_list_data[item])
        else:
            pass
            #如果不是日月一起召测，就只取与标识对应的值就可以了。（前3项）--- dt_m_list_data=['','','','','','']的前三项
        if dt_m_list_passrate[item] not in ['21330200','21310200','21320200']:
            dt_m_list_data=dt_m_list_data[:len(dt_m_list_passrate)]
        else:
            pass
    #生成的列表中没有空值，说明是获取到预期值了。
    if '' not in dt_m_list_data:
        getexceptvpassrow['real']=str(dt_m_list_data)
        getexceptvpassrow['result'] = '合格'
    else:
        getexceptvpassrow['real'] = 'fail'
        getexceptvpassrow['result'] = '不合格'
    logging.info('电压合格率召测标识dt_m_list_passrate:'+str(dt_m_list_passrate))
    logging.info('电压合格率 预期读取 结果dt_m_list_data:' + str(dt_m_list_data))
    return dt_m_list_passrate,dt_m_list_data

#定义电压合格率比较函数（单纯的5个值比较，会多次调用，所以单独做了一个函数，可以重复使用）
def vpassrate_com(expectvpass,realvpassdeal,errorlist,dt_m_vpass_list, vpassratecomrow):
    comparefalseflag = 0
    #如果预期值中有空值，说明预期值获取有误，结果肯定是不合格的。不用往下执行了。
    #日或者月中，预期值中有全空值，做一个标志位。
    daynullflag = 0
    monnullflag = 0
    if '' in expectvpass:
        vpassratecomrow['real'] = 'fail'
        vpassratecomrow['result'] = '不合格'
        return True
    else:pass
    if expectvpass[0] == [0,0,0,0,0]:
        daynullflag = 1
    else:pass
    if expectvpass[1] == [0,0,0,0,0]:
        monnullflag = 1
    else:pass
    for comnum in range(len(expectvpass)):
        # 如果召测的结果中有空值，预期值为全0，就认为合格，不需要再比较。日为空值，跳过日，月为空值跳过月。其它情况自动化可以判异常。#realvpassdeal[comnum] == ['',''] 在处理获取到的值时，空值可能被处理成多种情况。目前的改写方法比较简单，不用修改整个结构，只考虑几种特情况（月为空，日为空，或2者均为空）。
        if (daynullflag == 1) and ((realvpassdeal[comnum] == ['','']) or (realvpassdeal[comnum][0] == '') ) and  (comnum == 0) :
            if (realvpassdeal[comnum][0] == ''):
                realvpassdeal[1] = realvpassdeal[0][1:] + realvpassdeal[1]
            continue
        if (monnullflag == 1) and ((realvpassdeal[comnum] == [])  or (realvpassdeal[comnum] == ['']) ) and  (comnum == 1) :
            continue
        for numindex in range(len(expectvpass[comnum])):
            #如果日电压合格率的预期值为NULL，那么只比[0,0,0,0,0]中的后三项。
            if (comnum == 0) and (daynullflag == 1)  and (numindex == 0 or numindex == 1):
                continue
            # 如果月电压合格率的预期值为NULL，那么只比[0,0,0,0,0]中的后三项。
            if (comnum == 1) and (monnullflag == 1) and (numindex == 0 or numindex == 1):
                continue
            error = float(errorlist[numindex])
            logging.info('expectvpass[comnum][numindex]:' + str(expectvpass[comnum][numindex]))
            logging.info('realvpass[comnum][numindex]' + str(realvpassdeal[comnum][numindex]))
            if abs(expectvpass[comnum][numindex] - float(realvpassdeal[comnum][numindex])) > error:
                logging.info(dt_m_vpass_list[comnum] + ':' + str(numindex) + ':误差过大' + 40 * '*')
                comparefalseflag = 1
            else:
                logging.info(dt_m_vpass_list[comnum] + ':' + str(numindex) + ':比较通过')
    if comparefalseflag==1:
        vpassratecomrow['real']='fail'
        vpassratecomrow['result'] = '不合格'
    else:
        vpassratecomrow['real'] = 'OK'
        vpassratecomrow['result'] = '合格'
    return True

#电压合格率比较函数----电表电压合格率
def comparevpassratem(expectvpass,realvpass, vpassratecomrow,dt_m_vpass_list):
    realvpassdeal=[]
    realvpass=read6012list(realvpass[0])[4:]
    logging.info('expectvpass:' + str(expectvpass))
    logging.info('realvpass:'+str(realvpass))
    for realvpassitem in realvpass:
        realvpassdeal.append(realvpassitem.split(","))
    print('realvpassdeal:',realvpassdeal)
    errorlist=vpassratecomrow['expect'].split(',')
    vpassrate_com(expectvpass,realvpassdeal,errorlist,dt_m_vpass_list, vpassratecomrow)
    return True

#电压合格率比较函数---终端数据筛选电压合格率---5004,5005,5006,等
def comparevpassratem0(expectvpass,realvpass, vpassratecomrow,dt_m_vpass_list):
    realvpassdeal=[]
    realvpass=read6012list(realvpass[0])
    if dt_m_vpass_list[0] in ['21310200', '21320200', '21330200']:
        errorlist = vpassratecomrow['expect'].split(',')
        dt_m_vpass_list = [dt_m_vpass_list[0][:-1] + '1',dt_m_vpass_list[0][:-1] + '2']
        realvpass=realvpass[0].split(',')
        realvpass =[realvpass[0:5],realvpass[5:10]]
        logging.info('expectvpass:' + str(expectvpass))
        logging.info('realvpass:' + str(realvpass))
        vpassrate_com(expectvpass,realvpass,errorlist,dt_m_vpass_list, vpassratecomrow)
    else:
        realvpass_2021 = realvpass[0]
        expectvpass_2021 = expectvpass[0]
        #复位流程后，召测上次的数据，不再比较2021。
        if  vpassratecomrow['save'] == 'T_NCOM':
            pass
        else:
            if realvpass_2021 == expectvpass_2021:
                logging.info('realvpass_2021:' + str(realvpass_2021))
                logging.info('expectvpass_2021:' + str(expectvpass_2021))
                logging.info('电压合格率2021比较通过')
            else:
                logging.info('realvpass_2021:' + str(realvpass_2021))
                logging.info('expectvpass_2021:' + str(expectvpass_2021))
                logging.info('电压合格率2021误差大，比较失败')
                vpassratecomrow['real'] = '2021比较失败'
                vpassratecomrow['result'] = '不合格'
                return True
        realvpass = [realvpass[1:6],realvpass[6:11],realvpass[11:16]]
        expectvpass = expectvpass[1:]
        logging.info('expectvpass:' + str(expectvpass))
        logging.info('realvpass:'+str(realvpass))
        errorlist=vpassratecomrow['expect'].split(',')
        vpassrate_com(expectvpass,realvpass,errorlist,dt_m_vpass_list[1:], vpassratecomrow)
    return True


#能源控制器出厂参数设置和检查特殊判断合格不合格流程：
def factoryparamcheckssh(paramrow,readdataparam):
    if paramrow['param']=='devcfg -mac':
        datalen = paramrow['expect']
        lenexpect = int(datalen[-2:])
        checkdata = readdataparam[-17:]
        checkdatalist = checkdata.split(":")
        num00 = 0
        for itemcheack in checkdatalist:
            if itemcheack == "00":
                num00 += 1
            else:
                pass
        if num00 < 3 and len(readdataparam) == lenexpect:
            paramrow['real'] = readdataparam
            paramrow['result'] = '合格'
        else:
            paramrow['real'] = readdataparam
            paramrow['result'] = '不合格'
    elif paramrow['param']=='devctl -e':
        lenesn = paramrow['expect']
        if readdataparam[-18:][:-2] != '0000000000000000' and readdataparam[-23:-18] == "esn :" and len(readdataparam) == int(lenesn[-2:]):
            paramrow['real'] = readdataparam
            paramrow['result'] = '合格'
        else:
            paramrow['real'] = readdataparam
            paramrow['result'] = '不合格'
    elif paramrow['param']=='devcfg -sn':
            sn = paramrow['expect']
            expectsnlist = sn.split(',')
            snow = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            expectsnlist[2] = snow[:8]
            expectsnlist[3] = expectsnlist[3][-4:]
            expectsnstr = ''
            for item in expectsnlist:
                expectsnstr += item
            print('readdataparam[-26:-2]:',readdataparam[-26:-2])
            print('expectsnstr:',expectsnstr)
            if readdataparam[-26:-2] == expectsnstr:
                paramrow['real'] = readdataparam
                paramrow['result'] = '合格'
            else:
                paramrow['real'] = readdataparam
                paramrow['result'] = '不合格'
    elif paramrow['param']=='ps -aux|grep taskManager':
        paramrow['real'] = readdataparam
        if readdataparam.find(paramrow['expect']) >= 0:
            paramrow['result'] = "合格"
        else:
            paramrow['result'] = "不合格"
    elif paramrow['param'] == 'netstat -an':
        mqttexpect = paramrow['expect']
        paramrow['real'] = readdataparam
        if (readdataparam.find(mqttexpect) >= 0):
            paramrow['result'] = "合格"
        else:
            paramrow['result'] = "不合格"
    elif paramrow['param'] == 'devcfg -md':
        expectmdlist = readdataparam.split(':')
        snow = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        expectmdlist[1] = ":" + snow[:4] + "-" + snow[4:6] + "-" + snow[6:8]+expectmdlist[1][10:]
        expectmdstr = ''
        for itemmd in expectmdlist:
            expectmdstr += itemmd
        print("expectmdstr:",expectmdstr)
        print("readdataparam:", readdataparam)
        paramrow['real'] = readdataparam
        if expectmdstr == readdataparam:
            paramrow['result'] = "合格"
        else:
            paramrow['result'] = "不合格"
    elif paramrow['param'] == 'devcfg -mf':
       paramrow['real'] = readdataparam
       if  readdataparam[-8:-2] == paramrow['expect']:
           paramrow['result'] = "合格"
       else:
           paramrow['result'] = "不合格"



    else:print('增加出厂参数检查判断流程处理')
    return True


# def getdeviceresult(realrow,computerreply):
#     confindflag=0
#     realdata=""
#     confinddata=""
#     if realrow['save'].find('FACTORTSPECIALPARAM')>=0:
#         factoryparamcheckssh(realrow,computerreply)
#     elif realrow['save'] == 'STATICMSG':
#         realrow['real'] = computerreply
#         print({'aaa': realrow['expect']})
#         if realrow['expect'] == computerreply:
#             realrow['result'] = "合格"
#         else:
#             realrow['result'] = "不合格"
#     else:
#         expectlist=realrow['expect'].split(",")
#         for item in expectlist:
#             if computerreply.find(item) >= 0:
#                 wefind = True
#                 realdata+=item+","
#             #这个分支来判断几个容器或应用是否都在正常的运行状态，正常不做处理，异常，需要将结果置为不合格。
#             elif item.find("_") >= 0:
#                 statusandtimes=item.split("_")
#                 if computerreply.count(statusandtimes[0])== int(statusandtimes[1]):
#                     pass
#                 else:
#                     confindflag = 1
#                     confinddata +="容器状态错误，"
#             else:
#                 wefind = False
#                 confinddata += item+","
#                 confindflag = 1
#         if realrow['expect'] == 'NULL' :
#             if computerreply =='Zgdky@guest123\r\n\r\n[sudo] password for sysadm: \r\n':
#                 realrow['real'] = "OK"
#                 realrow['result'] = "合格"
#             else:
#                 realrow['real'] = computerreply
#                 if realrow['param'].find('appm')>=0:
#
#                     realrow['result'] = "不合格"
#                 else:
#                     realrow['result'] = "合格"
#         else:
#             if confindflag == 1:
#                 realrow['real']="查询失败项:"+confinddata[:-1]
#                 realrow['result'] = "不合格"
#             else:
#                 realrow['real'] = "查询到数据项："+realdata[:-1]
#                 realrow['result'] = "合格"
#     return True

# 检查配置文件etype oop mrtu
def readexcelappid(lfile, sname, etype):
    workbook = xlrd.open_workbook(lfile)
    rtem = {}
    for ns in workbook.sheet_names():
        sheet2 = workbook.sheet_by_name(ns)
        if ns == sname:
            for ir in range(0, sheet2.nrows):  # excel行数循环
                rowvalue = {}
                y = 0
                if etype == "appid":
                    rowlist = APPIDROWLIST
                    seq = 0
                    for y in rowlist:
                        rowvalue[y] = sheet2.cell(ir, seq).value
                        seq += 1
                    rtem[ir] = rowvalue
                elif etype == "rturead":
                    rowlist = RTUREADLIST
                    seq = 0
                    for y in rowlist:
                        rowvalue[y] = sheet2.cell(ir, seq).value
                        seq += 1
                    rtem[ir] = rowvalue
                elif etype == TASKROWLIST:
                    rowlist = TASKROWLIST
                    seq = 0
                    for y in rowlist:
                        rowvalue[y] = sheet2.cell(ir, seq).value
                        seq += 1
                    rtem[ir] = rowvalue
                elif etype == HARDWAREROWLIST:
                    rowlist = HARDWAREROWLIST
                    seq = 0
                    for y in rowlist:
                        rowvalue[y] = sheet2.cell(ir, seq).value
                        seq += 1
                    rtem[ir] = rowvalue
                else:
                    print('增加表格列类型')
    return rtem

#追加表格存储报告的流程，需要增加报告名称‘合格’、‘不合格’判断流程
def renameexcel(pathadd):
    if pathadd.find('不合格') >= 0:
        pass
    else:
        os.rename(pathadd,pathadd.replace('合格','不合格'))
    return None

def reportappend(testname,pathadd, allexcellist, TASKROWLIST,hardexhibitionname,hardjceachtimeflag,nowtimemin,path_experiment,APPENDREPORT,rtutypelist,test_acc_times,testtimes,appendexresult):
    if testname == "状态监测循环测试":
        write_excel_xlsx_append(pathadd, allexcellist, TASKROWLIST, hardexhibitionname, hardjceachtimeflag,nowtimemin, path_experiment,rtutypelist)
    elif testname in APPENDREPORT:
        write_excel_xlsx_append(pathadd, allexcellist, TASKROWLIST, '', 0,0 , '',rtutypelist)
    if (test_acc_times == testtimes) and (appendexresult == 1):
        renameexcel(pathadd)
    else:pass


def write_excel_xlsx_appid(filepath,allexcellist,exceltype,hardexhibitionname,hardjceachtimeflag,nowtimemin,path_experiment):
    sheetnum = 0
    try:
        workbook = openpyxl.load_workbook(filepath)
    except:
        workbook = openpyxl.Workbook()
        print('追加存储表格出现except')
    for worksheet in workbook:# 在工作簿中的所有表格中循环
        value = allexcellist[sheetnum]
        index = len(value)  # 获取需要写入数据的行数
        ##写入测试时间
        if (filepath.find('状态监测循环测试') >=0) or (filepath.find('停上电') >=0) or (filepath.find('硬件复位') >=0) :
            timerow = {index: {'nu':'测试时间：', 'rtutype': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:19], 'cut': '',
                     'name': '', 'secure': '', 'vadd': '', 'addrtype': '', 'caddr': '', 'op': '', 'oad_omd': '',
                     'param': '', 'save': '', 'delay': '', 'expect': '', 'real': '', 'result': ''}}
            value.update(timerow)
        else:pass
        index = len(value)  # 获取需要写入数据的行数
        ##
        maxRows = worksheet.max_row  # 获取表格中已存在的数据的行数
        # 追加写入数据，注意是从i+rows_old行开始写入
        for i in range(0, index):
            columnc = 0
            for j in exceltype:
                worksheet.cell(row=maxRows + i + 1, column=columnc + 1, value=str(value[i][j]))
                columnc += 1
                #暂时不需要几分钟存一次报告。先屏蔽，有需要再修改。
                # if (filepath.find('状态监测循环测试') >=0) and (value[i][j].find("!") >= 0) and ((hardjceachtimeflag == 1) or (nowtimemin%2== 0)):
                if (filepath.find('状态监测循环测试') >= 0) and (value[i][j].find("!") >= 0) and (value[i]['oad_omd']!= '60120300'):
                    # 展示硬件测试的结果
                    hardwareresultshow(value[i][j], filepath,'',hardexhibitionname,path_experiment)
        sheetnum += 1
    workbook.save(filepath)
    #这个用例中，实际值中有“异常”2个字，会用红色显示。
    if (filepath.find('状态监测循环测试') >=0) or (filepath.find('停上电') >=0) or (filepath.find('硬件复位') >=0) :
        write_excel_xlsx_style(filepath, allexcellist, exceltype)
    return None

def write_excel_xlsx_append(filepath,allexcellist,exceltype,hardexhibitionname,hardjceachtimeflag,nowtimemin,path_experiment,rtutypelist):
    sheetnum = 0
    try:
        workbook = openpyxl.load_workbook(filepath)
    except:
        workbook = openpyxl.Workbook()
        print('追加存储表格出现except')
    for worksheet in workbook:# 在工作簿中的所有表格中循环
        value = allexcellist[sheetnum]
        index = len(value)  # 获取需要写入数据的行数
        ##写入测试时间
        if (filepath.find('状态监测循环测试') >=0) or (filepath.find('停上电') >=0) or (filepath.find('硬件复位') >=0) :
            timerow = {index: {'nu':'测试时间：', 'rtutype': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:19], 'cut': '',
                     'name': '', 'secure': '', 'vadd': '', 'addrtype': '', 'caddr': '', 'op': '', 'oad_omd': '',
                     'param': '', 'save': '', 'delay': '', 'expect': '', 'real': '', 'result': ''}}
            value.update(timerow)
        else:pass
        index = len(value)  # 获取需要写入数据的行数
        ##
        maxRows = worksheet.max_row  # 获取表格中已存在的数据的行数
        # 追加写入数据，注意是从i+rows_old行开始写入
        # 实际要写入的行，与原来表格不一致了（只挑选对应终端类型的输出的最终报告中）。所以，新表格的行号用一个新的变量表示
        rownewexl = 0
        for i in range(0, index):
            if (value[i]['rtutype'] not in rtutypelist) and (i > 0) and (value[i]['nu'] != '测试时间：'):
                pass
            else:
                columnc = 0
                for j in exceltype:
                    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
                    value[i][j] = ILLEGAL_CHARACTERS_RE.sub(r'', str(value[i][j]))
                    worksheet.cell(row=maxRows + rownewexl + 1, column=columnc + 1, value=str(value[i][j]))
                    columnc += 1
                    #暂时不需要几分钟存一次报告。先屏蔽，有需要再修改。
                    # if (filepath.find('状态监测循环测试') >=0) and (value[i][j].find("!") >= 0) and ((hardjceachtimeflag == 1) or (nowtimemin%2== 0)):
                    if (filepath.find('状态监测循环测试') >= 0) and (value[i][j].find("!") >= 0)and (value[i]['oad_omd']!= '60120300'):
                        # 展示硬件测试的结果
                        hardwareresultshow(value[i][j], filepath,'',hardexhibitionname,path_experiment)
                rownewexl += 1
        sheetnum += 1
    workbook.save(filepath)
    #这个用例中，实际值中有“异常”2个字，会用红色显示。
    if (filepath.find('状态监测循环测试') >=0) or (filepath.find('停上电') >=0) or (filepath.find('硬件复位') >=0) :
        write_excel_xlsx_style(filepath, allexcellist, exceltype)
    return None



#当表格中的某些内容需要进行特殊处理时
def write_excel_xlsx_style(filepath,allexcellist,exceltype):
    styleexcel_no_style = readexcelappid(filepath, 'Sheet1', exceltype)
    allexcellist = [styleexcel_no_style]
    sheetnum = 0
    try:
        workbook = Workbook(filepath)
        ws0 = workbook.add_worksheet()
    except:
        workbook = openpyxl.Workbook()
        print('追加存储表格出现except')
    value = allexcellist[sheetnum]
    index = len(value)  # 获取需要写入数据的行数
    # maxRows = ws0. # 获取表格中已存在的数据的行数
    # 追加写入数据，注意是从i+rows_old行开始写入
    for i in range(0, index):
        columnc = 0
        for j in exceltype:
            if  filepath.find('状态监测') >=0 and j == 'real':
                red_word = ['异常']
                # 定义颜色
                red_color = workbook.add_format({'color': 'red', 'bold': True})
                item_flag, item_rich_string = str_2_rich_string_content(s=str(value[i][j]),
                                                                                rich_word_list=red_word,
                                                                                color_type=red_color)
                # 如果包含需要标记颜色的字段, 就以富文本（特殊格式）的形式写入
                if item_flag:
                    ws0.write_rich_string( i, columnc, *item_rich_string)
                else:
                    ws0.write( i, columnc,str(value[i][j]))
            elif ((filepath.find('状态监测') >=0)or (filepath.find('停上电') >=0) or (filepath.find('硬件复位') >=0)) and (j == 'result'):
                if value[i][j] == "不合格":
                    bold = workbook.add_format({'color': 'red','bold': 1,'bg_color': '#FFFF00'})
                    ws0.write(i, columnc,  str(value[i][j]), bold)
                else:
                    ws0.write(i, columnc, str(value[i][j]))
            else:
                ws0.write( i, columnc, str(value[i][j]))
            columnc += 1
    workbook.close()

def write_excel_xlsx_jccomrow(filepath,allexcellist,sheetname):
    sheetnum = 0
    try:
        workbook = openpyxl.load_workbook(filepath)
        print('打开指定路径下的表格')
    except:
        workbook = openpyxl.Workbook()
        print('创建了一个空白的Excel')
    # 不使用第一个sheet,所以先删除。
    if "Sheet" not in workbook.sheetnames:
        pass
    else:
        ws = workbook["Sheet"]
        workbook.remove(ws)
    if sheetname in workbook.sheetnames:
        worksheet = workbook.get_sheet_by_name(sheetname)
    else:
        worksheet = workbook.create_sheet(sheetname)
    maxRows = worksheet.max_row  # 获取表格中已存在的数据的行数
    #处理表格的模块本身有问题，不应该在没有行的情况下默认为1行，所以程序做了特殊处理，强制处理为初始行为0行。
    if maxRows == 1 and worksheet['A1'].value == None:
        maxRows = 0
    for i in range(len(allexcellist)):
        valuedata = allexcellist[i]
        index = len(valuedata)  # 获取需要写入数据的行数
        # 追加写入数据，注意是从i+rows_old行开始写入
        for j in range(0, index):
            try:
                worksheet.cell(row=maxRows + i + 1, column= j + 1, value=str(valuedata[j]))
            except:
                worksheet.cell(row=maxRows + i + 1, column=j + 1, value=str(valuedata[j]))
    workbook.save(filepath)
    try:
        #如果想在执行过程中查看报告，需要在temp_report路径中拷贝出来查看，否则程序会因为文件被占用而报错，导致数据漏存。
        welookpath = (filepath.split('展示')[0] + '展示可拷贝查看' + ".xlsx").replace('report','temp_report')
        workbook.save(welookpath)
    except:
        print('不是展示类的报告，不走这个流程。忽略')
    return None



def write_excel_xlsx_jccomcolumn(filepath,allexcellist,sheetname,jc_column):
    if allexcellist == [['']] :
        return None
    sheetnum = 0
    try:
        workbook = openpyxl.load_workbook(filepath)
    except:
        workbook = openpyxl.Workbook()
        print('追加存储表格出现except')
    if sheetname in workbook.sheetnames:
        worksheet = workbook.get_sheet_by_name(sheetname)
    else:
        worksheet = workbook.create_sheet(sheetname)
    # maxColumns = worksheet.max_column  # 获取表格中已存在的数据的列数
    maxRows = worksheet.max_row
    getRows = -1
    acc_time_row = 0
    for row in worksheet.iter_rows(min_row=1, max_col=1, max_row=maxRows):
        for cell in row:
            getRows += 1
            if cell.value == allexcellist[0][0]:
                acc_time_row = getRows
            else:pass
    for i in range(len(allexcellist)):
        valuedata = allexcellist[i]
        index = len(valuedata)  # 获取需要写入数据的行数
        # 追加写入数据，注意是从i+rows_old行开始写入
        for j in range(0, index):
            try:
                worksheet.cell(row=acc_time_row + i + 1, column=jc_column+ j + 1, value=str(valuedata[j]))
            except:
                worksheet.cell(row=acc_time_row + i + 1, column=jc_column+ j + 1, value=str(valuedata[j]))
    workbook.save(filepath)
    return None



def write_excel_xlsx_hardware(filepath,allexcellist,exceltype,sheetname):
    sheetnum = 0
    try:
        workbook = openpyxl.load_workbook(filepath)
    except:
        workbook = openpyxl.Workbook()
        print('追加存储表格出现except')
    worksheet = workbook.get_sheet_by_name(sheetname)
    maxRows = worksheet.max_row  # 获取表格中已存在的数据的行数
    for j in range(len(allexcellist)):
        value = allexcellist[j]
        index = len(value)  # 获取需要写入数据的行数
        # 追加写入数据，注意是从i+rows_old行开始写入
        for i in range(0, index):
            try:
                worksheet.cell(row=maxRows + i + 1, column= j + 1, value=float(value[i]))
            except:
                worksheet.cell(row=maxRows + i + 1, column=j + 1, value=str(value[i]))
    workbook.save(filepath)
    return None

#因表格模块处理本身有问题，在处理第一行时，会认为第一行已经有内容了，故对第一次追加时做特殊处理。
def write_excel_xlsx_register(filepath,allexcellist,sheetname):
    sheetnum = 0
    try:
        workbook = openpyxl.load_workbook(filepath)
        print('打开指定路径下的表格')
    except:
        workbook = openpyxl.Workbook()
        print('创建了一个空白的Excel')
    # 不使用第一个sheet,所以先删除。
    if "Sheet" not in workbook.sheetnames:
        pass
    else:
        ws = workbook["Sheet"]
        workbook.remove(ws)
    if sheetname in workbook.sheetnames:
        worksheet = workbook.get_sheet_by_name(sheetname)
    else:
        worksheet = workbook.create_sheet(sheetname)
    maxRows = worksheet.max_row  # 获取表格中已存在的数据的行数
    #处理表格的模块本身有问题，不应该在没有行的情况下默认为1行，所以程序做了特殊处理，强制处理为初始行为0行。
    if maxRows == 1 and worksheet['A1'].value == None:
        maxRows = 0
    for j in range(len(allexcellist)):
        value = allexcellist[j]
        index = len(value)  # 获取需要写入数据的行数
        # 追加写入数据，注意是从i+rows_old行开始写入
        for i in range(0, index):
            try:
                worksheet.cell(row=maxRows + i + 1, column= j + 1, value=str(value[i]))
            except:
                worksheet.cell(row=maxRows + i + 1, column=j + 1, value=str(value[i]))
    workbook.save(filepath)
    try:
        #如果想在执行过程中查看报告，需要在temp_report路径中拷贝出来查看，否则程序会因为文件被占用而报错，导致数据漏存。
        welookpath = (filepath.split('展示')[0] + '展示可拷贝查看' + ".xlsx").replace('report','temp_report')
        workbook.save(welookpath)
    except:
        print('不是展示类的报告，不走这个流程。忽略')
    return None

def listappend(listold,excel):
    for item in excel:
        if (excel[item]['appid'] != '') and (excel[item]['appid'] != 'APPID'):
            listold.append(excel[item]['appid'])
        else:
            pass
    return listold

def Repeatedjudgment():
    usedappidlist=[]
    newappidlist=[]
    appidrightflag = False
    #定义一个要向已送检表格中追加的所有新表的列表
    execeladd = []

    # 当前文件的路径
    pwd = os.getcwd()+ "\config"
    usedappidexcelsmiOS = readexcelappid(pwd + '/' + paramfilesappid['Usedappid'], 'smiOS', 'appid')
    usedappidlist = listappend(usedappidlist, usedappidexcelsmiOS)
    usedappidexcelmapManager = readexcelappid(pwd + '/' + paramfilesappid['Usedappid'], 'mapManager', 'appid')
    usedappidlist = listappend(usedappidlist, usedappidexcelmapManager)
    usedappidexcelwirelessDCM = readexcelappid(pwd + '/' + paramfilesappid['Usedappid'], 'wirelessDCM', 'appid')
    usedappidlist = listappend(usedappidlist, usedappidexcelwirelessDCM)
    usedappidexceldbCenter = readexcelappid(pwd + '/' + paramfilesappid['Usedappid'], 'dbCenter', 'appid')
    usedappidlist = listappend(usedappidlist, usedappidexceldbCenter)
    usedappidexcelccoRouter = readexcelappid(pwd + '/' + paramfilesappid['Usedappid'], 'ccoRouter', 'appid')
    usedappidlist = listappend(usedappidlist, usedappidexcelccoRouter)
    print('本次送检前已使用APPID:',len(usedappidlist),usedappidlist)

    newappidexcelsmiOS = readexcelappid(pwd + '/' + paramfilesappid['newappid'], 'smiOS', 'appid')
    newappidlist = listappend(newappidlist, newappidexcelsmiOS)
    newappidexcelmapManager = readexcelappid(pwd + '/' + paramfilesappid['newappid'], 'mapManager', 'appid')
    newappidlist = listappend(newappidlist, newappidexcelmapManager)
    newappidexcelwirelessDCM = readexcelappid(pwd + '/' + paramfilesappid['newappid'], 'wirelessDCM', 'appid')
    newappidlist = listappend(newappidlist, newappidexcelwirelessDCM)
    newappidexceldbCenter = readexcelappid(pwd + '/' + paramfilesappid['newappid'], 'dbCenter', 'appid')
    newappidlist = listappend(newappidlist, newappidexceldbCenter)
    newappidexcelccoRouter = readexcelappid(pwd + '/' + paramfilesappid['newappid'], 'ccoRouter', 'appid')
    newappidlist = listappend(newappidlist, newappidexcelccoRouter)
    print('本次送检使用APPID:',len(newappidlist), newappidlist)
    appidall = usedappidlist + newappidlist
    execeladd = [newappidexcelsmiOS,newappidexcelmapManager,newappidexcelwirelessDCM,newappidexceldbCenter,newappidexcelccoRouter]
    for eachappid in newappidlist:
        if appidall.count(eachappid) > 1:
            appidrightflag = False
            print(f'重复项:{eachappid}')
            #发现有重复就不再循环，并提示APPID有重复。
            break
        else:
            appidrightflag = True
    if appidrightflag == True:
        #APPID这里是用不到后面3个参数的。是给硬件自动化使用的。这里不用管后面的3个参数
        write_excel_xlsx_appid(pwd  + '/' + paramfilesappid['Usedappid'], execeladd,APPIDROWLIST,'',0,0,'')
    else:
        print('APPID有重复，请核实')
    print('appidrightflag:',appidrightflag)
    return appidrightflag

def saveasexceldirect(temfile, tlist):
    wb = xlwt.Workbook()
    ws0 = wb.add_sheet('sheet')
    rowlist = VERSIONLIST
    for key in tlist:
        counter = 0
        y = 0
        for y in rowlist:
            try:
                ws0.write(key, counter, tlist[key][y])
            except:
                if isinstance(tlist[key][y], list):
                    strcuv = ''
                    for i in range(len(tlist[key][y])):
                        strcuv += str(tlist[key][y][i]) + '\n'
                    ws0.write(key, counter, strcuv)
            counter += 1
    wb.save(temfile)




    #
    #     print(usedappidexcelsmiOS[item]['appid'])
#excel追加函数（如果已经存在，直接追加，如果不存在，先创建）
def saveasexceladd(filepathappid,rtuappidvalue):
    if os.path.exists(filepathappid) == True:
        workbook = xlrd.open_workbook(filepathappid)  # 打开工作簿
        sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
        worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
        rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
        new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
        ws0 = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
    else:
        rows_old = 0
        wb = xlwt.Workbook()  # 新建一个工作簿
        ws0 = wb.add_sheet('sheet')
    rowlist = RTUREADLIST
    counter = 0
    #y = 0
    for y in rowlist:
        try:
            ws0.write(rows_old, counter, rtuappidvalue)
        except:
            if isinstance(rtuappidvalue, list):
                strcuv = ''
                for i in range(len(rtuappidvalue)):
                    strcuv += str(rtuappidvalue[i]) + '\n'
                ws0.write(rows_old, counter, strcuv)
        counter += 1
    if os.path.exists(filepathappid) == True:
            new_workbook.save(filepathappid)
    else:
        wb.save(filepathappid)
    # if (danotinflag == 1) and (tlist[dealrow]['name'].find(u':') < 0):

    return  True
#本批次送检蓝牙是否有重复判断
def repeatcheckbule(bulerepeatrow):
    bulealllist = []
    bulerepeatflag = 0
    bulepath = os.getcwd() + "//report"
    buleexcel = readexcelappid(bulepath + '/' + '蓝牙名称和mac.xls', 'sheet', 'rturead')
    for itemrow in buleexcel:
        bulealllist.append(buleexcel[itemrow]['data'])
    print('bulealllist:',bulealllist)
    logging.info('bulealllist:'+str(bulealllist))
    for itembule in bulealllist:
        if bulealllist.count(itembule) > 1:
            bulerepeatflag = 1
            logging.info('重复蓝牙:'+ str(itembule))
            break
        else:
            bulerepeatflag = 0
    if bulerepeatflag == 1:
        bulerepeatrow['real'] = '有重复蓝牙'
        bulerepeatrow['result'] = '不合格'
    else:
        bulerepeatrow['real'] = '蓝牙正确，无重复'
        bulerepeatrow['result'] = '合格'
    return True


def appidrightcheck(appidreadlistall):
    usedappidlist=[]
    newappidlist=[]
    newappidlist1 = []
    newappidlist2 = []
    newappidlist3 = []
    newappidlist4 = []
    newappidlist5 = []
    #定义一个要向已送检表格中追加的所有新表的列表
    execeladd = []

    # 当前文件的路径
    pwd = os.getcwd()+ "\config"
    usedappidexcelsmiOS = readexcelappid(pwd + '/' + paramfilesappid['Usedappid'], 'smiOS', 'appid')
    usedappidlist = listappend(usedappidlist, usedappidexcelsmiOS)
    usedappidexcelmapManager = readexcelappid(pwd + '/' + paramfilesappid['Usedappid'], 'mapManager', 'appid')
    usedappidlist = listappend(usedappidlist, usedappidexcelmapManager)
    usedappidexcelwirelessDCM = readexcelappid(pwd + '/' + paramfilesappid['Usedappid'], 'wirelessDCM', 'appid')
    usedappidlist = listappend(usedappidlist, usedappidexcelwirelessDCM)
    usedappidexceldbCenter = readexcelappid(pwd + '/' + paramfilesappid['Usedappid'], 'dbCenter', 'appid')
    usedappidlist = listappend(usedappidlist, usedappidexceldbCenter)
    usedappidexcelccoRouter = readexcelappid(pwd + '/' + paramfilesappid['Usedappid'], 'ccoRouter', 'appid')
    usedappidlist = listappend(usedappidlist, usedappidexcelccoRouter)

    newappidexcelsmiOS = readexcelappid(pwd + '/' + paramfilesappid['newappid'], 'smiOS', 'appid')
    newappidlist1 = listappend(newappidlist1, newappidexcelsmiOS)
    newappidexcelmapManager = readexcelappid(pwd + '/' + paramfilesappid['newappid'], 'mapManager', 'appid')
    newappidlist2 = listappend(newappidlist2, newappidexcelmapManager)
    newappidexcelwirelessDCM = readexcelappid(pwd + '/' + paramfilesappid['newappid'], 'wirelessDCM', 'appid')
    newappidlist3 = listappend(newappidlist3, newappidexcelwirelessDCM)
    newappidexceldbCenter = readexcelappid(pwd + '/' + paramfilesappid['newappid'], 'dbCenter', 'appid')
    newappidlist4 = listappend(newappidlist4, newappidexceldbCenter)
    newappidexcelccoRouter = readexcelappid(pwd + '/' + paramfilesappid['newappid'], 'ccoRouter', 'appid')
    newappidlist5 = listappend(newappidlist5, newappidexcelccoRouter)
    appidall = usedappidlist + newappidlist
    execeladd = [newappidexcelsmiOS,newappidexcelmapManager,newappidexcelwirelessDCM,newappidexceldbCenter,newappidexcelccoRouter]
    newappidlistall=[newappidlist1,newappidlist2,newappidlist3,newappidlist4,newappidlist5]
    countall = 0
    #标记APPID出错的数量
    appidtrueflag = 0
    appnum=0
    for eachapp in newappidlistall:
        count=0
        for eachappid in eachapp:
            count+=1
            countall+=1
            if appidreadlistall[appnum].count(eachappid) == 1:
                print("APPID正确",count,countall)
                logging.info("APPID正确:"+str(count) +':'+str(countall))
            else:
                print('标准文件中APPID在本次送检文件中查找到的次数：',appidreadlistall[appnum].count(eachappid))
                print("APPID错误",count,countall)
                logging.info("APPID错误:" + str(count) +':'+ str(countall))
                appidtrueflag += 1
        appnum +=1
    print('appidtrueflag:',appidtrueflag)
    logging.info('appidtrueflag:'+str(appidtrueflag))
    return  appidtrueflag
#本批次送检APPID是否正确
def rightcheckappid(appidrightcheckrow):
    appidalllist = []
    appidwrongflag = 0
    smiOS = []
    wirelessDCM = []
    mapManager = []
    dbCenter = []
    ccoRouter = []
    appidpath = os.getcwd() + "//report"
    appidexcel = readexcelappid(appidpath + '/' + 'APPID.xls', 'sheet', 'rturead')
    for itemrow in appidexcel:
        appidalllist.append(strtolist(appidexcel[itemrow]['data']))
    print('appidalllist:',appidalllist)
    for appiditem in appidalllist:
        for itemapp in appiditem:
            if itemapp[0] == 'M-smiOS':
                smiOS.append(itemapp[3])
            elif itemapp[0] == 'M-wirelessDCM':
                wirelessDCM.append(itemapp[3])
            elif itemapp[0] == 'M-mapManager':
                mapManager.append(itemapp[3])
            elif itemapp[0] == 'M-dbCenter':
                dbCenter.append(itemapp[3])
            elif itemapp[0] == 'M-ccoRouter':
                ccoRouter.append(itemapp[3])
    appidallreadlist = [smiOS, mapManager, wirelessDCM,dbCenter, ccoRouter]
    appidwrongflag = appidrightcheck(appidallreadlist)
    if appidwrongflag >= 1:
        appidrightcheckrow['real'] = 'APPID有重复或错误，请核实'
        appidrightcheckrow['result'] = '不合格'
    else:
        appidrightcheckrow['real'] = 'APPID正确，无重复'
        appidrightcheckrow['result'] = '合格'
    return True
#APPID 版本和日期处理，客户特殊需求
def versionanddatedeal(rtuappidall,versionrow,versionconfigexcel,path_experiment,reportfloderpath):
    versionresult = []
    for item in rtuappidall:
        versionresult.append(item[0:3])
    for num in range(len(versionresult)):
        versionresult[num][2]=versionresult[num][2][:10]
    versionrow['real'] = str(versionresult)
    for eachversion in versionresult:
        for eachnum in versionconfigexcel:
            if str(versionconfigexcel[eachnum]['name']).find(eachversion[0])>=0:
                versionconfigexcel[eachnum]['version'] = eachversion[1]
                versionconfigexcel[eachnum]['date'] = eachversion[2]
            else:pass
    now = datetime.datetime.now()
    if versionrow['save'].find('NJUDGE') >= 0:
        pathversion = os.getcwd() + r"\report\\" + f'{path_experiment}' + '\\' +  f'{path_experiment}' + '模组及组件版本信息.xls'
        print('pathversion:',pathversion)
    elif reportfloderpath != '':
        pathversion = reportfloderpath+ '\\' + '模组及组件版本信息.xls'
        print('pathversion:',pathversion)
    else:
        pathversion =  os.getcwd() + "//report//" + '模组和组件版本'+ now.strftime('%Y%m%d%H%M') + '.xls'
    saveasexceldirect(pathversion,versionconfigexcel)
    rightflag = 0
    findnum = 0
    if len(strtolist(versionrow['real'])) == len(strtolist(versionrow['expect'])):
        for eachrealname in strtolist(versionrow['real']):
            for eachexpectname in strtolist(versionrow['expect']):
                if eachrealname[0] == eachexpectname[0]:
                    if eachrealname[1:] == eachexpectname[1:]:
                        findnum += 1
                        pass
                    else:
                        versionrow['real'] = versionrow['real'] + eachrealname[0] + '：版本错误'
                        rightflag = 1
                        break
    else:
        rightflag = 1
    if (rightflag == 1) or(findnum != len(strtolist(versionrow['expect']))) :
        versionrow['result'] = '不合格'
    else:
        versionrow['result'] = '合格'
    if versionrow['save'].find('NJUDGE') >= 0:
        versionrow['result'] = '合格'
    return True
#模组版本和日期
def moudleversion(readdata,moudlerow,versionconfigexcel):
    moudleversionresult = []
    for item in readdata:
        if item[1].find('GBH')>=0:
            moudleversionresult.append(['HPLC',item[3],item[4]])
        elif item[1].find('GBR')>=0:
            moudleversionresult.append(['485', item[3], item[4]])
        elif item[1].find('GRX')>=0:
            moudleversionresult.append(['回路巡检', item[3], item[4]])
        elif item[1].find('GXX')>=0:
            moudleversionresult.append(['遥信模块', item[3], item[4]])
        elif item[1].find('GY')>=0:
            moudleversionresult.append(['4G', item[3], item[4]])
        elif item[1].find('GKX')>=0:
            moudleversionresult.append(['控制模块', item[3], item[4]])
    moudlerow['real'] = str(moudleversionresult)
    for eachversion in moudleversionresult:
        for eachnum in versionconfigexcel:
            if str(versionconfigexcel[eachnum]['name']).find(eachversion[0])>=0:
                versionconfigexcel[eachnum]['version'] = eachversion[1]
                versionconfigexcel[eachnum]['date'] = eachversion[2]
            else:pass
    rightflag = 0
    findnum = 0
    if len(strtolist( moudlerow['real'])) == len(strtolist( moudlerow['expect'])):
        for eachrealname in strtolist( moudlerow['real']):
            for eachexpectname in strtolist( moudlerow['expect']):
                if eachrealname[0] == eachexpectname[0]:
                    if eachrealname[1:] == eachexpectname[1:]:
                        findnum += 1
                        pass
                    else:
                        moudlerow['real'] =  moudlerow['real'] + eachrealname[0] + '：版本错误'
                        rightflag = 1
                        break
    else:
        rightflag = 1
    print('findnum:',findnum)
    if (rightflag == 1) or (findnum != len(strtolist( moudlerow['expect']))):
        moudlerow['result'] = '不合格'
    else:
        moudlerow['result'] = '合格'
    if moudlerow['save'].find('NJUDGE') >= 0:
        moudlerow['result'] = '合格'
    return versionconfigexcel

#模组版本和日期
def atversion(readdata,moudlerow,versionconfigexcel):
    moudleversionresult = []
    for itemnum in range(len(readdata)):
        if itemnum == 0:
            moudleversionresult.append(['回路巡检核心板',readdata[itemnum][1],readdata[itemnum][2]])
        elif itemnum == 1:
            moudleversionresult.append(['CT1', readdata[itemnum][1], readdata[itemnum][2]])
        elif itemnum == 2:
            moudleversionresult.append(['CT2', readdata[itemnum][1], readdata[itemnum][2]])
        else:
            moudleversionresult.append(['CT3', readdata[itemnum][1], readdata[itemnum][2]])
    moudlerow['real'] = str(moudleversionresult)
    for eachversion in moudleversionresult:
        for eachnum in versionconfigexcel:
            if str(versionconfigexcel[eachnum]['name']).find(eachversion[0])>=0:
                versionconfigexcel[eachnum]['version'] = eachversion[1]
                versionconfigexcel[eachnum]['date'] = eachversion[2]
            else:pass
    if moudlerow['real'] == moudlerow['expect']:
        moudlerow['result'] = '合格'
    else:
        moudlerow['result'] = '不合格'
    if moudlerow['save'].find('NJUDGE') >= 0:
        moudlerow['result'] = '合格'
    return versionconfigexcel

def nameandnum(parameow698):
    numrightflag = 0
    nameeachrightflag = 0
    namerightflag = 0
    reallist = strtolist(parameow698['real'])
    expectlist = strtolist(parameow698['expect'])
    if len(expectlist) == len(reallist):
        numrightflag = 1
    #数量正确再继续判断名称是否都正确range(len(parameow698['expect']))
    if numrightflag == 1:
        for item in expectlist:
            for num in range(len(reallist)):
                if reallist[num][0] == item:
                    nameeachrightflag = 1
                    break
            #=0,说明没找到
            if nameeachrightflag == 0:
                parameow698['result'] = '不合格'
                return None
        if nameeachrightflag == 1:
            parameow698['result'] = '合格'
            return None
    else:
        parameow698['result'] = '不合格'
        return None


#正则表达式，匹配数字和字母
def numletter(data):
    compileX = re.compile(r"[a-zA-Z0-9]")
    m = compileX.findall(data)
    n = ''
    for item in m:
        n += item
    return n

def  maxget(oldmax,newdata):
    if oldmax < newdata:
        oldmax = newdata
    else:
        pass
    return round(oldmax,2)

def minget(oldmin,newdata,newtimes):
    if newtimes == 1:
        oldmin = round(newdata,2)
    elif oldmin > newdata:
        oldmin = round(newdata,2)
    else:
        pass
    return oldmin

def averget(newtimes,oldave,newdata):
    oldave = (oldave*(newtimes - 1) + newdata)/newtimes
    return round(oldave,2)

def Imaxminave(real0list,Imaina,rowdata):
    if rowdata['save'].find('8路') >= 0:
        Imaina[6] += 1
        for itemnum in range(len(real0list)):
            Imaina[0][itemnum] = maxget(Imaina[0][itemnum],real0list[itemnum][3])
            Imaina[1][itemnum] = minget(Imaina[1][itemnum], real0list[itemnum][3],Imaina[6])
            Imaina[2][itemnum] = averget(Imaina[6],Imaina[2][itemnum],real0list[itemnum][3])
        rowdata['real'] +=  f'\n最大值：{str(Imaina[0])}\n最小值：{str(Imaina[1])}\n平均值：{str(Imaina[2])}'
    elif rowdata['save'].find('16路') >= 0:
        Imaina[6] += 1
        for itemnum in range(len(real0list)):
            Imaina[3][itemnum] = maxget(Imaina[3][itemnum], real0list[itemnum][3])
            Imaina[4][itemnum] = minget(Imaina[4][itemnum], real0list[itemnum][3], Imaina[6])
            Imaina[5][itemnum] = averget(Imaina[6], Imaina[5][itemnum], real0list[itemnum][3])
        rowdata['real'] += f'\n最大值：{str(Imaina[3])}\n最小值：{str(Imaina[4])}\n平均值：{str(Imaina[5])}'
    logging.info('漏电流最大值，最小值和平均值:'+str(Imaina))
    return  Imaina




def ff140200deal(real0,expect0,rowdata,Imaina):
    real0list = strtolist(real0)
    expect0list = strtolist(expect0)
    resuflag = 0
    imax,imin = Placefunction.calfromparam(rowdata)
    for num  in range(len(real0list)):
        if rowdata['save'].find('TRIP_EXTEND') >= 0:
            if (real0list[num][1] != 1):
                resuflag = 1
                rowdata['real'] +=  f'第{num+1}路全跳状态，合闸状态异常；'
                print(f'第{num+1}路全跳状态，合闸状态异常；')
        else:
            if (real0list[num][1] != 0) and (real0list[num][1] != 1):
                resuflag = 1
                print(f'第{num+1}路合闸状态异常；')
                rowdata['real'] += f'第{num+1}路合闸状态异常；'
            #接入和未接入都要判漏电流
            if ((expect0list[num] == 0) and ((real0list[num][3] > imax)or(real0list[num][3] <imin)))or((expect0list[num] == 1) and  (real0list[num][3] != 0) ):
                resuflag = 1
                print(f'第{num+1}路漏电流检测异常')
                rowdata['real'] += f'第{num+1}路漏电流检测异常；'
            if real0list[num][2] != expect0list[num]:
                resuflag = 1
                print(f'第{num+1}路接入状态异常；')
                rowdata['real'] += f'第{num+1}路接入状态异常；'
    if resuflag == 1:
        rowdata['result'] = '不合格'
    else:
        rowdata['result'] = '合格'
    Imaina = Imaxminave(real0list,Imaina,rowdata)
    return Imaina
def ff140300deal(real1,rowdata):
    real1list = strtolist(real1)
    resuflag = 0
    for num  in range(len(real1list)):
        if ( (real1list[num][1] < 200) or (real1list[num][1] > 400) ):
            resuflag = 1
            print(f'第{num+1}路脉宽检测异常')
            rowdata['real'] += f'第{num+1}路脉宽检测异常；'
    if resuflag == 1:
        rowdata['result'] = '不合格'
    else:
        rowdata['result'] = '合格'
    return resuflag,rowdata

def addrres(realdata):
    num0 = 0
    for item in realdata:
        if item == '0':
            num0 += 1
    if (num0 != len(realdata)) and (len(realdata)%2 == 0):
        return True
    else:
        return False
def handlerow(checktrueflag,item,parameow698):
    if (parameow698['save'].find('triprow') >= 0) and (parameow698['oad_omd'] in ['80008200']):
        num = 0
        num = int(parameow698['save'].split(',')[-1].split(':')[-1])
        item += num
    else:
        if parameow698['real'] == '有扩展模块档案':
            checktrueflag = 1
    return checktrueflag,item

#出厂参数中，698规约需要特殊处理结果标识，在这个函数中处理
def factoryparamcheck698(parameow698,versionconfigexcel,addr_experiment,path_experiment,h_l_temperatureflag,reportfloderpath,checktrueflag,ecumodverexpect,Imaina,fujianjcdata,dt_list_realData,dt_m_list,item):
    #2940稳定性测试的预期值要求比较多，单独做处理。
    if parameow698['save'].find('2940CON') >= 0:
        Placefunction.deal2940con(parameow698)
    else:
        if parameow698['oad_omd'] == 'F20D0200':
            idtypetrueflag = 0
            unqualified = ''
            realdata=strtolist(parameow698['real'])
            if parameow698['save'].find('FACTORTSPECIALPARAM_VERSION')>=0:
                versionconfigexcel=moudleversion(realdata,parameow698,versionconfigexcel)
            elif parameow698['save'].find('QUALITY_LOW') >= 0:
                parameow698['result'] = '合格'
                if len(realdata) == 5:
                    ecumodverexpect = parameow698['real']
                    checktrueflag = 1
                    parameow698['result'] = '合格'
                #质量底线连续召测3次模组版本，最后一次召测后，将checktrueflag清0，避免影响后续测试。
                if parameow698['save'].find('QLAST') >= 0:
                    checktrueflag = 0
                    ecumodverexpect = parameow698['real']
                    if len(realdata) != 5:
                        parameow698['result'] = '不合格'
                print('QLAST:',checktrueflag)
            elif parameow698['expect'].find('QUALITY_LOW') >= 0:
                if parameow698['real'] == ecumodverexpect:
                    checktrueflag = 1
                    parameow698['result'] = '合格'
                else:
                    parameow698['result'] = '合格'
                    checktrueflag = 0
                # 质量底线连续召测3次模组版本，最后一次召测后，将checktrueflag清0，避免影响后续测试。
                if parameow698['expect'].find('QLAST') >= 0:
                    checktrueflag = 0
                    if parameow698['real'] == ecumodverexpect:
                        parameow698['result'] = '合格'
                print('QLAST:', checktrueflag)
            else:
                #比较ID类型是否正确
                expectdata=strtolist(parameow698['expect'])
                if  realdata[0][2][18:22] != expectdata[0][2][18:22]:
                    idtypetrueflag = 1
                    unqualified += '槽位1模组类型错误！'
                del realdata[0][2]
                if  realdata[1][2][18:22] != expectdata[1][2][18:22]:
                    idtypetrueflag = 1
                    unqualified += '槽位2模组类型错误！'
                del realdata[1][2]
                if realdata[2][2][18:22] != expectdata[2][2][18:22]:
                    idtypetrueflag = 1
                    unqualified += '槽位3模组类型错误！'
                del realdata[2][2]
                if  realdata[3][2][18:22] != expectdata[3][2][18:22]:
                     idtypetrueflag = 1
                     unqualified += '槽位4模组类型错误！'

                del realdata[3][2]
                #防止4G没识别到
                try:
                    if  realdata[4][2][18:22] != expectdata[4][2][18:22]:
                        idtypetrueflag = 1
                        unqualified += '槽位5模组类型错误！'
                    del realdata[4][2]
                    del expectdata[4][2]
                except:
                    unqualified += '4G模块识别异常，请界面检查'
                del expectdata[0][2]
                del expectdata[1][2]
                del expectdata[2][2]
                del expectdata[3][2]
                if realdata==expectdata and idtypetrueflag == 0:
                    parameow698['result'] = '合格'
                else:
                    parameow698['result'] = '不合格：' + unqualified
        elif parameow698['oad_omd'] == 'F20F0200':
            rtuappid = strtolist(parameow698['real'])
            #APPID是否重复及文档更新使用
            if parameow698['save'].find('APPID_REPEAT_CHECK')>=0:
                appidrepeat = False
                appidrepeat = Repeatedjudgment()
                if appidrepeat == True:
                    parameow698['result'] = '合格'
                else:
                    parameow698['result'] = '不合格：APPID有重复项，请检查！！！并修改为正确的APPID后重新执行该用例，进行本次送检APPID的检查和录入工作'
            #APPID 是否正确及是否有重复（与标准文件比）
            elif parameow698['save'].find('APPID_RIGHT_CHECK')>=0:
                rightcheckappid(parameow698)
            #从APPID中获取版本号  日期信息
            elif parameow698['save'].find('FACTORTSPECIALPARAM_VERSION')>=0:
               versionanddatedeal(rtuappid,parameow698,versionconfigexcel,path_experiment,reportfloderpath)
            elif parameow698['save'].find('NAME_NUM') >= 0:
                nameandnum(parameow698)

            else:
                wirelessexpect = int(parameow698['save'].split(":")[1])
                if len(rtuappid) == wirelessexpect:
                    parameow698['result'] = '组件数量正确，不合格'
                    pathappid = os.getcwd() +  "\\report\\" + 'APPID' + '.xls'
                    #每次执行F20F0200时，都把读到的APPID信息存入APPID文件中。
                    saveasexceladd(pathappid,parameow698['real'])
                else:
                    parameow698['result'] = '组件数量不正确，不合格'
        elif (parameow698['oad_omd'] == '45100500') and ( len(parameow698['expect']) >0 ):
            num0 = 0
            for itemt in parameow698['real']:
                if itemt == "0":
                    num0 += 1
                else:
                    pass
            if num0 < 6 and len(parameow698['real']) == int(parameow698['expect'][4:6]):
                parameow698['result'] = '合格'
            else:
                parameow698['result'] = '不合格'
        elif parameow698['oad_omd'] == 'F20B0200':
            realbule = strtolist(parameow698['real'])
            if parameow698['save'].find('APPID_REPEAT_CHECK') >= 0:
                repeatcheckbule(parameow698)
            else:
                nameblue = realbule[0][1][0]
                markeblue = realbule[0][1][1]
                if len(nameblue) == 14 and  len(markeblue) == 12 and nameblue != '00000000000000' and markeblue!= '000000000000' and reversebit(nameblue[-6:]) == markeblue [:6]:
                    parameow698['result'] = '格式正确，需人工查看11台终端是否有重复的蓝牙名称，不合格'
                    pathbule = os.getcwd() + "\\report\\" + '蓝牙名称和mac' + '.xls'
                    # 每次执行F20B0200时，都把读到的蓝牙信息存入《蓝牙名称和mac》文件中。
                    saveasexceladd(pathbule, realbule)
                else:
                    parameow698['result'] = '格式不正确，不合格'
        elif parameow698['oad_omd'] == '40000200':
            if len(parameow698['real']) == 16 and parameow698['real'] != '00000000000000':
                if str(parameow698['save']).find('<=62') >= 0:
                    systime = parameow698['expect']
                    err_rtutime = timecompare(systime, parameow698['real'].strip("'"))
                    if (differseconds(err_rtutime) <=  62) and int(parameow698['real'].strip("'"))>=int(systime):
                        parameow698['result'] = '合格'
                    else:
                        parameow698['result'] = '不合格'
                elif str(parameow698['save']).find('AUTO_SYSTEM_TIME') >= 0:
                    systime = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
                    print(f'systime:{systime}')
                    err_rtutime = timecompare(systime, parameow698['real'].strip("'"))
                    if differseconds(err_rtutime) <= int(parameow698['expect']):
                        parameow698['result'] = '合格'
                    else:
                        parameow698['result'] = '不合格'
                else:
                    parameow698['result'] = '合格'
            else:
                parameow698['result'] = '不合格'
        elif (parameow698['oad_omd']) == '40020200':
            if len(parameow698['real']) == 14:
                parameow698['result'] = '合格'
            else:
                parameow698['result'] = '不合格'
        elif (parameow698['oad_omd']) == '40040200':
            if parameow698['real'] != parameow698['expect']:
                parameow698['result'] = '合格'
            else:
                parameow698['result'] = '不合格'
        elif (parameow698['oad_omd']) == '40010200':
            if parameow698['expect'].find("_常温") >= 0:
                h_l_temperatureflag = 1
            elif (parameow698['expect'].find("_高温") >= 0) or (parameow698['expect'].find("_低温") >= 0):
                h_l_temperatureflag = 2
            else:pass
            if (addrres(parameow698['real']) == True) and (parameow698['expect'].find("实验") >= 0):
                parameow698['result'] = '合格'
                addr_experiment = parameow698['real'].strip("'") + parameow698['expect']
                path_experiment = addr_experiment
                print(" addr_experiment:", addr_experiment)
            else:
                parameow698['result'] = '不合格'
        elif (parameow698['oad_omd']) == 'F1000200':
            addr_experiment = parameow698['real'].strip("'") + parameow698['expect']
            path_experiment = numletter(addr_experiment)
            print(" addr_experiment:", path_experiment)
            judge = readmeterresult()
            if judge.recresult(parameow698['real']) == True:
                parameow698['result'] = '合格'
        elif parameow698['oad_omd'] == '40400500':
            idtypetrueflag = 0
            unqualified = ''
            realdata = strtolist(parameow698['real'])
            if parameow698['save'].find('FACTORTSPECIALPARAM_VERSION')>=0:
                versionconfigexcel=atversion(realdata,parameow698,versionconfigexcel)
            else:
                # 比较ID类型是否正确
                expectdata = strtolist(parameow698['expect'])
                if realdata[0][0][18:22] != expectdata[0][0][18:22]:
                    idtypetrueflag = 1
                    unqualified += '槽位1模组类型错误！'
                del realdata[0][0]
                if realdata[1][0][18:22] != expectdata[1][0][18:22]:
                    idtypetrueflag = 1
                    unqualified += '槽位2模组类型错误！'
                del realdata[1][0]
                if realdata[2][0][18:22] != expectdata[2][0][18:22]:
                    idtypetrueflag = 1
                    unqualified += '槽位3模组类型错误！'
                del realdata[2][0]
                if realdata[3][0][18:22] != expectdata[3][0][18:22]:
                    idtypetrueflag = 1
                    unqualified += '槽位4模组类型错误！'
                del realdata[3][0]

                del expectdata[0][0]
                del expectdata[1][0]
                del expectdata[2][0]
                del expectdata[3][0]
                if realdata == expectdata and idtypetrueflag == 0:
                    parameow698['result'] = '合格'
                else:
                    parameow698['result'] = '不合格：' + unqualified
        elif parameow698['oad_omd'] == '23010B00' and parameow698['expect'].find("<") >= 0:
            if float(parameow698['real']) < float(parameow698['expect'].strip('<')):
                parameow698['result'] = '合格'
            else:
                parameow698['result'] = '不合格'
        #时段功控，依据不同时段的定值来获取不同的预期值（能源使用）
        elif parameow698['oad_omd'] == '23011100' and parameow698['save'].find("ACC_T") >= 0:
            tcexpect = strtolist(parameow698['expect'])
            tcexpect[0] =dlreaddata.tc_convalue()
            parameow698['expect'] = str(tcexpect).replace(' ','')
            if parameow698['real'] == parameow698['expect']:
                parameow698['result'] = '合格'
            else:
                parameow698['result'] = '不合格'
          #电控时，不再判功控定值
        elif parameow698['oad_omd'] == '23011100' and parameow698['save'].find("ECTL") >= 0:
            tcexpect = strtolist(parameow698['expect'])
            del tcexpect[0]
            tcreal = strtolist(parameow698['real'])
            del tcreal[0]
            if tcexpect == tcreal:
                parameow698['result'] = '合格'
            else:
                parameow698['result'] = '不合格'
        elif parameow698['oad_omd'] == '24010F00':
            print(parameow698['real'],parameow698['expect'])
            if float(strtolist(parameow698['real'])[0]) == float(parameow698['expect']):
                parameow698['result'] = '合格'
            else:
                parameow698['result'] = '不合格'
        elif (parameow698['oad_omd'] == '43000300')and (parameow698['param'].find('APLAY_R') < 0):
            if parameow698['expect'].find('len')>= 0:
                if len(strtolist(parameow698['real'])) == 6:
                    parameow698['result'] = '合格'
                else:
                    parameow698['result'] = '不合格'
            else:
                parameow698['result'] = '合格'
                parameow698['real'] = parameow698['real'].replace("\x00","")
                visave(parameow698['real'],path_experiment,'终端版本信息：厂商代码、软件版本号、软件版本日期、硬件版本号、硬件版本日期、厂商扩展信息')
        elif (parameow698['oad_omd'] == '45000500') and (str(parameow698['param']).find('APLAY_R') < 0):
            parameow698['result'] = '合格'
            parameow698['real'] = parameow698['real'].replace("\x00", "")
            visave(parameow698['real'],path_experiment,'右模块版本信息：厂商代码、软件版本号、软件版本日期、硬件版本号、硬件版本日期、厂商扩展信息')
        elif parameow698['oad_omd'] == 'F2090201':
            parameow698['result'] = '合格'
            parameow698['real'] = parameow698['real'].replace("\x00", "")
            visave(str(strtolist(parameow698['real'])[-1]),path_experiment,'左模块版本信息：厂商代码、芯片代码、版本日期、软件版本')
        elif parameow698['save'].find('ALWAYS_RIGHT') >= 0:
            parameow698['result'] = '合格'
        elif parameow698['oad_omd'] == 'FF140200':
            Imaina = ff140200deal(parameow698['real'],parameow698['expect'],parameow698,Imaina)
        elif parameow698['oad_omd'] == 'FF140300':
            ff140300deal(parameow698['real'],parameow698)
        elif parameow698['save'].find('FUJIANJC') >= 0:
            fujianjcdata = Placefunction.fujianjcdatadeal(parameow698,fujianjcdata)
            dt_list_realData, dt_m_list = fujianjcdata,['20000201','20000202','20000203','20010201','20010202','20010203','20040201','20040202','20040203','20040204']
        elif (parameow698['save'].find('EXTEND') >= 0) and (parameow698['oad_omd'] == '60000200'):
            parameow698['result'] = '合格'
            if (parameow698['real'].find('40360200') >= 0)and (parameow698['real'].find('F2010202') >= 0) and (parameow698['real'].find('516') >= 0):
                parameow698['real'] = '有扩展模块档案'
        else:
            print('请增加需要特殊处理的标识')
    if parameow698['save'].find('CONFIG_PARAM') >= 0:
        checktrueflag,item = handlerow(checktrueflag,item,parameow698)
    #6505自动化，读取到值就判合格。
    if parameow698['param'].find('APLAY_R') >= 0:
        if parameow698['result'] not in ['通讯异常','出现异常']:
            parameow698['result'] ='合格'
    return versionconfigexcel,addr_experiment,path_experiment,h_l_temperatureflag,checktrueflag,ecumodverexpect,Imaina,fujianjcdata,dt_list_realData,dt_m_list,item

def quoteresult(quoteresultflag,quoterow):
    if quoteresultflag == 1:
        quoterow['real'] = 'false'
        quoterow['result'] = '不合格'
    else:
        quoterow['real'] = '成功'
        quoterow['result'] = '合格'

    return True
def regionparamget(regionsetrow,regionparamlistold):
    regionparamnew=[]
    regionsetrowlist=strtolist(regionsetrow['param'])
    regionoad=regionsetrowlist[0]
    regioninterval=regionsetrowlist[-1]
    print('regioninterval:',regioninterval)
    if regioninterval.find('分')>=0:
        regioninterval=int(regioninterval[:-1])*60
    elif regioninterval.find('秒')>=0:
        regioninterval = int(regioninterval[:1])
    else:
        print("请增加区间统计除了分和秒以外的时间间隔计算")
    regionparamnew=[regionoad,regioninterval]
    regionparamlistold=covervalue(regionsetrow,regionparamlistold,regionparamnew)
    logging.info('regionparamlistold:'+str(regionparamlistold))
    return regionparamlistold

def region123get(regiongive,compinterval):
    if regiongive != 0:
        regionlist = [regiongive * compinterval,regiongive ]
    else:
        regionlist = [0, 0]
    return regionlist



#定义区间统计的预期读取存储流程 regionexrow,regiongetparam：当前参数和设置时存储好的[{'oad/omd': 'REGION_PARAM_1', 'value': ['200A0201', 60]}, {'oad/omd': 'REGION_PARAM_2', 'value': ['200A0202', 60]}, {'oad/omd': 'REGION_PARAM_3', 'value': ['200A0203', 60]}, {'oad/omd': 'REGION_PARAM_4', 'value': ['200A0204', 60]}]
def regionexpectsingle(regionexrowparam,regiongetparam):
    comoad=''
    comparamfindflag=0
    regionexpectlistsingle=[]
    regionparamlist =regionexrowparam.split(",")
    comparam=regionparamlist[0]
    for itemparam in regiongetparam:
        if itemparam['oad/omd'] == comparam:
            comgetparam=itemparam['value']
            comparamfindflag = 1
        else:pass
    if comparamfindflag == 0:
        print('请检查用例设计是否正确！区间参数没有配置！')
        return regionexpectlistsingle
    else:pass
    region1=int(regionparamlist[1][3:])
    region2 = int(regionparamlist[2][3:])
    region3 = int(regionparamlist[3][3:])
    comoad = comgetparam[0]
    cominterval = comgetparam[1]
    regionlist1=region123get(region1, cominterval)
    regionlist2 = region123get(region2, cominterval)
    regionlist3 = region123get(region3, cominterval)
    regionexpectlistsingle=[comoad,[regionlist1,regionlist2,regionlist3]]
    logging.info('regionexpectlistsingle:'+str(regionexpectlistsingle))
    return regionexpectlistsingle
#regionexrow,regiongetparam：当前行，[{'oad/omd': 'REGION_PARAM_1', 'value': ['200A0201', 60]}, ]
def regionexpect(regionexrow,regiongetparam):
    regionexpectlist=[]
    if regionexrow['param'].find(';')>=0:
        regionexpectlistmul=[]
        mulparamgivelist=regionexrow['param'].split(';')
        for mulparamgiveitem in mulparamgivelist:
            regionexpectlistmul.append(regionexpectsingle(mulparamgiveitem, regiongetparam, ))
        print('regionexpectlistmul:',regionexpectlistmul)
        for ecxeptitemnum in range(len(regionexpectlistmul[0][1])):
            if regionexpectlistmul[0][1][ecxeptitemnum] == [0,0]:
                regionexpectlistmul[0][1][ecxeptitemnum]=regionexpectlistmul[1][1][ecxeptitemnum]
            else:
                pass
        regionexpectlist=regionexpectlistmul[0]
    else:
        regionexpectlist = regionexpectsingle(regionexrow['param'], regiongetparam)
    logging.info('regionexpectlist:' + str(regionexpectlist))
    return regionexpectlist
#regioncomparerow,regionexpect  当前行和预期的区间值
def regioncompare(regioncomparerow,regionexpect):
    comfalseflag=0
    try:
        regionreallist = strtolist(regioncomparerow['real'])
    except NameError:
        print("区间统计等待时间需加长")

    logging.info('regionexpect:'+str(regionexpect))
    logging.info('regionreallist:' + str(regionreallist))
    if regionreallist[0] != regionexpect[0]:
        regioncomparerow['result'] = '不合格'
        logging.error('区间统计比较oad有误!')
        return True
    else:pass
    regiontolerancelist=strtolist(regioncomparerow['expect'])
    #根据预期值的不同，进行不同的处理。
    if regiontolerancelist[0].find("≤")>=0:
        regiontolerancelist[0]=int(regiontolerancelist[0].strip('≤±')[:-1])
        regiontolerancelist[1]=int(regiontolerancelist[1].strip('≤±'))
    else:
        regiontolerancelist[0]=int(regiontolerancelist[0][:1])
    for comnum in range(len(regionexpect[1])):
        # 存储某个区间比价后的误差值
        regionlisterr = []
        logging.info('预期区间'+str(comnum+1)+':'+ str(regionexpect[1][comnum]))
        logging.info('真实区间'+str(comnum+1)+':' + str(regionreallist[1][comnum]))
        for itemnum in range(len(regionexpect[1][comnum])):
            abserror=abs(regionexpect[1][comnum][itemnum]-regionreallist[1][comnum][itemnum])
            regionlisterr.append(abserror)
            if abserror>regiontolerancelist[itemnum]:
                comfalseflag=1
                logging.info('误差大，区间'+str(comnum+1)+'比较失败'+'*'*40)
            else:
                pass
        logging.info('区间' + str(comnum + 1) + '误差值:' + str(regionlisterr))
    if comfalseflag==1:
        regioncomparerow['result'] = '不合格'
    else:
        regioncomparerow['result'] = '合格'
    return True













def getexpectandcompare(regioncomrow,regionparam):
    #区间统计预期值获取流程
    try:
        regionexpectlist=regionexpect(regioncomrow,regionparam)
        regioncompare(regioncomrow,regionexpectlist)
    except IndexError:
        logging.info('区间统计有参数未配置或参数获取错误')
    return True
#极值统计：根据预期值来获取对应值
def expectget(ext_min_data_rtu_get,expect_paramnum):
    expect_data=''
    for item in ext_min_data_rtu_get:
        if item['oad/omd'] == expect_paramnum:
            expect_data = item['value']
            #找到了就跳出循环
            break
    return expect_data



def extgetexpect(extcomrow,ext_min_data_rtu_get,ext_max_data_rtu_get,ext_min_time_rtu_get,ext_max_time_rtu_get):
    ext_expect=['','','','']
    ext_expect_result=['','','','','']
    expect_param=extcomrow['param'].split(',')
    for itemparamnum in range(len(expect_param)):
        if expect_param[itemparamnum].find('MIN_DATA')>=0:
            ext_expect[itemparamnum]=expectget(ext_min_data_rtu_get,expect_param[itemparamnum])
        elif expect_param[itemparamnum].find('MAX_DATA')>=0:
            ext_expect[itemparamnum ] = expectget(ext_max_data_rtu_get, expect_param[itemparamnum])
        elif expect_param[itemparamnum].find('MIN_TIME') >= 0:
            ext_expect[itemparamnum ] = expectget(ext_min_time_rtu_get, expect_param[itemparamnum])
        elif expect_param[itemparamnum].find('MAX_TIME') >= 0:
            ext_expect[itemparamnum ] = expectget(ext_max_time_rtu_get, expect_param[itemparamnum])
        else:print('请检查极值统计用例 比较行 参数是否有误')
    if ext_expect[0][0]!=ext_expect[2][0]:
        print('极值统计用例设计有误')
        return ext_expect_result
    else:
        ext_expect_result=[ext_expect[0][0],float(ext_expect[0][1]),ext_expect[1].strip("'"),float(ext_expect[2][1]),ext_expect[3].strip("'")]
    logging.info('ext_expect_result:'+str(ext_expect_result))
    return ext_expect_result
#定义一个处理极值统计预期值格式的函数，exttolerancelistrow：当前比较行中的预期值列表
def extexpectdeal(exttolerancelistrow):
    for itemexnum  in range(len(exttolerancelistrow)):
        if exttolerancelistrow[itemexnum].find("%")>=0:
            exttolerancelistrow[itemexnum]=float(exttolerancelistrow[itemexnum].strip('≤±').strip('%'))/100
        else:
            exttolerancelistrow[itemexnum] = float(exttolerancelistrow[itemexnum].strip('≤±')[:-1])
    print('exttolerancelist:'+str(exttolerancelistrow))
    return exttolerancelistrow
#当前比较行，存储好的预期比较列表
def extcompare(extcomparerow, ext_expect_list):
    comfalseflag = 0
    exterrorlist = ['', '', '', '']
    try:
        extreallist = strtolist(extcomparerow['real'])
    except NameError:
        print("极值统计等待时间需加长")
    logging.info('extexpect:' + str(ext_expect_list))
    logging.info('extreallist:' + str(extreallist))
    if extreallist[0][:-1] != ext_expect_list[0][:-1]:
        extcomparerow['result'] = '不合格'
        logging.error('极值统计比较oad有误!')
        return True
    else:
        pass
    #极值和均值时，因台体零线电流波动，统计结果跳动很大，无法走正常判断流程。只要零线电流小于0.5A 即可判合格。
    if ext_expect_list[0] == '20010400':
        if ext_expect_list[1] < 0.5 and ext_expect_list[3] < 0.5:
            extcomparerow['result'] = '合格'
        else:
            extcomparerow['result'] = '不合格'
        return True
    else:
        pass
    exttolerancelist = strtolist(extcomparerow['expect'])
    exttolerancelist = extexpectdeal(exttolerancelist)
    logging.info('比较OAD为：' + ext_expect_list[0])
    for comnum in range(len(exttolerancelist)):
        logging.info('ext_expect_list:' + str(ext_expect_list[comnum + 1]))
        logging.info('extreallist:' + str(extreallist[comnum + 1]))
        if isinstance(ext_expect_list[comnum + 1], float):
            if ext_expect_list[comnum + 1] != 0:
                exterrorlist[comnum] = float(
                    '%.4f' % (abs(ext_expect_list[comnum + 1] - extreallist[comnum + 1]) /abs(ext_expect_list[comnum + 1])) )
            else:
                exterrorlist[comnum] = extreallist[comnum + 1]
            if exterrorlist[comnum] > exttolerancelist[comnum]:
                if comnum == 0:
                    logging.info("极大值比较失败,该项误差为：" + str(exterrorlist[comnum]) + ' 大于允许误差' + '*' * 20)
                else:
                    logging.info("极小值比较失败,该项误差为：" + str(exterrorlist[comnum]) + ' 大于允许误差' + '*' * 20)

                comfalseflag = 1
            else:
                if comnum == 0:
                    logging.info("极大值比较通过,该项误差为：" + str(exterrorlist[comnum]))
                else:
                    logging.info("极小值比较通过,该项误差为：" + str(exterrorlist[comnum]))
        else:
            exterrorlist[comnum] = differseconds(timecompare(ext_expect_list[comnum + 1], extreallist[comnum + 1]))
            if exterrorlist[comnum] > exttolerancelist[comnum]:
                if comnum == 1:
                    logging.info("极大值发生时间比较失败,该项误差为：" + str(exterrorlist[comnum]) + '秒 大于允许误差' + '*' * 20)
                else:
                    logging.info("极小值发生时间比较失败,该项误差为：" + str(exterrorlist[comnum]) + '秒 大于允许误差' + '*' * 20)

                comfalseflag = 1
            else:
                if comnum == 1:
                    logging.info("极大值发生时间比较通过,该项误差为：" + str(exterrorlist[comnum]) + '秒')
                else:
                    logging.info("极小值发生时间比较通过,该项误差为：" + str(exterrorlist[comnum]) + '秒')
    logging.info('预期误差：' + str(exttolerancelist))
    logging.info('实际误差：' + str(exterrorlist))
    if comfalseflag == 1:
        extcomparerow['result'] = '不合格'
    else:
        extcomparerow['result'] = '合格'
    return True

def extexpectsave(extexpectrow,ext_min_data_rtu,ext_max_data_rtu,ext_min_time_rtu,ext_max_time_rtu):
    if extexpectrow['save'].find(ext_min_data) >= 0:
        ext_min_data_rtu = covervalue(extexpectrow, ext_min_data_rtu,
                                              [extexpectrow['oad_omd'], extexpectrow['real']])
        logging.info('ext_min_data_rtu' + str(ext_min_data_rtu))
    else:
        pass
    if extexpectrow['save'].find(ext_max_data) >= 0:
        ext_max_data_rtu = covervalue(extexpectrow, ext_max_data_rtu,
                                              [extexpectrow['oad_omd'],extexpectrow['real']])
        logging.info('ext_max_data_rtu' + str(ext_max_data_rtu))
    else:
        pass
    if extexpectrow['save'].find(ext_min_time) >= 0:
        ext_min_time_rtu = covervalue(extexpectrow, ext_min_time_rtu, extexpectrow['real'])
        logging.info('ext_min_time_rtu' + str(ext_min_time_rtu))
    else:
        pass
    if extexpectrow['save'].find(ext_max_time) >= 0:
        ext_max_time_rtu =covervalue(extexpectrow, ext_max_time_rtu, extexpectrow['real'])
        logging.info('ext_max_time_rtu' + str(ext_max_time_rtu))
    else:
        pass
    if extexpectrow['real'] !='' and extexpectrow['real'] !=[]:
        extexpectrow['result']='合格'
    return ext_min_data_rtu,ext_max_data_rtu,ext_min_time_rtu,ext_max_time_rtu

#遥信、控制或回路巡检状态异常处理：异常状态标志位、当前行、哪个状态异常、存储报告路径、当前时间
def rsctlpabnormal(popup_use,statusabnomalflag,tasklistrow,typercl,path,nowt,hardexhibitionname,path_experiment):
    if os.path.exists(path) == True:
        pass
    else:
        hardwaresavepath = excelcopy(hardexhibitionname, '硬件自动化结果展示模板.xlsx',path_experiment)
    if typercl == '遥信':
        if tasklistrow['real'] == '通讯异常':
            pass
        else:
            statusabnomalflag[0] = 1
    elif  typercl == '控制':
        if tasklistrow['real'] == '通讯异常':
            pass
        else:
            statusabnomalflag[1] = 1
    elif  typercl == '回路巡检':
        if tasklistrow['real'] == '通讯异常':
            pass
        else:
            statusabnomalflag[2] = 1
    elif  typercl == '门节点':
        if tasklistrow['real'] == '通讯异常':
            pass
        else:
            statusabnomalflag[3] = 1
    tasklistrow['result'] = "不合格"
    if tasklistrow['real'] == '通讯异常':
        # write_excel_xlsx_hardware(path,
        #                           [['4G通讯异常，' + '异常发生时间：' + nowt]],
        #                           HARDWAREROWLIST, "测试结论")
        pass
    else:
        write_excel_xlsx_hardware(path,
                              [[typercl+'有变位:' + tasklistrow['real'] +'变位发生时间：'+ nowt ]],
                              HARDWAREROWLIST,"测试结论")
        if popup_use == 0:
            if win32api.MessageBox(0, nowt + "：" + typercl + "状态异常，请检查。异常状态：" + tasklistrow['real'] + "是否继续执行", "警告",
                                   win32con.MB_YESNO) == 6:
                pass
            else:
                os._exit(0)
        else:pass
    return statusabnomalflag, tasklistrow['result']

def realtimeget(rx_out_time):
    dt = bcdtodatetime(rx_out_time)
    delta = datetime.timedelta(seconds=30)
    realtimestart = (str(dt - delta))
    return '终端掉线时间：' + realtimestart

#计算掉线时长。暂时不用。
def lengthoftime(online_off_time,rx_out_time):
        try:
            #掉线时间
            offtime = ''
            rx_out_realtime = ''
            #等待超时时间为30秒，故检测到掉线会滞后30秒。该时间后续可能会调整。
            longtime = 30
            rx_out_realtime =  realtimeget(rx_out_time)
            offtime = rx_out_time
            onlitime = online_off_time[-1].replace('-','').replace(" ",'').replace(":",'')
            timestytle = timecompare(onlitime, offtime)
            secondtotal = differseconds(timestytle) + longtime
            online_off_time[-1] = '4G掉线时长约为' +str(secondtotal) + "秒；"+ rx_out_realtime +'；上线时间:' + online_off_time[-1]
            #掉线时间计算完成，将掉线发生时间清空，避免重复计算。
            rx_out_time = ''
        except:
            logging.error("4G上线时间记录出现异常")

        return online_off_time,rx_out_time


def specialdeal(tasklistrow,pulsevalue,checktrueflag,rsexpect,ctexpect,lpexpect,trytimes,trytimesflag,hardexhibitionname,statusabnomalflag,plusenumrightflag,hplcfailnum,rsfailnum,popup_use,path_experiment,bnm,h_l_temperatureflag,sta_agent,stagentsusconfig,reportfloderpath,repeatresult,longframeflag,drexpect):
    if str(tasklistrow['save']).find('PULSE_CONSTANT')>=0:
        if tasklistrow['op'] == '读取一个对象属性':
            if tasklistrow['real'] != '[]':
                pulsevalue = strtolist(tasklistrow['real'])[0][2]
                checktrueflag = 1
            else:
                checktrueflag = 0
        elif tasklistrow['op'] == '设置一个对象属性':
            pulsevalue = strtolist(tasklistrow['param'])[0][2]
        print('pulsevalue',pulsevalue)
        #获取打脉冲数量并依据脉冲常数计算出正向有功
    elif str(tasklistrow['save']).find('PULSENUM:')>=0:
        pulsenum = int(tasklistrow['save'].split(":")[1])
        paenergyexpect =  strtolist(tasklistrow['expect'])
        paenergyexpect[0] = (pulsenum/pulsevalue)*10000
        paenergyexpect[1] = (pulsenum / pulsevalue) * 10000
        tasklistrow['expect'] = str(paenergyexpect)
        if str(tasklistrow['param']).find('TRY_PM_TIMES:') >= 0:
            judge = readmeterresult()
            trytimes += 1
            trytimes_max = int(tasklistrow['param'].split(':')[1])
            if (trytimes <= trytimes_max) and (judge.recresult(tasklistrow['real']) == False):
                trytimesflag = 1
                if trytimes == trytimes_max:
                    # 达到重试次数后，不再重试，将对应标志位清0.
                    trytimes = 0
                    trytimesflag = 0
                    if tasklistrow['oad_omd'] == "24010700":
                        plusenumrightflag.append(str(1) + ":" + str(tasklistrow['real']))
                    elif tasklistrow['oad_omd'] == "24020700":
                        plusenumrightflag.append(str(1) + ":" + str(tasklistrow['real']))
                    elif tasklistrow['oad_omd'] == "24030700":
                        plusenumrightflag.append(str(1) + ":" + str(tasklistrow['real']))
                    elif tasklistrow['oad_omd'] == "24040700":
                        plusenumrightflag.append(str(1) + ":" + str(tasklistrow['real']))
                    else:pass
                else:
                    pass
            else:
                trytimes = 0
                trytimesflag = 0
                #只判总，不需要判在哪个费率上。
                if strtolist(tasklistrow['real'])[0] == paenergyexpect[0] :
                    tasklistrow['result'] = '合格'
                    if tasklistrow['oad_omd'] == "24010700":
                        plusenumrightflag.append(str(0) + ":" + str(strtolist(tasklistrow['real'])[0]) )
                    elif tasklistrow['oad_omd'] == "24020700":
                        plusenumrightflag.append(str(0) + ":" + str(strtolist(tasklistrow['real'])[0]) )
                    elif tasklistrow['oad_omd'] == "24030700":
                        plusenumrightflag.append(str(0) + ":" + str(strtolist(tasklistrow['real'])[0]) )
                    elif tasklistrow['oad_omd'] == "24040700":
                        plusenumrightflag.append(str(0) + ":" + str(strtolist(tasklistrow['real'])[0]) )
                    else:pass
                else:
                    if tasklistrow['oad_omd'] == "24010700":
                        plusenumrightflag.append(str(1) + ":" + str(strtolist(tasklistrow['real'])[0]))
                    elif tasklistrow['oad_omd'] == "24020700":
                        plusenumrightflag.append(str(1) + ":" + str(strtolist(tasklistrow['real'])[0]))
                    elif tasklistrow['oad_omd'] == "24030700":
                        plusenumrightflag.append(str(1) + ":" + str(strtolist(tasklistrow['real'])[0]))
                    elif tasklistrow['oad_omd'] == "24040700":
                        plusenumrightflag.append(str(1) + ":" + str(strtolist(tasklistrow['real'])[0]))
                    else:pass
                    tasklistrow['result'] = '不合格'
    elif (str(tasklistrow['save']).find('CONFIG_PARAM') >= 0) and (tasklistrow['oad_omd'] != '20500200'):
        if tasklistrow['oad_omd'] == 'F2030400':
            judge = readmeterresult()
            if judge.recresult(tasklistrow['real']) == False:
                checktrueflag = 0
            else:
                checktrueflag = 1
                tasklistrow['result'] = '合格'
        elif tasklistrow['oad_omd'] == '24010300' or tasklistrow['oad_omd'] == '24020300':
            if strtolist(tasklistrow['real'])[0] == 0 or strtolist(tasklistrow['real'])[1] == 0:
                checktrueflag = 0
            else:
                checktrueflag = 1
    elif str(tasklistrow['save']).find('EXPECT_STATE') >= 0:
        if tasklistrow['oad_omd'] == 'F2030200':
            rsexpect = tasklistrow['real']
            tasklistrow['result'] = '合格'
            print(f'rsexpect:{rsexpect}')
        elif tasklistrow['oad_omd'] == 'F2030600':
            drexpect = tasklistrow['real']
            tasklistrow['result'] = '合格'
        elif tasklistrow['oad_omd'] == 'F2050200':
            ctexpect = tasklistrow['real']
            tasklistrow['result'] = '合格'
        elif tasklistrow['oad_omd'] == '20500200':
            try:
                if isinstance(strtolist(tasklistrow['real']),list):
                    lpexpect = tasklistrow['real']
                    tasklistrow['result'] = '合格'
                else:
                    lpexpect = "回路巡检测试出现异常"
                    tasklistrow['result'] = '不合格'
            except:
                lpexpect = "回路巡检测试出现异常"
                tasklistrow['result'] = '不合格'

        else:pass
    elif str(tasklistrow['expect']).find('EXPECT_STATE') >= 0:
        mkdir(os.getcwd() + r"\report\\"+f'{path_experiment}\\')
        path = os.getcwd() + r"\report\\" +f'{path_experiment}\\'+ hardexhibitionname
        nowt = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:23]
        if tasklistrow['oad_omd'] == 'F2030200':
            tasklistrow['expect'] = rsexpect
            if tasklistrow['real'] == tasklistrow['expect']:
                tasklistrow['result'] = '合格'
                if  statusabnomalflag[0] == '不执行':
                    statusabnomalflag[0] = 0
                else:
                    pass
                if h_l_temperatureflag != 0:
                    if os.path.exists(path) == True:
                        pass
                    else:
                        hardwaresavepath = excelcopy(hardexhibitionname, '硬件自动化结果展示模板.xlsx', path_experiment)
                    write_excel_xlsx_hardware(path,
                                              [["遥信状态与准备工作一致" + tasklistrow['real'] + '记录时间：' + nowt]],
                                              HARDWAREROWLIST, "测试结论")

            else:
                statusabnomalflag, tasklistrow['result'] = rsctlpabnormal(popup_use,statusabnomalflag, tasklistrow, "遥信", path,
                                                                          nowt,hardexhibitionname,path_experiment)
        elif tasklistrow['oad_omd'] == 'F2030600':
            tasklistrow['expect'] = drexpect
            if tasklistrow['real'] == tasklistrow['expect']:
                tasklistrow['result'] = '合格'
                if  statusabnomalflag[3] == '不执行':
                    statusabnomalflag[3] = 0
                else:
                    pass
                if h_l_temperatureflag != 0:
                    if os.path.exists(path) == True:
                        pass
                    else:
                        hardwaresavepath = excelcopy(hardexhibitionname, '硬件自动化结果展示模板.xlsx', path_experiment)
                    write_excel_xlsx_hardware(path,
                                              [["门节点状态与准备工作一致" + tasklistrow['real'] + '记录时间：' + nowt]],
                                              HARDWAREROWLIST, "测试结论")

            else:
                statusabnomalflag, tasklistrow['result'] = rsctlpabnormal(popup_use,statusabnomalflag, tasklistrow, "门节点", path,
                                                                          nowt,hardexhibitionname,path_experiment)

        elif tasklistrow['oad_omd'] == 'F2050200':
            tasklistrow['expect'] = ctexpect
            if tasklistrow['real'] == tasklistrow['expect']:
                tasklistrow['result'] = '合格'
                if  statusabnomalflag[1] == '不执行':
                    statusabnomalflag[1] = 0
                else:
                    pass
            else:
                statusabnomalflag, tasklistrow['result'] = rsctlpabnormal(popup_use,statusabnomalflag, tasklistrow, "控制", path,
                                                                          nowt,hardexhibitionname,path_experiment)
        elif tasklistrow['oad_omd'] == '20500200':
            tasklistrow['expect'] = lpexpect
            if tasklistrow['real'] == "'NULL'":
                checktrueflag = 0
            else:
                checktrueflag = 1
            if (tasklistrow['real'] == tasklistrow['expect']) or ((tasklistrow['real'] == "'NULL'") and (tasklistrow['save'].find('CONFIG_PARAM') >= 0))  :
                tasklistrow['result'] = '合格'
                if  statusabnomalflag[2] == '不执行':
                    statusabnomalflag[2] = 0
                else:
                    pass
            else:
                statusabnomalflag, tasklistrow['result'] = rsctlpabnormal(popup_use,statusabnomalflag,tasklistrow,"回路巡检",path,nowt,hardexhibitionname,path_experiment)
        else:pass
    elif str(tasklistrow['save']).find('RS_FIRST_R') >= 0:
        if len(strtolist(tasklistrow['real'])) == 8:
            tasklistrow['result'] = '合格'
        else:
            tasklistrow['result'] = '不合格'
    elif str(tasklistrow['save']).find('FAIL_TIMES') >= 0:
        judge = readmeterresult()
        trytimes += 1
        trytimes_max =  int(tasklistrow['save'].split(':')[1].split(',')[0])
        if (trytimes <= trytimes_max) and (judge.recresult(tasklistrow['real']) == False):
            trytimesflag = 1
            if trytimes == trytimes_max:
                tasklistrow['result'] = "不合格"
                if tasklistrow['expect'].find('底线') >= 0:
                    pass
                else:
                    win32api.MessageBox(0, "HPLC组网异常，请检查", "警告", win32con.MB_OK)
                    os._exit(0)
            else:pass
        else:
            trytimes = 0
            trytimesflag = 0
        if  judge.recresult(tasklistrow['real']) == False:
            if tasklistrow['save'] == "FAIL_TIMES:0,#A相":
                if tasklistrow['real']  == '通讯异常':
                    pass
                else:
                    hplcfailnum += 1
                rsfailnum[4] = 1
            elif tasklistrow['save'] == "FAIL_TIMES:0,#B相":
                rsfailnum[9] = 1
                if tasklistrow['real']  == '通讯异常':
                    pass
                else:
                    rsfailnum[11] += 1
            elif tasklistrow['save'] == "FAIL_TIMES:0,#C相":
                rsfailnum[10] = 1
                if tasklistrow['real'] == '通讯异常':
                    pass
                else:
                    rsfailnum[12] += 1
            elif tasklistrow['save'] == "FAIL_TIMES:0,HPLC#单相07":
               sta_agent[4] += 1
               sta_agent[5] += 1
            elif tasklistrow['save'] == "FAIL_TIMES:0,HPLC#单相698":
               sta_agent[6] += 1
               sta_agent[7] += 1
            elif tasklistrow['save'] == "FAIL_TIMES:0,HPLC#三相07":
               sta_agent[8] += 1
               sta_agent[9] += 1
            elif tasklistrow['save'] == "FAIL_TIMES:0,HPLC#三相698":
               sta_agent[10] += 1
               sta_agent[11] += 1
            else:pass
            tasklistrow['result'] = "不合格"
        else:
            tasklistrow['result'] = "合格"
            if tasklistrow['save'] == "FAIL_TIMES:0,#A相":
                rsfailnum[4] = 1
            elif tasklistrow['save'] == "FAIL_TIMES:0,#B相":
                rsfailnum[9] = 1
            elif tasklistrow['save'] == "FAIL_TIMES:0,#C相":
                rsfailnum[10] = 1
            elif tasklistrow['save'] == "FAIL_TIMES:0,HPLC#单相07":
                sta_agent[5] += 1
            elif tasklistrow['save'] == "FAIL_TIMES:0,HPLC#单相698":
                sta_agent[7] += 1
            elif tasklistrow['save'] == "FAIL_TIMES:0,HPLC#三相07":
                sta_agent[9] += 1
            elif tasklistrow['save'] == "FAIL_TIMES:0,HPLC#三相698":
                sta_agent[11] += 1
            else:pass
    elif str(tasklistrow['expect']).find('FAIL_TIMES') >= 0:
        judge = readmeterresult()
        trytimes += 1
        trytimes_max =  int(tasklistrow['expect'].split(':')[1].split(',')[0])
        if (trytimes <= trytimes_max) and (judge.recresult(tasklistrow['real']) == False):
            trytimesflag = 1
            if str(tasklistrow['name']).find('CAN') >= 0:
                if win32api.MessageBox(0, str(tasklistrow['name'])+"：CAN抄读异常，请检查接线或参数配置", "提示", win32con.MB_YESNO) == 6:
                    pass
                else:
                    os._exit(0)
            else:
                if win32api.MessageBox(0, str(tasklistrow['name']) + "：485抄读异常，请检查接线或参数配置", "提示",
                                       win32con.MB_YESNO) == 6:
                    pass
                else:
                    os._exit(0)
            if trytimes == trytimes_max:
                if str(tasklistrow['name']).find('CAN') >= 0:
                    win32api.MessageBox(0, "CAN口通讯异常，请检查", "警告", win32con.MB_OK)
                else:
                    win32api.MessageBox(0, "485通讯异常，请检查", "警告", win32con.MB_OK)
                tasklistrow['result'] = "不合格"
                os._exit(0)
            else:pass
        else:
            trytimes = 0
            trytimesflag = 0
        if (judge.recresult(tasklistrow['real']) == False):
            if tasklistrow['expect'] == "FAIL_TIMES:0,485#1":
                rsfailnum[5] = 1
                if tasklistrow['real']  == '通讯异常':
                    pass
                else:
                    rsfailnum[0] += 1
            elif (tasklistrow['expect'] == "FAIL_TIMES:0,485#2"):
                rsfailnum[6] = 1
                if tasklistrow['real']  == '通讯异常':
                    pass
                else:
                    rsfailnum[1] += 1
            elif (tasklistrow['expect'] == "FAIL_TIMES:0,485#3"):
                rsfailnum[7] = 1
                if tasklistrow['real']  == '通讯异常':
                    pass
                else:
                    rsfailnum[2] += 1
            elif tasklistrow['expect'] == "FAIL_TIMES:0,485#4":
                rsfailnum[8] = 1
                if tasklistrow['real']  == '通讯异常':
                    pass
                else:
                    rsfailnum[3] += 1
            elif tasklistrow['expect'] == "FAIL_TIMES:0,CAN":
                rsfailnum[13] = 1
                if tasklistrow['real']  == '通讯异常':
                    pass
                else:
                    rsfailnum[14] += 1
            elif tasklistrow['expect'] == "FAIL_TIMES:0,485#07":
               sta_agent[0] += 1
               sta_agent[1] += 1
            elif tasklistrow['expect'] == "FAIL_TIMES:0,485#698":
               sta_agent[2] += 1
               sta_agent[3] += 1
            else: pass
            tasklistrow['result'] = "不合格"
        else:
            tasklistrow['result'] = "合格"
            if (tasklistrow['expect'] == "FAIL_TIMES:0,485#1") or (tasklistrow['expect'] =='01,FAIL_TIMES:0,485#1'):
                rsfailnum[5] = 1
            elif (tasklistrow['expect'] == "FAIL_TIMES:0,485#2") or (tasklistrow['expect'] =='01,FAIL_TIMES:0,485#2'):
                rsfailnum[6] = 1
            elif (tasklistrow['expect'] == "FAIL_TIMES:0,485#3") or (tasklistrow['expect'] =='01,FAIL_TIMES:0,485#3'):
                rsfailnum[7] = 1
            elif tasklistrow['expect'] == "FAIL_TIMES:0,485#4":
                rsfailnum[8] = 1
            elif tasklistrow['expect'] == "FAIL_TIMES:0,CAN":
                rsfailnum[13] = 1
            elif tasklistrow['expect'] == "FAIL_TIMES:0,485#07":
               sta_agent[1] += 1
            elif tasklistrow['expect'] == "FAIL_TIMES:0,485#698":
               sta_agent[3] += 1
            else:pass
    elif str(tasklistrow['save']).find('COMSPECIAL') >= 0:
        if (len(tasklistrow['real']) > 0) and (tasklistrow['real'] != u'终端没响应') and (tasklistrow['real'] != u'其它') and (
                tasklistrow['real'] != u'拒绝读写'):
            tasklistrow['result'] = u'合格'
        else:
            tasklistrow['result'] = u'不合格'
    elif str(tasklistrow['addrtype']).find('广播地址') >= 0 and str(tasklistrow['op']).find('设置一个对象属性') >= 0:
        tasklistrow['result'] = u'合格'
    elif str(tasklistrow['save']).find('AA_TIME') >= 0 :
        if tasklistrow['real'][:-4] == tasklistrow['expect'][:-4]:
            tasklistrow['result'] = u'合格'
        else:
            tasklistrow['result'] = u'不合格'
    elif str(tasklistrow['save']).find('AA_NOT_TIME') >= 0 :
        if tasklistrow['real'][:-4] == tasklistrow['expect'][:-4]:
            tasklistrow['result'] = u'不合格'
        else:
            tasklistrow['result'] = u'合格'
    elif str(tasklistrow['save']).find('SAVENAME_MAC') >= 0:
        nameblue = ''
        markeblue = ''
        nameblue = strtolist(tasklistrow['real'])[0][1][0]
        markeblue = strtolist(tasklistrow['real'])[0][1][1]
        # if len(nameblue) == 14 and len(
        #     markeblue) == 12 and nameblue != '00000000000000' and markeblue != '000000000000' :
        #     bnm = ['蓝牙名称：' + nameblue, '蓝牙MAC：'+ markeblue  ]
        #     tasklistrow['result'] = '合格'
        if len(markeblue) == 12 and nameblue != '00000000000000' and markeblue != '000000000000':
            bnm = ['蓝牙名称：' + nameblue, '蓝牙MAC：' + markeblue]
            tasklistrow['result'] = '合格'

        else:
            bnm = ['蓝牙名称读取有异常'+ nameblue , '蓝牙MAC读取有异常'+markeblue]
            tasklistrow['result'] = '不合格'
        print(f'bnm:{bnm}')
    #多次重复执行的行，不合格标志位被置一，结果就为不合格
    elif (str(tasklistrow['save']).find('REPEAT_EXECUTION') >= 0) and (repeatresult == 1):
        tasklistrow['result'] = '不合格'
    #长帧召测是否合格判断，并记录。
    elif str(tasklistrow['save']).find('TESTTIMES') >= 0:
        datapointnum  = int(str(tasklistrow['save']).split(":")[1])
        puncnum = tasklistrow['real'].count('!')
        framenum = tasklistrow['real'].count('\n')
        if (tasklistrow['real'].find('NULL') >= 0) or ((puncnum + framenum)!= (datapointnum - 1)) :
            tasklistrow['result'] = '不合格'
            longframeflag = 1
        else:
            tasklistrow['result'] = '合格'
            if longframeflag == 1:
                pass
            else:
                longframeflag = 0
        print('puncnum:',puncnum,'framenum:',framenum,'puncnum + framenum:',puncnum+framenum,)
    else:
        pass
    if (str(tasklistrow['real']).find('通讯异常') >= 0) and  (str(tasklistrow['expect']).find('通讯异常') < 0):
        tasklistrow['result'] = '通讯异常'
    else:
        pass
    #能源稳定性测试，用来处理代理成功率
    if (str(tasklistrow['expect']).find('FAIL_TIMES') >= 0) or (str(tasklistrow['save']).find('FAIL_TIMES') >= 0):
        stagentsusconfig = dlreaddata.sta_succ_ratesave(sta_agent,tasklistrow,stagentsusconfig,reportfloderpath)
    return pulsevalue,checktrueflag,rsexpect,ctexpect,lpexpect,trytimes,trytimesflag,statusabnomalflag,plusenumrightflag,hplcfailnum,rsfailnum,bnm,sta_agent,stagentsusconfig,longframeflag,drexpect



def comspecialdeal(tasktestrow,dt_m_list,testname,jc_column,jc_path,reportfloderpath):
    #将OAD加上注释
    for meani in range(len(dt_m_list)):
        if dt_m_list[meani] in OADTOMEAN.keys():
            dt_m_list[meani] = dt_m_list[meani] + OADTOMEAN[dt_m_list[meani]]
        else:
            pass
    if str(tasktestrow['save']).find('JC_FIRST') >= 0:
        nowtime = datetime.datetime.now()
        jc_path = reportfloderpath + testname +'展示'+ nowtime.strftime('%Y%m%d%H%M') + ".xlsx"
        write_excel_xlsx_jccomrow(jc_path,[dt_m_list,tasktestrow['real'].split(";")],testname)
        jc_column[0] = len(dt_m_list)
    elif str(tasktestrow['save']).find('JC_ONE_COLUMN') >= 0:
        write_excel_xlsx_jccomrow(jc_path, [tasktestrow['real'].split(";")], testname)
    elif str(tasktestrow['save']).find('JC_SECOND') >= 0:
        write_excel_xlsx_jccomcolumn(jc_path, [dt_m_list], testname,jc_column[0])
        write_excel_xlsx_jccomcolumn(jc_path, [tasktestrow['real'].split(";")], testname,jc_column[0])
        jc_column[1] = jc_column[0] + len(dt_m_list)
    elif str(tasktestrow['save']).find('JC_TWO_COLUMN') >= 0:
        write_excel_xlsx_jccomcolumn(jc_path, [tasktestrow['real'].split(";")], testname, jc_column[0])
    elif str(tasktestrow['save']).find('JC_THIRD') >= 0:
        write_excel_xlsx_jccomcolumn(jc_path, [dt_m_list], testname,jc_column[1])
        write_excel_xlsx_jccomcolumn(jc_path, [tasktestrow['real'].split(";")], testname,jc_column[1])
        jc_column[2] = jc_column[1] + len(dt_m_list)
    elif str(tasktestrow['save']).find('JC_THREE_COLUMN') >= 0:
        write_excel_xlsx_jccomcolumn(jc_path, [tasktestrow['real'].split(";")], testname, jc_column[1])
    elif str(tasktestrow['save']).find('JC_FOURTH') >= 0:
        write_excel_xlsx_jccomcolumn(jc_path, [dt_m_list], testname, jc_column[2])
        write_excel_xlsx_jccomcolumn(jc_path, [tasktestrow['real'].split(";")], testname, jc_column[2])
    elif str(tasktestrow['save']).find('JC_F_COLUMN') >= 0:
        write_excel_xlsx_jccomcolumn(jc_path, [tasktestrow['real'].split(";")], testname, jc_column[2])
    else:pass
    return jc_path,jc_column

#稳定性测试，处理上报数据使用
def freportshowre(sta_path,testname,dt_m_list,readlistdata,sheetname,reptype,reportfloderpath):
    nowtime = datetime.datetime.now()
    #检测到不是同一天的报告，就单独再保存一个路径。报告名称单位取到分钟就可以，没有用例可以测试那么快，达到分钟级别的。为了方便测试，暂用名称到分
    if (sta_path == '') or (sta_path[-19:-11] != nowtime.strftime('%Y%m%d%H%M')[:8]) :
        sta_path = reportfloderpath + testname + reptype + nowtime.strftime('%Y%m%d%H%M%S') + ".xlsx"
    else:
        pass
    try:
        workbook = openpyxl.load_workbook(sta_path)
        if (sheetname not in workbook.sheetnames) :
            write_excel_xlsx_jccomrow(sta_path, [dt_m_list,readlistdata], sheetname)
            # 分钟冻结，除首次外
        else:
            write_excel_xlsx_jccomrow(sta_path, [readlistdata], sheetname)
    except:
        write_excel_xlsx_jccomrow(sta_path, [dt_m_list, readlistdata], sheetname)
    return sta_path

def freportshow(sta_path,testname,dt_m_list,readlistdata,sheetname,reptype,reportfloderpath,oadlistold):
    nowtime = datetime.datetime.now()
    #检测到不是同一天的报告，就单独再保存一个路径。报告名称单位取到分钟就可以，没有用例可以测试那么快，达到分钟级别的。为了方便测试，暂用名称到分
    if (sta_path == '') or (sta_path[-19:-11] != nowtime.strftime('%Y%m%d%H%M')[:8]) :
        sta_path = reportfloderpath + testname + reptype + nowtime.strftime('%Y%m%d%H%M%S') + ".xlsx"
    else:
        pass
    try:
        workbook = openpyxl.load_workbook(sta_path)
        if (sheetname not in workbook.sheetnames) or (oadlistold != dt_m_list):
            write_excel_xlsx_jccomrow(sta_path, [dt_m_list,readlistdata], sheetname)
            # 分钟冻结，除首次外
        else:
            write_excel_xlsx_jccomrow(sta_path, [readlistdata], sheetname)
    except:
        write_excel_xlsx_jccomrow(sta_path, [dt_m_list, readlistdata], sheetname)
    oadlistold = dt_m_list
    return sta_path,oadlistold

def getnumdeal(tasktestrow,alltestname,getnum):
    if tasktestrow['save'].find("NUM0") >= 0:
        getnum[0] = len(alltestname)
        tasktestrow['result'] = str(getnum[0])
    elif tasktestrow['save'].find("NUM1") >= 0:
        getnum[1] = len(alltestname)
        tasktestrow['result'] = str(getnum[1])
    elif tasktestrow['save'].find("NUM2") >= 0:
        getnum[2] = len(alltestname)
        tasktestrow['result'] = str(getnum[2])
    elif tasktestrow['save'].find("NUM3") >= 0:
        getnum[3] = len(alltestname)
        tasktestrow['result'] = str(getnum[3])
    elif tasktestrow['save'].find("NUM4") >= 0:
        getnum[4] = len(alltestname)
        tasktestrow['result'] = str(getnum[4])
    elif tasktestrow['save'].find("NUM5") >= 0:
        getnum[5] = len(alltestname)
        tasktestrow['result'] = str(getnum[5])
    elif tasktestrow['save'].find("NUM6") >= 0:
        getnum[6] = len(alltestname)
        tasktestrow['result'] = str(getnum[6])
    elif tasktestrow['save'].find("NUM7") >= 0:
        getnum[7] = len(alltestname)
        tasktestrow['result'] = str(getnum[7])
    else:print('增加其它情况')

    return getnum


def totalnumget(tasktestrow,getnum,item,tasklist):
    totaldata = 0
    if tasktestrow['save'].find("TOTALSUCCESS") >= 0:
        for itemdata in range(item-8,item):
            totaldata += int(tasklist[itemdata]['expect'])
        getnum[8] = (getnum[0] + getnum[1] + getnum[2] + getnum[3] + getnum[4] + getnum[5] +
                     getnum[6] + getnum[7]) / (totaldata)
        tasktestrow['real'] = '{:.2%}'.format(getnum[8])
        if tasktestrow['real'] == str(tasktestrow['expect']):
            tasktestrow['result'] = '合格'
        else:
            tasktestrow['result'] = '不合格'
    getnum = [0,0,0,0,0,0,0,0,0]
    logging.info('totaldata'+str(totaldata))
    return getnum



def stabldataread(tasktestrow,dt_m_list,sta_expath,getnum,reportfloderpath,staecandpceventnum,oadlistold,plannumdict):
    #稳定性的特殊处理，全部放在一起，方便查看。
    if tasktestrow['save'].find('EVENTTIMES') >= 0:
        if tasktestrow['save'].find('STA_EC_EVENTTIMES_') >= 0:
            staecandpceventnum[0] = int(tasktestrow['save'][-1])
        elif tasktestrow['save'].find('STA_PC_EVENTTIMES_') >= 0:
            staecandpceventnum[1] = int(tasktestrow['save'][-1])
        else:print('稳定性事件次数统计save列不正确')
        print('staecandpceventnum:',staecandpceventnum)
    elif (tasktestrow['save'].find('STA_POWER') >= 0) and (tasktestrow['result'] == '不合格'):
        staecandpceventnum[2] = 1
    elif (tasktestrow['save'].find('STA_ENERGY') >= 0) and (tasktestrow['result'] == '不合格'):
        staecandpceventnum[3] = 1
    else:
        #将OAD加上注释
        freezetype = ''
        alltestname =[]
        if dt_m_list[0].find('500') >= 0 or dt_m_list[0] == '':
            freezetype = dt_m_list[0]
            del dt_m_list[0]
        else:pass
        if (tasktestrow['real'] != '' and tasktestrow['real'] != "[[]]" and tasktestrow['real'] != '[]') and ((freezetype != '') or (tasktestrow['save'].find('实时数据存储') >= 0)):
            readlist, alltestname = dlreaddata.stabilityread(tasktestrow['real'], freezetype)
        testname = ''
        #统计HPLC抄表成功率使用
        if tasktestrow['save'].find('NUM')>= 0:
            getnum = getnumdeal(tasktestrow,alltestname,getnum)
            return sta_expath,getnum
        else:pass
        if len(alltestname) > 0:
            testname =  dlreaddata.chiname(alltestname[0])
        else:
            testname = '冻结数据召测异常'
        for meani in range(len(dt_m_list)):
            if dt_m_list[meani] in OADTOMEAN.keys():
                dt_m_list[meani] = dt_m_list[meani] + OADTOMEAN[dt_m_list[meani]]
            else:
                pass
        if tasktestrow['save'].find('PLACE_') >= 0:
            dt_m_list.append(f"{'方案编号:' + tasktestrow['expect']}")
        if str(tasktestrow['save']).find('STA_') >= 0:
            nowtime = datetime.datetime.now()
            for sheetname in alltestname:
                sheetname = dlreaddata.selsheetnameget(sheetname,tasktestrow['param'])
                plannumdict, oadlistold = Placefunction.titlenew(oadlistold, sheetname, tasktestrow['expect'],plannumdict)
                # 地方抄表用例，需要在加过展示中增加方案编号。
                for datanum in range(len(readlist)):
                    if sheetname.find(readlist[datanum][0][2:]) >= 0:
                        if freezetype == '50020200':
                            sta_expath[0],oadlistold = freportshow(sta_expath[0], testname, dt_m_list, readlist[datanum],
                                                        sheetname,
                                                        '展示',reportfloderpath,oadlistold)
                        elif freezetype == '50010200':
                            sta_expath[6],oadlistold = freportshow(sta_expath[6], testname, dt_m_list, readlist[datanum],
                                                        sheetname,
                                                        '展示',reportfloderpath,oadlistold)
                        elif freezetype == '50040200':
                            sta_expath[1],oadlistold = freportshow(sta_expath[1], testname, dt_m_list, readlist[datanum],
                                                        sheetname,
                                                        '展示',reportfloderpath,oadlistold)
                        elif freezetype == '50050200':
                            sta_expath[2],oadlistold = freportshow(sta_expath[2], testname, dt_m_list, readlist[datanum],
                                                        sheetname,
                                                        '展示',reportfloderpath,oadlistold)
                        elif freezetype == '50060200':
                            sta_expath[3],oadlistold = freportshow(sta_expath[3], testname, dt_m_list, readlist[datanum],
                                                        sheetname,
                                                        '展示',reportfloderpath,oadlistold)
                        elif freezetype == '':
                            sta_expath[4],oadlistold = freportshow(sta_expath[4], testname, dt_m_list, readlist[datanum],
                                                        sheetname,
                                                        '展示',reportfloderpath,oadlistold)
                        elif freezetype == '50030200':
                            sta_expath[5],oadlistold = freportshow(sta_expath[5], testname, dt_m_list, readlist[datanum],
                                                        sheetname,
                                                        '展示',reportfloderpath,oadlistold)
                        else:print('请新增稳定性冻结数据展示类型')
                    else:pass
        else:pass
    return sta_expath,getnum,staecandpceventnum,oadlistold,plannumdict
#数据标识、路径、上报处理好的real，上报通道
def reportdataread(dt_m_list,sta_expath,report_data,report_way,reportfloderpath):
    #将OAD加上注释
    freezetype = ''
    if dt_m_list[0].find('500') >= 0 or dt_m_list[0] == '':
        freezetype = dt_m_list[0]
        del dt_m_list[0]
    else:pass
    if report_data != '' and report_data != "[[]]" and report_data != '[]' and freezetype != '' :
        readlist, alltestname = dlreaddata.stabilityread(report_data, freezetype)
        testname = ''
        if len(alltestname) > 0:
            testname =  dlreaddata.chiname(alltestname[0])
        else:
            testname = '冻结数据召测异常'
        for meani in range(len(dt_m_list)):
            if dt_m_list[meani] in OADTOMEAN.keys():
                dt_m_list[meani] = dt_m_list[meani] + OADTOMEAN[dt_m_list[meani]]
            else:
                pass
        nowtime = datetime.datetime.now()
        for sheetname in alltestname:
            for datanum in range(len(readlist)):
                if sheetname.find(readlist[datanum][0][2:]) >= 0:
                    #分钟冻结，首次需要第一行标识行
                    if freezetype == '50020200':
                        sta_expath[0] = freportshowre(sta_expath[0], testname, dt_m_list, readlist[datanum], sheetname,
                                                    '上报展示'+report_way,reportfloderpath)
                    # 分钟冻结，首次需要第一行标识行
                    elif freezetype == '50040200':
                        sta_expath[1] = freportshowre(sta_expath[1], testname, dt_m_list, readlist[datanum],
                                                    sheetname,
                                                    '上报展示' + report_way,reportfloderpath)
                    else:print('请新增稳定性冻结数据上报展示类型')
    else:pass
    return sta_expath

def reportEventdataread(eventoad,dt_m_list,sta_expath,report_data,report_way,reportfloderpath):
    #将OAD加上注释
    if dt_m_list[0] == '':
        del dt_m_list[0]
    else:pass
    if report_data != '' and report_data != "[[]]" and report_data != '[]' :
        readlist, alltestname = dlreaddata.stabilityreport(report_data, eventoad)
        testname = ''
        if len(alltestname) > 0:
            testname =  '稳定性事件记录'
        else:
            testname = '事件上报记录召测异常'
        for meani in range(len(dt_m_list)):
            if dt_m_list[meani] in EVENTMENA.keys():
                dt_m_list[meani] = dt_m_list[meani] + EVENTMENA[dt_m_list[meani]]
            else:
                pass
        nowtime = datetime.datetime.now()
        for sheetname in alltestname:
            #事件记录，首次需要第一行标识行
                #当路径为空时，说明该冻结类型的表格还没有生成。
            if (sta_expath[0] == '') or (sta_expath[0][-17:-9] != nowtime.strftime('%Y%m%d%H%M')[:8]):
                sta_expath[0] = reportfloderpath + testname + '上报展示'+ report_way + nowtime.strftime('%Y%m%d%H%M') + ".xlsx"
            else:pass
            # 打开表格，如果sheet也存在，说明该冻结类型的这个地址已经生成过sheet页了，说明第一行备注已经加好。如果打开出现异常，恰好说明表格没有新建好，就是第一次生成某个sheet页，需要备注。
            try:
                workbook = openpyxl.load_workbook(sta_expath[0])
                if sheetname not in workbook.sheetnames:
                    write_excel_xlsx_jccomrow(sta_expath[0],[dt_m_list,readlist],sheetname)
                    # 事件上报记录，除首次外
                else:
                    write_excel_xlsx_jccomrow(sta_expath[0], [readlist], sheetname)
            except:
                write_excel_xlsx_jccomrow(sta_expath[0], [dt_m_list, readlist], sheetname)
    else:pass
    return sta_expath
#百分比的形式比较的流程 预期值和真实值
def percentage(expect,real):
    errorresult=0
    if expect != 0:
        errorresult = float('%.4f' % (abs(expect - real) / abs(expect)))
    else:
        errorresult = real
    return errorresult

def extgetexpectandcompare(extcomrow,ext_min_data_rtu_get,ext_max_data_rtu_get,ext_min_time_rtu_get,ext_max_time_rtu_get):
   #预期值处理流程
    ext_expect_list=extgetexpect(extcomrow,ext_min_data_rtu_get,ext_max_data_rtu_get,ext_min_time_rtu_get,ext_max_time_rtu_get)
   #比较流程
    extcompare(extcomrow, ext_expect_list)
    return True

#百分比的形式比较的流程 expect,real：预期值和真实值
def percentage(expect,real):
    errorresult=0
    if expect != 0:
        errorresult = float('%.4f' % (abs(expect - real) / abs(expect)))
    else:
        errorresult = real
    return errorresult

def aveexpectsave(aveexpectrow,ave_data_rtu):
    if aveexpectrow['save'].find(ave_param) >= 0:
        ext_min_data_rtu = covervalue(aveexpectrow, ave_data_rtu,
                                      [aveexpectrow['oad_omd'], aveexpectrow['real']])
        logging.info('ave_data_rtu' + str(ave_data_rtu))
    else:
        pass
    if aveexpectrow['real'] !='' and aveexpectrow['real'] !=[]:
        aveexpectrow['result']='合格'
    return ave_data_rtu

def avegetexpect(aveexpectgetrow, ave_data_rtu_get):
    sum=0
    ave=0
    comaveoad=''
    avesave=aveexpectgetrow['param'].split('+')
    print('avesave:',avesave)
    for item in avesave:
        for findoad in ave_data_rtu_get:
          if  findoad ['oad/omd']==item.strip('()'):
              sum+=float(findoad ['value'][1])
              comaveoad=findoad ['value'][0]
              break
          else:pass
    ave=float('%.4f' %(sum/len(avesave)))
    logging.info('累加平均预期：'+str([comaveoad,float('%.4f' %sum),ave]))
    return [comaveoad,float('%.4f' %sum),ave]

def avecompare(avecomresultrow, ave_expect_list_result):
    comfalseflag = 0
    avereallist=[]
    aveerror=float(avecomresultrow['expect'].strip('"').strip("'").strip('≤±').strip('%'))/100
    try:
        avereallist = strtolist(avecomresultrow['real'])
    except NameError:
        print("均值统计等待时间需加长")
    logging.info('aveexpect:' + str(ave_expect_list_result))
    logging.info('avereallist:' + str(avereallist))
    if avereallist[0][:-1] != ave_expect_list_result[0][:-1]:
        avecomresultrow['result'] = '不合格'
        logging.error('累加平均统计比较oad有误!')
        return True
    else:
        pass
        # 极值和均值时，因台体零线电流波动，统计结果跳动很大，无法走正常判断流程。只要零线电流小于0.5A 即可判合格。
    if avereallist[0] == '20010400':
        if avereallist[1] < 0.5 and avereallist[2] < 0.5:
            avecomresultrow['result'] = '合格'
        else:
            avecomresultrow['result'] = '不合格'
        return True
    else:
        pass
       # 均值时，因台体电压波动，统计结果跳动很大，无法走正常判断流程。只要20260200小于0.08， 即可判合格。
    if avereallist[0] == '20260200' or avereallist[0] == '20270200':
        if avereallist[1] < 0.09 and avereallist[2] < 0.09:
            avecomresultrow['result'] = '合格'
        else:
            avecomresultrow['result'] = '不合格'
        return True
    else:
        pass
    exttolerancelist = strtolist(avecomresultrow['expect'])
    logging.info('比较OAD为：' + ave_expect_list_result[0])
    for avecomnum in range(1,len(ave_expect_list_result),1):
        aveerrorresult=percentage(ave_expect_list_result[avecomnum],avereallist[avecomnum])
        if aveerrorresult>aveerror:
            comfalseflag = 1
            if avecomnum==1:
                logging.info('累加预期值:' + str(ave_expect_list_result[avecomnum]))
                logging.info('累加真实值:' + str(avereallist[avecomnum]))
                logging.info('累加值比较失败，该项误差为' + str(aveerrorresult)+'*'*20)
            else:
                logging.info('均值预期值:' + str(ave_expect_list_result[avecomnum]))
                logging.info('均值真实值:' + str(avereallist[avecomnum]))
                logging.info('均值比较失败，该项误差为' + str(aveerrorresult)+'*'*20)
        else:
            if avecomnum==1:
                logging.info('累加预期值:' + str(ave_expect_list_result[avecomnum]))
                logging.info('累加真实值:' + str(avereallist[avecomnum]))
                logging.info('累加值比较通过，该项误差为' + str(aveerrorresult))
            else:
                logging.info('均值预期值:' + str(ave_expect_list_result[avecomnum]))
                logging.info('均值真实值:' + str(avereallist[avecomnum]))
                logging.info('均值比较失败通过，该项误差为' + str(aveerrorresult))
    if comfalseflag == 1:
        avecomresultrow['result'] = '不合格'
    else:
        avecomresultrow['result'] = '合格'
    return True






def avegetexpectandcompare(avecomrow,ave_data_rtu):
    ave_expect_list = avegetexpect(avecomrow, ave_data_rtu)
    avecompare(avecomrow, ave_expect_list)
    return True








def hardpopup( hardtestname,rowresult,prepareoflag):
    # 准备工作时，出现不合格或通讯异常，便弹窗提示，并提示不合格项。。
    if ((hardtestname == '状态监测准备工作') or (hardtestname == '首件检验实验工作'))and (prepareoflag == 0):
        problemname = "第" + str(rowresult['nu']) + "行，" + rowresult['name'] + "：测试中出现异常。是否继续执行"
        if win32api.MessageBox(0, problemname, "警告", win32con.MB_YESNO) == 6:
            pass
        else:
            os._exit(0)
    else:
        pass
    return None


# 根据每一行的测试结果来统计最终报告的测试结果
def totalresult(rowresult,unflag,testcasename,receiveouttimeflag,prepareoflag):
    #判断结果是否有异常或不合格
    if (rowresult['result'].find('不合格') >= 0) or (rowresult['result'].find('通讯异常') >= 0)or (rowresult['result'].find('出现异常') >= 0) :
        unflag = 1
    else:
        pass
    #循环测试时用来标记出现通讯异常。
    if (testcasename == '状态监测循环测试') and ((rowresult['result'].find('通讯异常') >= 0)or (rowresult['result'].find('出现异常') >= 0)) :
        receiveouttimeflag = 1
    else:
        pass
    #准备工作时，出现不合格或通讯异常，便弹窗提示，并提示不合格项。。
    if ((testcasename == '状态监测准备工作') or (testcasename == '首件检验实验工作')) and (prepareoflag == 0) :
        if (rowresult['result'].find('不合格') >= 0) or (rowresult['result'].find('异常') >= 0):
            problemname = "第"+str(rowresult['nu']) +"行，" + rowresult['name'] + "：测试中出现异常。是否继续执行"
            if win32api.MessageBox(0, problemname, "警告",win32con.MB_YESNO) == 6:
                pass
            else:
                os._exit(0)
    else:
        pass
    return unflag,receiveouttimeflag

 #首件检用例的准备工作用例中，包含了一部分不需要判断是否合格的用例，那么只需要将该检查的检查完，就可以弹窗提示了。其它EMC实验仍旧保持原流程。
def preparepopup(h_l_temperatureflag):
    if h_l_temperatureflag != 0:
        win32api.MessageBox(0, "准备工作已完成，请进行遥信变位操作后，点击“确定”开始试验。", "提示", win32con.MB_OK)
    else:
        win32api.MessageBox(0, "准备工作已完成，请点击“确定”后开始试验", "提示", win32con.MB_OK)
    return None

#最终用例执行结果及报告输出处理函数
def resultandreport(excelname,pathtempdelet0,unqualitfieldflag0,path_experiment,h_l_temperatureflag,testname,reportfloderpath,appendexresult,prepareoflag):
    now = datetime.datetime.now()
    if testname == '能源稳定性测试' or testname == '交采秘钥自动化测试用例集' :
        path1 = reportfloderpath
    else:
        path1 = os.getcwd() + "\\report\\"
    if excelname.find('状态监测') >= 0:
        mkdir(path1 + f"{path_experiment}\\")
        path1 = path1 + f"{path_experiment}\\"
    else:
        pass
    if excelname.find('出厂参数检验') >= 0:
        excelname = f"{path_experiment}" + excelname
    if unqualitfieldflag0 == 1:
        logging.info(excelname+u'测试不合格'+'\n'+'\n'+'\n'+'\n'+'\n'+'\n''\n'+'\n'+'\n'+'\n')
        path1 += excelname + '-不合格' + now.strftime('%Y-%m-%d %H-%M-%S-%f')[2:16] + '.xlsx'
        # 显示结论到平台
    else:
        logging.warning(excelname+u'测试合格'+'\n'+'\n'+'\n'+'\n'+'\n'+'\n''\n'+'\n'+'\n'+'\n')
        path1 += excelname + '-合格' + now.strftime('%Y-%m-%d %H-%M-%S-%f')[2:16] + '.xlsx'
        # 显示结论到平台
    #如果最终报告生成，就将实时产生的报告删除掉。
    tempexcelname = excelname + '.xlsx'
    #先判断实时生成的文件是否存在，存在了再执行删除。否则会出现异常。
    if os.path.exists(pathtempdelet0 + tempexcelname) == True:
        os.remove(os.path.join(pathtempdelet0 + tempexcelname))
    else:pass
    print(excelname+"测试结束")
    #首件检用例的准备工作用例中，包含了一部分不需要判断是否合格的用例，那么只需要将该检查的检查完，就可以弹窗提示了。其它EMC实验仍旧保持原流程。
    if (excelname.find("状态监测准备工作")>= 0) and (prepareoflag == 0):
        if h_l_temperatureflag != 0:
            win32api.MessageBox(0, "准备工作已完成，请进行遥信变位操作后，点击“确定”开始试验。", "提示", win32con.MB_OK)
        else:
            win32api.MessageBox(0, "准备工作已完成，请点击“确定”后开始试验", "提示", win32con.MB_OK)
    else:pass
    #只要出现不合格，就会将标志位置1
    if path1.find('不合格') >= 0:
        appendexresult = 1
    return path1,appendexresult
#电压合格率
def v_sources_get(sourceparamrow):
    v_m0_all=[]
    v_param = sourceparamrow.split(';')
    for itemsorce in v_param:
        if itemsorce.find('Un=')>=0:
            v_Un=float(itemsorce[3:])
        else:pass
    for itemsorce in v_param:
        if itemsorce.find('Ua=')>=0:
            if itemsorce.find('%')>=0:
                v_a =float(itemsorce[3:-3])*v_Un/100
            else:
                v_a = v_Un
        elif itemsorce.find('Ub=')>=0:
            if itemsorce.find('%')>=0:
                v_b =float(itemsorce[3:-3])*v_Un/100
            else:
                v_b = v_Un
        elif itemsorce.find('Uc=')>=0:
            if itemsorce.find('%')>=0:
                v_c =float(itemsorce[3:-3])*v_Un/100
            else:
                v_c = v_Un
        else:pass
    v_m0_all=[{'oad/omd': 'V_M0_A', 'value': v_a}, {'oad/omd': 'V_M0_B', 'value': v_b}, {'oad/omd': 'V_M0_C', 'value': v_c}]
    logging.info(v_m0_all)
    return v_m0_all

#计量电量计算
def jc_ele_get(jcsourceparamrow):
    ele = 0
    i_f = 0
    v_Un = 0
    i_In = 0
    v_param = jcsourceparamrow['param'].split(';')
    t_jc = float(jcsourceparamrow['delay'])  + 6
    for itemsorce in v_param:
        if itemsorce.find('Un=')>=0:
            v_Un=float(itemsorce[3:])
        else:pass
        if itemsorce.find('In=')>=0:
            i_In=float(itemsorce[3:])
        else:pass
    for itemsorce in v_param:
        if itemsorce.find('Ua=')>=0:
            if itemsorce.find('%')>=0:
                v_a =float(itemsorce[3:-3])*v_Un/100
            else:
                v_a = v_Un
        elif itemsorce.find('Ia=')>=0:
            if itemsorce.find('%')>=0:
                i_a =float(itemsorce[3:-3])*i_In/100
            else:
                i_a = i_In
        elif itemsorce.find('IA1=')>=0:
            i_f=float(itemsorce[4:])
        else:pass
    ele = "%.2f" % (3*(v_Un*i_In)*(cmath.cos(i_f/30*(3.141592653/6)))*(t_jc/3600/1000)).real
    logging.info(f'计量走字电量ele为：{ele}')
    return  ele

#处理 60120300 或者方法筛选结果为空值、需要比的数据，只要有值就判合格，具体值对不对，在比较行体现
def resultdeal(datarow,timeparam,addnullflag):
    #追加空记录关闭时，如果预期值是'NULL',则强制修改为'[]'。又新增了一个逻辑：四川专变（或其它地区），分钟冻结和小时冻结，无论
    #回空还是回“带NULL”的，都判合格。
    exold = datarow['expect']
    if (addnullflag == 1) and (datarow['expect'] == 'NULL') and (datarow['save'].find('NO_JUDGE') < 0) :
        datarow['expect'] = '[]'
    else:pass
    if ((datarow['real'] != '') and (datarow['real'] != str([]))) or (datarow['save'].find('NO_JUDGE') >= 0):
        if datarow['expect'] == 'NULL' :
            if (datarow['real'] != str([])):
                datarow['result'] = nulldata(datarow['real'])
            if datarow['result'] == True:
                datarow['result'] = u'合格'
            elif (datarow['save'].find('NO_JUDGE') >= 0) and datarow['real'] == str([])  :
                datarow['result'] = u'合格'
            else:
                datarow['result'] = u'不合格'
        elif datarow['expect'] == '[]':
            datarow['result'] = u'不合格'
            #稳定性终端数据筛选时，用例计算特殊的预期值
        elif datarow['save'].find('RTU_SELECT') >= 0:
            expectlist = datarow['expect'].split(';')
            expect2021 = timeparam[0][0][2]
            expectlist[0] = expect2021
            readdatalist = datarow['real'].split(';')
            readdatalist[0] = readdatalist[0].strip("'").strip('"')
            datarow['result'] = u'合格'
            datarow['expect'] = str(expectlist)
            if readdatalist[0] == expectlist[0]:
                print('datarow[real]:',readdatalist)
                print('expectlist:',expectlist)
            else:
                print('datarow[real]:', readdatalist)
                print('expectlist:', expectlist)
                datarow['result'] = u'不合格'
            if dlreaddata.selc_err(readdatalist) == False:
                datarow['result'] = u'不合格'
            else:pass
        else:
            datarow['result'] = u'合格'
    elif ((datarow['real'] == '') or datarow['real'] == str([])) and datarow['expect'] == '[]':
        datarow['result'] = u'合格'
        # 电压合格率回空值的判断流程
    elif datarow['expect'].find('M=') >= 0:
        if datarow['expect'].find('M=0') >= 0:
            if datarow['real'] == str([]):
                datarow['result'] = u'合格'
            else:
                pass
        else:
            resultfailflag = 0
            for resultitem in strtolist(datarow['real']):
                if resultitem != 'NULL':
                    resultfailflag = 1
                else:
                    pass
            if resultfailflag == 0:
                datarow['result'] = u'合格'
    else:
        datarow['result'] = u'不合格'
    datarow['expect'] = exold
    return None

#定义时间处理函数，知道起始时间和时间间隔，来进行对时。例如，起始时间为某日0点，时间间隔为15分钟，则下次对时时间为15分，下下次为30分等。
def timeparamdeal(timestart,timeintervel,timeparamrow,paramold,rtu_current_time):
    paramold = timeparamrow['param']
    if timeparamrow['save'].find('TIMESPECIAL') >= 0:
        Placefunction.timeparamsp(timeparamrow,rtu_current_time)
    else:
        if timeparamrow['param'].find('STARTTIME') >= 0 and timeparamrow['param'].find('STARTTIME') >= 0:
            numre = re.compile(r'\d+')
            paramlist = timeparamrow['param'].split(',')
            intervelre = numre.search(paramlist[1])
            # 时间间隔
            timeintervel = int(intervelre.group())
            paramori = numre.search(paramlist[0])
            #累加计算得到新的对时时间
            if paramori == None:
                dt = bcdtodatetime(timestart)
                delta = datetime.timedelta(minutes=0)
                if paramlist[1].find("分钟") >= 0 :
                    delta = datetime.timedelta(minutes=timeintervel)
                    timestart = (str(dt + delta).replace('-', '').replace(" ", '').replace(":", ''))
                elif paramlist[1].find("日") >= 0 :
                    delta = datetime.timedelta(days=timeintervel)
                    timestart = (str(dt + delta).replace('-', '').replace(" ", '').replace(":", ''))
                elif paramlist[1].find("小时") >= 0 :
                    delta = datetime.timedelta(hours=timeintervel)
                    timestart = (str(dt + delta).replace('-', '').replace(" ", '').replace(":", ''))
                elif paramlist[1].find("月") >= 0 :
                    nowmonthtime = bulidarrowobject(timestart)
                    #先算出下2个月，然后推到下2个月的第一天，再减去一天，可以得到下月最后一天。
                    lasttwomonth = nowmonthtime.shift(months=timeintervel+1).format("YYYYMMDDHHmmss")
                    lastmonthfirstday = lasttwomonth[:6] + "01" + lasttwomonth[-6:]
                    dt = bcdtodatetime(lastmonthfirstday)
                    delta = datetime.timedelta(days=1)
                    timestart = (str(dt - delta).replace('-', '').replace(" ", '').replace(":", ''))
                else:print('可继续扩展为其它时间间隔')

            #首次计算时间
            else:
                timestart = paramori.group()
            timeparamrow['param'] = timestart
            print('timestart, timeintervel:',timestart, timeintervel)
        if (timeparamrow['save'].find('STARTTIME') >= 0)  :
            timestart = paramold
    return timestart, timeintervel,timeparamrow['param'],paramold


def agentdata(dt_m_list,dt_list_realData,ldatarow):
    oadlist = []
    valuelist = []
    if dt_list_realData[0].find("超时") >= 0:
        ldatarow['result']  = "不合格"
        print("485抄表异常")
    else:
        data = strtolist(dt_list_realData[0])[1:]
        print('data:',data)
        for item in data:
            oadlist.append(item[0].split(":")[0])
            valuelist.append(item[0].split(":")[1])
    return oadlist,valuelist

def split_list_by_rich_word(word_list, rich_word):
    """根据rich_word将文本列表进一步拆分"""
    rt = []
    for word in word_list:
        if rich_word in word:
            tmp = word.split(rich_word)
            rt_tmp = []
            flag = False
            for t in tmp:
                if not flag:
                    flag = True
                    rt_tmp.append(t)
                else:
                    rt_tmp.append(rich_word)
                    rt_tmp.append(t)

            rt += rt_tmp
        else:
            rt.append(word)

    # print('word_list:', word_list)
    return rt


def str_2_rich_string_content(s, rich_word_list, color_type):
    """构建字体颜色"""
    wl = [s]
    # 是否包含rich_word_list
    flag = False
    for r_w in rich_word_list:
        if r_w in s:
            flag = True
            wl = split_list_by_rich_word(wl, r_w)
    if not flag:
        return False, []
    # 添加颜色
    for i in range(len(wl) - 1, -1, -1):
        if wl[i] in rich_word_list:
            wl.insert(i, color_type)
    if '' in wl:
        wl.remove('')
    return True, wl

#获取一个列表中绝对值最大的项
def maxabsgetlist(listx,oldmax):
    maxlist1 = oldmax
    for i in range(len(listx)):
        if abs(listx[i]) > abs(oldmax):
            maxlist1 = listx[i]
        else:pass
    return maxlist1

def maxabsget(list1,list2,err_var_max):
    if err_var_max[0] == '不执行':
        err_var_max = [0,0]
    else:pass
    err_var_max[0] = maxabsgetlist(list1,err_var_max[0])
    err_var_max[1] = maxabsgetlist(list2,err_var_max[1])
    return err_var_max

def txtwrite(pathid,idnum):
    f = open(pathid, 'w')
    f.write(idnum )
    f.close()

def kwtow(dt_m0_list_data,dt_list_realData):
    dt_m0_list_data[6] = float(dt_m0_list_data[6]) * 1000
    dt_m0_list_data[7] = float(dt_m0_list_data[7]) * 1000
    dt_m0_list_data[8] = float(dt_m0_list_data[8]) * 1000
    dt_list_realData[6] = float(dt_list_realData[6]) * 1000
    dt_list_realData[7] = float(dt_list_realData[7]) * 1000
    dt_list_realData[8] = float(dt_list_realData[8]) * 1000
    print(dt_m0_list_data,':dt_m0_list_data')
    print(dt_list_realData, ':dt_list_realData')

    return dt_m0_list_data,dt_list_realData


#BEFORE_TEST_DATA
def compareagentdata(dt_m0_list_data, dt_list_realData, tasklistagentrow,dt_m_list,error_hardwaresave,hardjcq,hardjceachtimeflag,err_var_max,h_l_temperatureflag,testtime,rtuaddr):
    jcdatapath = os.getcwd() + "\\jc_error_record\\" + rtuaddr + '交采常温误差值记录.txt'
    # 实验中数据误差
    error_hardwaretest = []
    #实验前和实验中的变化量
    variation = []
    #存储最终的结果，带描述的
    realdata = []
    error = 0
    describerrowlist = ['Ua误差:', 'Ub误差:', 'Uc误差:', 'Ia误差:', 'Ib误差:', 'Ic误差:', 'Pa误差:', 'Pb误差:', 'Pc误差:']
    describvalist = ['Ua变化量:', 'Ub变化量:', 'Uc变化量:', 'Ia变化量:', 'Ib变化量:', 'Ic变化量:', 'Pa变化量:', 'Pb变化量:', 'Pc变化量:']
    #电压电流功率的额定值
    hardwarerating = [220,220,220,1.5,1.5,1.5,330,330,330]
    expecterror = float(tasklistagentrow['expect'].strip('≤±').strip('%')) / 100
    del dt_m0_list_data[6]
    #硬件测试要求功率的单位为W，所以在此处进行转换
    dt_m0_list_data,dt_list_realData = kwtow(dt_m0_list_data,dt_list_realData)
    #分为实验前存好数据和实验中读取数据。（写入txt，从txt文本中读取）
    if ((tasklistagentrow['save'].find('BEFORE_TEST_DATA') >= 0) or (h_l_temperatureflag == 1)) and (h_l_temperatureflag != 2) :
        error_hardwaresave = []
        print("存储测试开始前的误差值")
        for i in range(len(dt_m0_list_data)):
            error_hardwaresave.append(float('%.4f' % ((float(dt_m0_list_data[i]) - float(dt_list_realData[i]))/hardwarerating[i])))
            realdata.append(describerrowlist[i] + str(error_hardwaresave[i]))
            if abs(error_hardwaresave[i]) >= expecterror:
                error = 1
                realdata[i] = realdata[i] +"异常"
            else:pass
            #硬件自动化展示结果用
        realdata = str(realdata) +'!' + str(dt_m0_list_data)+'!'+ str(dt_list_realData)+'!'+ str(error_hardwaresave)
        # if h_l_temperatureflag == 1 and testtime == 1:
        #先改成误差每次都存。覆盖前面的结果。(常温下才会存)
        if h_l_temperatureflag == 1 :
             txtwrite(jcdatapath,str(error_hardwaresave))
        else:pass
    else:
        if  h_l_temperatureflag == 2:
            f = open(jcdatapath, "r")  # 设置文件对象
            error_hardwaresave = strtolist(f.read() ) # 直接将文件中按行读到list里，效果与方法2一样
            print(f'error_hardwaresave:{error_hardwaresave}')
            f.close()
        else:pass
        variationexcel = []
        for i in range(len(dt_m0_list_data)):
            error_hardwaretest.append(float('%.4f' % ((float(dt_m0_list_data[i]) - float(dt_list_realData[i]))/hardwarerating[i])))
            variation.append(float('%.4f' %(error_hardwaretest[i]-error_hardwaresave[i])))
            realdata.append(describvalist[i] + str(variation[i]))
            if abs(error_hardwaretest[i]) >= expecterror:
                error = 1
                realdata[i] = realdata[i] +" 误差异常；"
            else:pass
            if abs(variation[i]) >= expecterror:
                error = 1
                realdata[i] = realdata[i] +" 变化量异常"
            else:pass
        err_var_max = maxabsget(error_hardwaretest,variation,err_var_max)
        realdata = str(realdata) + '!' + str(dt_m0_list_data) + '!' + str(dt_list_realData) + '!' + str(error_hardwaretest)+ '!' + str(variation)
    tasklistagentrow['real'] = str(realdata)

    if error == 1:
        tasklistagentrow['result'] = "不合格"
        hardjcq = 1
        hardjceachtimeflag = 1
    else:
        hardjceachtimeflag = 0
        tasklistagentrow['result'] = "合格"
    return error_hardwaresave,hardjcq,hardjceachtimeflag,err_var_max


#如果电压电流异常，就停止硬件测试。
def rateduicheck(Rtu698datalistSum):
    uwrongflag = 0
    iwrongflag = 0
    # 判断源是否正常标准
    rated_u = ['200', '200', '200']
    rated_i = [ '1', '1', '1']

    datau= Rtu698datalistSum[0]['value'][:3]
    datai = Rtu698datalistSum[0]['value'][3:6]
    for i in range(len(rated_u)):
        if abs(float(datau[i])) < float(rated_u[i]):
            uwrongflag = 1
            break
        else:
            pass
    for j in range(len(rated_i)):
        if abs(float(datai[j])) < float(rated_i[j]):
            iwrongflag = 1
            break
        else:
            pass
    if uwrongflag == 1 and iwrongflag == 1:
        # ##提醒OK消息框
        if win32api.MessageBox(0, "电压、电流均异常，是否继续执行程序？", "提示", win32con.MB_YESNO) == 6:
            pass
        else:
            os._exit(0)
    elif uwrongflag == 1:
        # ##提醒OK消息框
        if win32api.MessageBox(0, "电压异常，是否继续执行程序？", "提示", win32con.MB_YESNO) == 6:
            pass
        else:
            os._exit(0)
    elif iwrongflag == 1:
        # ##提醒OK消息框
        if win32api.MessageBox(0, "电流异常，是否继续执行程序？", "提示", win32con.MB_YESNO) == 6:
            pass
        else:
            os._exit(0)
    else:pass

    return None

#将  硬件自动化结果展示模板  从config中复制到 report中(新路径中)
def excelcopypath(hardwarereportname,oldname) :
    path = os.getcwd()
    if hardwarereportname.find('合格') >= 0:
        source = path + r"\report" + '/' + oldname
    else:
        source = path + r"\config" + '/' + oldname
    target = path + r"\report" + hardwarereportname
    try:
        copyfile(source, target)
    except:
        print('copy file faild')
    return target

#将  测试报告重命名
def excelcopy(hardwarereportname,oldname,path_experiment) :
    path = os.getcwd()
    if hardwarereportname.find('合格') >= 0:
        source = path + r"\report" + '/' +  f"{path_experiment}\\" + oldname
    else:
        source = path + r"\config" + '/' + oldname
    target = path +r"\report" + '/' +  f"{path_experiment}\\" + hardwarereportname
    # try:
    copyfile(source, target)
    # except:
    #     print('copy file faild')
    return target

#将 自动化测试报告模板  从config中复制到 report中
def excelcopyconfig(hardwarereportname,oldname) :
    path = os.getcwd()
    if hardwarereportname.find('合格') >= 0:
        source = path + r"\report" + '/' + oldname
    else:
        source = path + r"\config" + '/' + oldname
    target = path + r"\report" + hardwarereportname
    try:
        copyfile(source, target)
    except:
        print('copy file faild')
    return target
# 输入目录路径，输出最新文件完整路径
def find_new_file(dir):
    '''查找目录下最新的文件'''
    file_lists = os.listdir(dir)
    file_lists.sort(key=lambda fn: os.path.getmtime(dir + "\\" + fn)
    if not os.path.isdir(dir + "\\" + fn) else 0)
    print('最新的文件为： ' + file_lists[-1])
    file = os.path.join(dir, file_lists[-1])
    print('完整路径：', file)
    return file


def logcopy(sourcepath,targetpath):
    dir = sourcepath
    source = find_new_file(dir)
    try:
       shutil.copy(source, targetpath)
    except IOError as e:
       print("Unable to copy file. %s" % e)
    except:
       print("Unexpected error:", sys.exc_info())
    return None

def highlowreportdeal(rtuaddr,newdst):
    highlowpath = os.getcwd() + r"\report\\" + rtuaddr + r'高低温\\'
    mkdir(highlowpath)
    source = newdst + r'\\'
    dest1 = highlowpath
    files = os.listdir(source)
    for f in files:
        shutil.move(source + f, dest1)
    shutil.rmtree(source)
    return None



def hardwarefinalmerge(path_experiment,hardexhibitionname,h_l_temperatureflag,rtuaddr):
    # 测试结束时，对整个文件夹中的文件进行合并处理。
    ultimatename = ''
    excelpath = ''
    #标记测试结果
    resultright = 0
    ultimatename = excelcopy((hardexhibitionname.split(".")[0] + '报告汇总' + '.xlsx'), '硬件自动化测试报告汇总模板.xlsx', path_experiment)
    # ultimatename = ultimatename.split(".")[0] + '报告汇总'+'.xlsx'
    excelpath = os.getcwd() + r"\report\\" + f'{path_experiment}'
    resultright = Exceldeal.hardwareexcelmerge(excelpath, ultimatename,resultright)
    oldsrc = excelpath
    #将测试报告对应的log也放到一个文件夹中
    sourcepath = os.getcwd() +'\log'
    targetpath = excelpath + '\\' +  f'{path_experiment}' + '.log'
    logcopy(sourcepath, targetpath)
    #存好log后再重命名文件夹，否则会找不到文件夹。
    if resultright == 0:
        newdst = excelpath+'合格'
    else:
        newdst = excelpath + '不合格'
    os.rename(oldsrc, newdst)
    #高低温测试时，需要将常温、高温、低温时的报告放在以终端名称命名（000000000001高低温）的文件夹中。
    if h_l_temperatureflag != 0:
        highlowreportdeal(rtuaddr,newdst)
    return None

#处理一下不执行的项，结果中不展示
def reportoptimize(columnlist1, columnlist2):
    indexlist1 = 0
    #交采
    if columnlist1[1].find('不执行') >= 0 :
        del columnlist1[1]
        del columnlist2[1]
        indexlist1 += 1
    #遥信
    if columnlist1[2 - indexlist1] == '':
        del columnlist1[2 - indexlist1]
        del columnlist2[2- indexlist1]
        indexlist1 += 1
    #控制
    if columnlist1[3 - indexlist1] == '':
        del columnlist1[3- indexlist1]
        del columnlist2[3- indexlist1]
        indexlist1 += 1
    #回路巡检
    if columnlist1[4 - indexlist1] == '':
        del columnlist1[4 - indexlist1]
        del columnlist2[4- indexlist1]
        indexlist1 += 1
    # 门节点
    if columnlist1[5 - indexlist1] == '':
        del columnlist1[5 - indexlist1]
        del columnlist2[5 - indexlist1]
        indexlist1 += 1
    #脉冲
    if columnlist1[6 - indexlist1].find('不执行') >= 0:
        del columnlist1[6 - indexlist1]
        del columnlist2[6- indexlist1]
        indexlist1 += 1
    #HPLC:A相
    if columnlist1[7 - indexlist1] == '':
        del columnlist1[7 - indexlist1]
        del columnlist2[7- indexlist1]
        indexlist1 += 1
    # HPLC B相
    if columnlist1[8 - indexlist1] == '':
        del columnlist1[8 - indexlist1]
        del columnlist2[8 - indexlist1]
        indexlist1 += 1
    # HPLC C相
    if columnlist1[9 - indexlist1] == '':
        del columnlist1[9 - indexlist1]
        del columnlist2[9 - indexlist1]
        indexlist1 += 1
    # 485#1
    if columnlist1[10 - indexlist1] == '':
        del columnlist1[10 - indexlist1]
        del columnlist2[10 - indexlist1]
        indexlist1 += 1
    # 485#2
    if columnlist1[11 - indexlist1] == '':
        del columnlist1[11 - indexlist1]
        del columnlist2[11 - indexlist1]
        indexlist1 += 1
    #485#3
    if columnlist1[12 - indexlist1] == '':
        del columnlist1[12 - indexlist1]
        del columnlist2[12 - indexlist1]
        indexlist1 += 1
    # 485#4
    if columnlist1[13 - indexlist1] == '':
        del columnlist1[13 - indexlist1]
        del columnlist2[13 - indexlist1]
        indexlist1 += 1
    # CAN口
    if columnlist1[14 - indexlist1] == '':
        del columnlist1[14 - indexlist1]
        del columnlist2[14 - indexlist1]
        indexlist1 += 1
    # 蓝牙名称和MAC
    if columnlist1[15 - indexlist1] == '不执行':
        del columnlist1[15 - indexlist1]
        del columnlist2[15 - indexlist1]
        indexlist1 += 1
    # 4G心跳和登录情况
    if columnlist1[16 - indexlist1] == '不执行':
        del columnlist1[16 - indexlist1]
        del columnlist2[16 - indexlist1]
        indexlist1 += 1
    #长帧
    if columnlist1[17 - indexlist1] == '':
        del columnlist1[17 - indexlist1]
        del columnlist2[17 - indexlist1]
        indexlist1 += 1
    # 扩展模块是否掉线
    if columnlist1[18 - indexlist1] == '':
        del columnlist1[18 - indexlist1]
        del columnlist2[18 - indexlist1]
        indexlist1 += 1
    return columnlist1, columnlist2

#定义处理硬件自动化结果函数
def hardwconclusion( status,conclusionflag,columnlist1,columnlist2):
    pulsestatus = ''
    testtimes = status[7]
    hplctesttimes = testtimes *3
    hplc_afalsetimes = status[6]
    if status[0] == 1:
        status[0] = "交采误差或变化量有异常，请查看:误差最大值为：" + str(status[9][0]) + "；变化最大值为：" +str(status[9][1])
        conclusionflag = 1
    else:
        status[0] = "交采误差或变化量无异常。误差最大值为：" + str(status[9][0]) + "；变化最大值为：" +str(status[9][1])
    #遥信判断流程
    if status[1][0] == 1:
        status[1][0] = '实验中，“遥信”状态有异常：'
        conclusionflag = 1
    elif status[1][0] == 0:
        status[1][0] =  '实验中，“遥信”状态无异常'
    else:
        status[1][0] =''

    # 控制判断流程
    if status[1][1] == 1:
        status[1][1] = '实验中，“控制”状态有异常：'
        conclusionflag = 1
    elif status[1][1] == 0:
        status[1][1] = '实验中，“控制”状态无异常'
    else:
        status[1][1] = ''
    # 回路巡检判断流程
    if status[1][2] == 1:
        status[1][2] = '实验中，“回路巡检”状态有异常：'
        conclusionflag = 1
    elif status[1][2] == 0:
        status[1][2] = '实验中，“回路巡检”状态无异常'
    else:
        status[1][2] = ''
    #门节点判断流程
    if status[1][3] == 1:
        status[1][3] = '实验中，“门节点”状态有异常：'
        conclusionflag = 1
    elif status[1][3] == 0:
        status[1][3] =  '实验中，“门节点”状态无异常'
    else:
        status[1][3] =''
    #脉冲状态判断流程(如果是公变，不测试脉冲，就不进行处理)
    if status[5] == []:
        status[5].append("脉冲不执行或出现异常")
        status[5].append("脉冲不执行或出现异常")
        pass
    else:
        for i in range(len(status[5])):
            if int(status[5][i].split(":")[0]) == 1:
                conclusionflag = 1
                pulsestatus += "第"+str(i+1)+f"脉冲有异常，实际脉冲数量为{status[5][i].split(':')[1]}；"
            else:
                pulsestatus += "第" + str(i+1) + f"脉冲无异常，实际脉冲数量为{status[5][i].split(':')[1]}；"
        if pulsestatus.find("有异常") >= 0:
            conclusionflag = 1
            status[5][0] = "实验后脉冲数量有异常"
        else:
            status[5][0] = "实验后脉冲数量无异常"
        try:
            status[5][1] = pulsestatus[:-1]
        except:
            status[5].append(pulsestatus[:-1])
    #HPLC结果输出:循环一次，HPLC点抄3次，所以点抄总次数是循环次数的3倍。
    # 重新赋值status[6],用来存储HPLC点抄合格率
    status[6] = [[], [], []]
    #HPLC：A相点抄成功率统计
    if status[8][4] == 1:
        if hplc_afalsetimes /hplctesttimes > 0:
            conclusionflag = 1
            status[6][0] = ["A相HPLC点抄次数为：" + str(hplctesttimes)+"；HPLC点抄失败次数为："+ str(hplc_afalsetimes) + "；有异常","点抄成功率"+'{:.2%}'.format(1 - hplc_afalsetimes/hplctesttimes)]
        else:
            status[6][0] = ["A相HPLC点抄次数为：" + str(hplctesttimes) + "；HPLC点抄失败次数为：" + str(hplc_afalsetimes) + "；无异常",
                         "点抄成功率" + '{:.2%}'.format(1 - hplc_afalsetimes/hplctesttimes)]
    else:
        status[6][0]=['','']
    # HPLC：B相点抄成功率统计
    if status[8][9] == 1:
        if status[8][11] / hplctesttimes > 0:
            conclusionflag = 1
            status[6][1] = ["B相HPLC点抄次数为：" + str(hplctesttimes) + "；HPLC点抄失败次数为：" + str(status[8][11]) + "；有异常",
                            "点抄成功率" + '{:.2%}'.format(1 - status[8][11] / hplctesttimes)]
        else:
            status[6][1] = ["B相HPLC点抄次数为：" + str(hplctesttimes) + "；HPLC点抄失败次数为：" + str(status[8][11]) + "；无异常",
                            "点抄成功率" + '{:.2%}'.format(1 - status[8][11] / hplctesttimes)]
    else:
        status[6][1] = ['', '']
    # HPLC：C相点抄成功率统计
    if status[8][10] == 1:
        if status[8][12] / hplctesttimes > 0:
            conclusionflag = 1
            status[6][2] = ["C相HPLC点抄次数为：" + str(hplctesttimes) + "；HPLC点抄失败次数为：" + str(status[8][12]) + "；有异常",
                            "点抄成功率" + '{:.2%}'.format(1 - status[8][12] / hplctesttimes)]
        else:
            status[6][2] = ["C相HPLC点抄次数为：" + str(hplctesttimes) + "；HPLC点抄失败次数为：" + str(status[8][12]) + "；无异常",
                            "点抄成功率" + '{:.2%}'.format(1 - status[8][12] / hplctesttimes)]
    else:
        status[6][2] = ['', '']

    #RS485抄表结果输出
    #RS485 #1 抄表结果输出
    #重新赋值status[7],用来存储485抄表合格率、扩展一个CAN口
    status[7] = [[],[],[],[],[]]
    if status[8][5] == 1:
        if status[8][0] / testtimes > 0:
            conclusionflag = 1
            status[7][0] = ["RS485#1点抄次数为：" + str(testtimes) + "；RS485#1点抄失败次数为：" + str(status[8][0]) + "；有异常",
                         "点抄成功率" + '{:.2%}'.format(1 - status[8][0] / testtimes)]
        else:
            status[7][0] = ["RS485#1点抄次数为：" + str(testtimes) + "；RS485#1点抄失败次数为：" + str(status[8][0]) + "；无异常",
                         "点抄成功率" + '{:.2%}'.format(1 - status[8][0] / testtimes)]
            # RS485 #2 抄表结果输出
    else:
        status[7][0]=['','']
    if status[8][6] == 1:
        if status[8][1] / testtimes > 0:
            conclusionflag = 1
            status[7][1] = ["RS485#2点抄次数为：" + str(testtimes) + "；RS485#2点抄失败次数为：" + str(status[8][1]) + "；有异常",
                            "点抄成功率" + '{:.2%}'.format(1 - status[8][1] / testtimes)]
        else:
            status[7][1] = ["RS485#2点抄次数为：" + str(testtimes) + "；RS485#2点抄失败次数为：" + str(status[8][1]) + "；无异常",
                            "点抄成功率" + '{:.2%}'.format(1 - status[8][1] / testtimes)]
            # RS485 #3 抄表结果输出
    else:
        status[7][1] = ['', '']
    if status[8][7] == 1:
        if status[8][2] / testtimes > 0:
            conclusionflag = 1
            status[7][2] = ["RS485#3点抄次数为：" + str(testtimes) + "；RS485#3点抄失败次数为：" + str(status[8][2]) + "；有异常",
                            "点抄成功率" + '{:.2%}'.format(1 - status[8][2] / testtimes)]
        else:
            status[7][2] = ["RS485#3点抄次数为：" + str(testtimes) + "；RS485#3点抄失败次数为：" + str(status[8][2]) + "；无异常",
                            "点抄成功率" + '{:.2%}'.format(1 - status[8][2] / testtimes)]
            # RS485 #4 抄表结果输出
    else:
        status[7][2] = ['', '']
    if status[8][8] == 1:
        if status[8][3] / testtimes > 0:
            conclusionflag = 1
            status[7][3] = ["RS485#4点抄次数为：" + str(testtimes) + "；RS485#4点抄失败次数为：" + str(status[8][3]) + "；有异常",
                            "点抄成功率" + '{:.2%}'.format(1 - status[8][3] / testtimes)]
        else:
            status[7][3] = ["RS485#4点抄次数为：" + str(testtimes) + "；RS485#4点抄失败次数为：" + str(status[8][3]) + "；无异常",
                            "点抄成功率" + '{:.2%}'.format(1 - status[8][3] / testtimes)]
    else:
        status[7][3] = ['', '']
    if status[8][13] == 1:
        if status[8][14] / testtimes > 0:
            conclusionflag = 1
            status[7][4] = ["CAN口点抄次数为：" + str(testtimes) + "；CAN口点抄失败次数为：" + str(status[8][14]) + "；有异常",
                            "点抄成功率" + '{:.2%}'.format(1 - status[8][14] / testtimes)]
        else:
            status[7][4] = ["CAN口点抄次数为：" + str(testtimes) + "；CAN口点抄失败次数为：" + str(status[8][14]) + "；无异常",
                            "点抄成功率" + '{:.2%}'.format(1 - status[8][14] / testtimes)]
    else:
        status[7][4] = ['', '']
    if status[18] == 1:
        status[18] = "长帧测试有异常"
        conclusionflag = 1
    elif status[18] == 0:
        status[18] = "长帧测试合格"
    else:
        status[18] = ""
    if status[19] == '':
        status[19] = ''
    elif status[19] > 0:
        status[19] = "扩展模块有异常->掉线次数为:" + str(status[19])
        conclusionflag = 1
    else:
        status[19] = "扩展模块未出现掉线"
    #不执行的状态、抄表就不在最终的报告中显示，这里要做处理
    columnlist1 = ["实验结论:",status[0],status[1][0],status[1][1],status[1][2],status[1][3],status[5][0],status[6][0][0],status[6][1][0],status[6][2][0],status[7][0][0],status[7][1][0],status[7][2][0],status[7][3][0],status[7][4][0],status[14][0],status[15][0],status[18],status[19]]
    columnlist2 = ['','','"遥信"实验前状态:'+str(status[2]),'"控制"实验前状态:'+str(status[3]),'"回路巡检"实验前状态:'+str(status[4]),'"门节点"实验前状态:'+str(status[20]),str(status[5][1]),status[6][0][1],status[6][1][1],status[6][2][1],status[7][0][1],status[7][1][1],status[7][2][1],status[7][3][1],status[7][4][1],status[14][1],status[15][1],'','']
    columnlist1, columnlist2 = reportoptimize(columnlist1, columnlist2)
    return status,conclusionflag,columnlist1,columnlist2

def hardwarereport(path,conclusionflag,hardexhibitionname,reportflag,path_experiment,rxoutflag):
    name = ''
    if (conclusionflag == 0) and (reportflag == 0) and (rxoutflag == 0):
        name = '合格'
    else:
        name = '不合格'
    if hardexhibitionname.find("实验") >= 0:
        hardexhibitionnamenew = hardexhibitionname.split("实验")[0] + name + hardexhibitionname.split("实验")[1]
    else:
        hardexhibitionnamenew = hardexhibitionname.split('.xlsx')[0]+name+'.xlsx'
    excelcopy(hardexhibitionnamenew, hardexhibitionname,path_experiment)
    os.remove(path)
    return None
#
# def renamecirtest(oldreportpath,path_experiment):
#     if oldreportpath.find('状态监测循环测试')>=0:
#         excelcopy(oldreportpath.split('合格')[0] + "不合格",oldreportpath,path_experiment)
#     else:pass

def file_name(file_dir):
    for root, dirs, files in os.walk(file_dir):
        # print(root)  # 当前目录路径
        # print(dirs)  # 当前路径下所有子目录
        print(files)  # 当前路径下所有非目录子文件
        return files

def cirrename(path_experiment):
    timeend = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    allreportname = file_name(os.getcwd() + r"\report\\" + f'{path_experiment}')
    for item in allreportname:
        if item.find('状态监测循环测试合格')>= 0 :
            os.rename(os.getcwd() + r"\report\\" + f'{path_experiment}\\'+item, os.getcwd() + r"\report\\" + f'{path_experiment}\\'+ item.split('合格')[0] + '不合格'+ timeend +'.xlsx')
        else:pass
    return None





def hardwareresultshow(datashow,oldreportpath,status,hardexhibitionname,path_experiment):
    #现将模板文件从config中复制到report中。且只有在重新执行"状态监测准备工作"时，才会进行COPY
    #硬件自动化结果展示最终路径和名字
    mkdir(os.getcwd() + r"\report\\" + f'{path_experiment}\\')
    path = os.getcwd() + r"\report\\" + f'{path_experiment}' + hardexhibitionname
    if oldreportpath.find("状态监测准备工作") >= 0:
        hardwaresavepath = excelcopy(hardexhibitionname,'硬件自动化结果展示模板.xlsx',path_experiment)
    else:pass
    #展示交采结果后，输出测试结论。
    if oldreportpath.find("状态监测结束工作") >= 0:
        #最终报告中的第1列数据
        columnlist1 = []
        # 最终报告中的第2列数据
        columnlist2 = []
        conclusionflag = 0
        if os.path.exists(path) == True:
            pass
        else:
            hardwaresavepath = excelcopy(hardexhibitionname, '硬件自动化结果展示模板.xlsx',path_experiment)
        status,conclusionflag,columnlist1,columnlist2 = hardwconclusion( status,conclusionflag,columnlist1,columnlist2)
        if conclusionflag == 1 or status[13] == 1 :
            cirrename(path_experiment)
        else:pass
        write_excel_xlsx_hardware(path, [columnlist1,columnlist2], HARDWAREROWLIST,"测试结论")
        write_excel_xlsx_hardware(path, [status[10], status[11]], [], '4G登录记录')
        #EMC和首件检模板不同（硬件自动化结果展示模板），这样写程序，可以由程序通过模板自动识别是否需要存4G登录记录。比较方便 20210922
        try:
            write_excel_xlsx_hardware(path, [status[16], status[17]], [], '以太网登录记录')
        except:
            logging.info("不是首件检，不关注以太网登录情况")
        hardwarereport(path,conclusionflag,hardexhibitionname,status[12],path_experiment,status[13])
    else:
        hardwarerating = [220, 220, 220, 1.5, 1.5, 1.5, 330, 330, 330]
        snow = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:23]
        timelist = [snow]
        datalist = datashow.split("!")[1:]
        rtudatalist = strtolist(datalist[0])
        meterdatalist = strtolist(datalist[1])
        errorlist = strtolist(datalist[2])
        #实验中（循环测试时）才会有变化量，故要做处理
        if ((oldreportpath.find("状态监测循环测试") >= 0) or (oldreportpath.find("_高温") >= 0)or (oldreportpath.find("_低温") >= 0)) and (oldreportpath.find("_常温") <= 0):
            valist = strtolist(datalist[3])
        else:
            valist = []
        # hardwareshowexcel = readexcelappid(path,'Sheet1',HARDWAREROWLIST)
        if oldreportpath.find("状态监测循环测试") >= 0:
            write_excel_xlsx_hardware(path, [['实验中','额定值'], ['','终端交采'],  ['','标准表'],['','误差'], ['','变化量'],['','召测时间']], HARDWAREROWLIST,"测试结论")
        else:pass
        write_excel_xlsx_hardware(path, [hardwarerating,rtudatalist,meterdatalist,errorlist,valist,timelist], HARDWAREROWLIST,"测试结论")
        return True


def autoexhibition(name):
    nowtime = datetime.datetime.now()
    hardwarereportname = '/' + name + nowtime.strftime('%Y-%m-%d %H-%M-%S-%f')[2:16] + '.xlsx'
    return  hardwarereportname

def hardexhibition(name,path_experiment):
    nowtime = datetime.datetime.now()
    hardwarereportname = '/' + name + nowtime.strftime('%Y%m%d%H%M')[4:] + '.xlsx'
    path_experiment = path_experiment + nowtime.strftime('%Y%m%d%H%M')[4:]
    return  hardwarereportname,path_experiment
def reportspecialdeal(reportfloderpath,timesaved,overflag,staecandpceventnum):
    #rownum:测试到最后一行用例时，无论是什么时间，都再最后处理一次报告
    thour = int((datetime.datetime.now().strftime('%Y%m%d%H%M%S'))[-6:-4])
    min = int((datetime.datetime.now().strftime('%Y%m%d%H%M%S'))[-4:-2])
    # 当时间为XX小时的整数倍对报告进行分析。
    if len(reportfloderpath) > 0:
        if ((thour % 6 == 0) and (thour != timesaved) and ( min > 20)) or (overflag == 1) :
            stableanaly.anlaysispathreport(reportfloderpath,staecandpceventnum,overflag)
            timesaved = thour
    return timesaved
def staleoveranlay(testname,reportfloderpath,timesaved,staecandpceventnum):
    if testname == "能源稳定性测试":
        timesaved = reportspecialdeal(reportfloderpath, timesaved, 1,staecandpceventnum)
    #如果是计量冻结报告，直接调用函数即可
    elif testname in ['交采负荷记录','交采日冻结','计量结算日冻结','交采月冻结']:
        stableanaly.anlaysispathreport(reportfloderpath, staecandpceventnum, 1)
    return None


def autotestreport(autotestplan_excelrow,unqualitfieldflag,firsttestexcel,autoreportpath):
    if unqualitfieldflag == 1:
        result = "不合格"
    else:
        result = "合格"
    if firsttestexcel == 1:
        reportname = autoexhibition('自动化测试报告')
        autoreportpath = os.getcwd() + r"\report" + reportname
        excelcopyconfig(reportname,'自动化测试报告模板.xlsx')
    else:pass
    write_excel_xlsx_hardware(autoreportpath, [[autotestplan_excelrow['test_name']], [autotestplan_excelrow['test_sheet']], [result]],
                              HARDWAREROWLIST, "Sheet1")
    return autoreportpath

#定义一个计算2个值误差大于0.06就返回False的类。计量中，电量误差大于0.02就判不合格。会多处引用。
class Jcerror():
    def __init__(self):
        self.abserror = 0
    def erroresult(self,data1,data2):
        self.data1 = data1
        self.data2 = data2
        self.abserror = abs(self.data1 - self.data2)
        if self.abserror  > 0.06:
            return False
        else:
            return True
    def __str__(self):
        return f'预期值与实际值之差的绝对值为{"%.2f" % (self.abserror)}'
#定义一个计算2个值误差大于0.02就返回False的类。计量中，电量误差大于0.02就判不合格。会多处引用。
class Jccomerror():
    def __init__(self):
        self.abserror = 0
    def erresult(self,data1,data2):
        self.data1 = data1
        self.data2 = data2
        self.abserror = abs(self.data1 - self.data2)
        if self.abserror  > 0.02:
            return False
        else:
            return True
    def __str__(self):
        return f'预期值与实际值之差的绝对值为{"%.2f" % (self.abserror)}'

def ratejudge(ele_calculation,dealrow):
    elerightflag = 0
    for eachelenum in range(len(ele_calculation)):
        elecal = Jcerror()
        elere = elecal.erroresult(ele_calculation[eachelenum], strtolist(dealrow['real'])[eachelenum])
        if elere == False:
            logging.info(elecal)
            elerightflag = 1
            break
    return elerightflag
def data0get(dealrow,data):
    dataall = 0
    if dealrow['save'].find('FIX') >= 0:
        dataall = data
    else:
        if dealrow['param'].find('FIRST') >= 0:
            dataall = data
        elif dealrow['param'].find('SECOND') >= 0:
            dataall = data * 2
        elif dealrow['param'].find('THIRD') >= 0:
            dataall = data * 3
        elif dealrow['param'].find('FOURTH') >= 0:
            dataall = data * 4
        else:
            print('表格中填入参数错误')
    return dataall

def demanddataget(realdata):
    data = []
    for item in realdata:
        data.append(item[0])
    return data

def demdateget(realdata):
    date = []
    for item in realdata:
        date.append(item[1])
    return date


#这个函数用来判断电量是否发生倒走
def norewalkju(tasklist,dealrow,flagdata):
    expectdata = []
    realdata = []
    uflag = 0
    for itemrow in tasklist:
        if (tasklist[itemrow]['save'] == dealrow['save']) and (tasklist[itemrow]['param'] == flagdata):
            expectdata = strtolist(tasklist[itemrow]['real'])
            realdata = strtolist(dealrow['real'])
            if dealrow['save'].find('FIX') >= 0:
                realdata = demanddataget(realdata)
                expectdata = demanddataget(expectdata)
            for num in range(len(expectdata)):
                if realdata[num] < expectdata[num]:
                    uflag = 1
                    break
                else:pass

            if uflag == 1:
                logging.info('出现倒走')
                return '不合格'

            else:
                logging.info('没有倒走')
                return '合格'




def resultget(dataex,tasklist,dealrow):
    logging.info('走字预期值'+ str(dataex))
    #首先看是否倒走，没有倒走进一步判断值是否与预期在一定误差范围内
    dataerrflag = 0
    if dealrow['param'].find('FIRST') >= 0:
        pass
    else:
        if dealrow['param'].find('SECOND') >= 0:
            dealrow['result'] = norewalkju(tasklist,dealrow,'FIRST')
        elif dealrow['param'].find('THIRD') >= 0:
            dealrow['result'] = norewalkju(tasklist,dealrow,'SECOND')
        elif dealrow['param'].find('FOURTH') >= 0:
            dealrow['result'] = norewalkju(tasklist,dealrow,'THIRD')
        if dealrow['result'] == '不合格':
            return None
    if dealrow['save'].find('FIX') >= 0:
        # 处理需量值和时间
        realdata0 = strtolist(dealrow['real'])
        realdata = demanddataget(realdata0)
        demandate = demdateget(realdata0)
        for itemnum in range(len(dataex)):
            if abs(dataex[itemnum] - realdata[itemnum]) > 0.5:
                dataerrflag = 1
                dealrow['result'] = '不合格'
                logging.info('需量走字费率异常')
                break
            if (realdata[itemnum] > 0) and (demandate[itemnum] == '20000101000000') :
                dataerrflag = 1
                dealrow['result'] = '不合格'
                logging.info('需量发生时间异常')
                break
    else:
        realdata = strtolist(dealrow['real'])
        for itemnum in range(len(dataex)):
            if itemnum == 0:
                if abs(dataex[itemnum] - realdata[itemnum]) > 0.081:
                    dataerrflag = 1
                    dealrow['result'] = '不合格'
                    logging.info('总电量走字费率异常')
                    break
            else:
                if abs(dataex[itemnum] - realdata[itemnum]) > 0.021:
                    dataerrflag = 1
                    dealrow['result'] = '不合格'
                    logging.info('分费率电量走字费率异常')
                    break
    if dataerrflag == 0:
        dealrow['result'] = '合格'
    return None





def ratemovecal(dealrow,tasklist):
    dataex = [0,0,0,0,0]
    data = float(dealrow['save'].split(':')[1])
    dataex[0] = data0get(dealrow,data)
    if dealrow['expect'].find("费率1") >= 0:
        dataex[1] = data
    elif dealrow['expect'].find("费率2") >= 0:
        dataex[1] = data
        dataex[2] = data
    elif dealrow['expect'].find("费率3") >= 0:
        dataex[1] = data
        dataex[2] = data
        dataex[3] = data
    elif dealrow['expect'].find("费率4") >= 0:
        dataex[1] = data
        dataex[2] = data
        dataex[3] = data
        dataex[4] = data
    else:
        pass
    resultget(dataex,tasklist,dealrow)
    return  None



def comtonomalresult(dealrow,ele_calculation,ele,tasklist):
    if str(dealrow['save']).find('JC_E') >= 0:
        data1 = float(dealrow['expect'])
        data2 = float(strtolist(dealrow['real'])[0])
        #创建实例化对象
        jcerr = Jccomerror()
        result = jcerr.erresult(data1,data2)
        logging.info(jcerr)
        if result == True:
            dealrow['result'] = '合格'
        else:
            dealrow['result'] = '不合格'
    elif str(dealrow['save']).find('ELE_GET') >= 0:
        elerightflag = 0
        if dealrow['expect'].find("费率1") >= 0:
            ele_calculation[0] = ele_calculation[0] + float(ele)
            ele_calculation[1] = ele_calculation[1] + float(ele)
            elerightflag = ratejudge(ele_calculation, dealrow)
        elif dealrow['expect'].find("费率2") >= 0:
            ele_calculation[0] = ele_calculation[0] + float(ele)
            ele_calculation[2] = ele_calculation[2] + float(ele)
            elerightflag = ratejudge(ele_calculation, dealrow)
        elif dealrow['expect'].find("费率3") >= 0:
            ele_calculation[0] = ele_calculation[0] + float(ele)
            ele_calculation[3] = ele_calculation[3] + float(ele)
            elerightflag = ratejudge(ele_calculation, dealrow)
        elif dealrow['expect'].find("费率4") >= 0:
            ele_calculation[0] = ele_calculation[0] + float(ele)
            ele_calculation[4] = ele_calculation[4] + float(ele)
            elerightflag = ratejudge(ele_calculation, dealrow)
        else:pass
        if elerightflag == 1:
            dealrow['result'] = '不合格'
        else:
            dealrow['result'] = '合格'
            ele_calculation = strtolist(dealrow['real'])
        print(f'ele_calculation:{ele_calculation}')
        logging.info(f'ele_calculation:{ele_calculation}')
        #交采计量数据初始化，清空之前存好的数据，从0开始重新计算。
    elif str(dealrow['save']).find('CLEAR_ELE') >= 0:
        ele_calculation = [0,0,0,0,0]
        if dealrow['real'] == dealrow['expect']:
            dealrow['result'] = '合格'
        else:
            dealrow['result'] = '不合格'
    elif str(dealrow['save']).find('RATEMOVE') >= 0:
        ratemovecal(dealrow,tasklist)
    else:
        pass

    return ele_calculation

def sametobefore(tasklist,dealrow):
    for item in tasklist:
        try:
            if tasklist[item]['save'].find(',')>= 0:
                if tasklist[item]['save'].split(',')[1] == dealrow['expect'].split(',')[1]:
                    if dealrow['real'] != tasklist[item]['real']:
                        dealrow['result'] = '不合格'
                    else:
                        dealrow['result'] = '合格'
        except:
            dealrow['result'] = '不合格'
            continue
    return None



def comresult(dealrow,ele_calculation,ele,tasklist):
    if str(dealrow['save']).find('DEFAULT_VALUE') >= 0:
        if dealrow['real'] == dealrow['expect']:
            dealrow['result'] = '合格'
        else:
            dealrow['result'] = '不合格'
    elif str(dealrow['save']).find('SPECIALPARAM') >= 0:
        if (dealrow['oad_omd']) == '40020200':
            if len(dealrow['real']) == 14 :
                dealrow['result'] = '合格'
            else:
                dealrow['result'] = '不合格'
        elif dealrow['oad_omd'] == '40000200':
            if len(dealrow['real']) == 16 and dealrow['real'] != '00000000000000':
                if str(dealrow['save']).find('AUTO_SYSTEM_TIME') >= 0:
                    systime = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
                    print(f'systime:{systime}')
                    err_rtutime = timecompare(systime, dealrow['real'].strip("'"))
                    if differseconds(err_rtutime) <= int(dealrow['expect']):
                        dealrow['result'] = '合格'
                    else:
                        dealrow['result'] = '不合格'
                else:
                    dealrow['result'] = '合格'
            else:
                dealrow['result'] = '不合格'
                #地理坐标，不是全0,判合格。
        elif dealrow['oad_omd'] == '40040200':
            if dealrow['real'] != dealrow['expect']:
                dealrow['result'] = '合格'
            else:
                dealrow['result'] = '不合格'
    elif str(dealrow['save']).find('JC_E') >= 0:
        data1 = float(dealrow['expect'])
        data2 = float(strtolist(dealrow['real'])[0])
        #创建实例化对象
        jcerr = Jcerror()
        result = jcerr.erroresult(data1,data2)
        logging.info(jcerr)
        if result == True:
            dealrow['result'] = '合格'
        else:
            dealrow['result'] = '不合格'
    elif str(dealrow['save']).find('ELE_GET') >= 0:
        elerightflag = 0
        if dealrow['expect'].find("费率1") >= 0:
            ele_calculation[0] = ele_calculation[0] + float(ele)
            ele_calculation[1] = ele_calculation[1] + float(ele)
            elerightflag = ratejudge(ele_calculation, dealrow)
        elif dealrow['expect'].find("费率2") >= 0:
            ele_calculation[0] = ele_calculation[0] + float(ele)
            ele_calculation[2] = ele_calculation[2] + float(ele)
            elerightflag = ratejudge(ele_calculation, dealrow)
        elif dealrow['expect'].find("费率3") >= 0:
            ele_calculation[0] = ele_calculation[0] + float(ele)
            ele_calculation[3] = ele_calculation[3] + float(ele)
            elerightflag = ratejudge(ele_calculation, dealrow)
        elif dealrow['expect'].find("费率4") >= 0:
            ele_calculation[0] = ele_calculation[0] + float(ele)
            ele_calculation[4] = ele_calculation[4] + float(ele)
            elerightflag = ratejudge(ele_calculation, dealrow)
        else:pass
        if elerightflag == 1:
            dealrow['result'] = '不合格'
        else:
            dealrow['result'] = '合格'
            ele_calculation = strtolist(dealrow['real'])
        print(f'ele_calculation:{ele_calculation}')
        logging.info(f'ele_calculation:{ele_calculation}')
        #交采计量数据初始化，清空之前存好的数据，从0开始重新计算。
    elif str(dealrow['save']).find('CLEAR_ELE') >= 0:
        ele_calculation = [0,0,0,0,0]
        if dealrow['real'] == dealrow['expect']:
            dealrow['result'] = '合格'
        else:
            dealrow['result'] = '不合格'
    elif str(dealrow['save']).find('JUDGE_') >= 0:
        sametobefore(tasklist,dealrow)
    else:
        if (len(dealrow['real']) > 0) and (dealrow['real'] != u'终端没响应') and (dealrow['real'] != u'其它') and (
                dealrow['real'] != u'拒绝读写'):
            dealrow['result'] = u'合格'
        else:
            dealrow['result'] = u'不合格'

    return ele_calculation

#因广播地址时有特殊需求，故做处理。（AA\终端当前地址\非当前终端地址）
def addrget(ldata,iniRtu,frm):
    #'单地址': 0, '组地址': 2, '通配地址': 1, '广播地址': 3
    if frm['TSA_TYPE'] == 3 and ldata['save'] != 'NOTAA' and ldata['save'] != 'NOTADDR':
        frm['TSA_AD'] = 'AA'
    elif frm['TSA_TYPE'] == 1 :
        frm['TSA_AD'] = 'AAAAAAAAAAAA'
    elif frm['TSA_TYPE'] == 2 :
        frm['TSA_AD'] = '112222222222'
    else:
        if ldata['save'] == 'NOTADDR':
            frm['TSA_AD'] = '888111999666'
        else:
            frm['TSA_AD'] = iniRtu['ADDR']
    #先做死，给2940检测用，后面可以优化。
    if ldata['vadd'] == '终端通道5':
        frm['TSA_AD'] = Placefunction.ex2940tsaad()[0]
    elif ldata['vadd'] == '终端通道6':
        frm['TSA_AD'] = Placefunction.ex2940tsaad()[1]
    return frm['TSA_AD']

#在指定的位置创建文件夹
def mkdir(path):
    #  去除首位空格
    path=path.strip()
    #  去除尾部 \ 符号
    path=path.rstrip("\\")
    # 判断路径是否存在 # 存在 True # 不存在 False
    isExists=os.path.exists(path) # 判断结果
    # 如果不存在则创建目录 　# 创建目录操作函数
    if not isExists:
        os.makedirs(path)
        print(path+' 创建成功')
        return True
    else: # 如果目录存在则不创建，并提示目录已存在
        print(path+' 目录已存在')
        return False

#485抄表或者载波抄表时，出现汉字，说明抄表出现异常，回不合格。
class readmeterresult():
    def __init__(self):
        self.wrong1 = r'[\u4e00-\u9fa5]'
        self.wrong2 = r'NULL'
    def recresult(self,str):
        if re.search(self.wrong1, str) == None and re.search(self.wrong2, str) == None and len(str) > 0:
            return True
        else:
            return False
    def __del__(self):
        print("抄表结果异常判断流程readmeterresult，已经执行过了！")

def tempparamget(tempparamrow):
    host = '192.168.124.37'
    port = 50065
    host = strtolist(tempparamrow)[0][0]
    port = strtolist(tempparamrow)[0][1]
    return host,port

def visibletohex(s): # 任意字符以gb2312的编码格式，组帧成16进制
    cc = bytes(s, 'gb2312')
    ee = ""
    for x in cc:
        dd = str(hex(x))
        ee += dd[2:].zfill(2)
    return ee
def comresult_AT(testex_at,trytimesflag,trytimes):
    rxfff4 = 4
    txfff6 = 5
    if testex_at['param'].find(':MAC') >= 0:
        conflag = 0
        alldatalist = testex_at['real'].split("\r\n")
        for eachdata in alldatalist:
            # print(eachdata)
            if eachdata.find('FFF4') >= 0:
                rxfff4 = eachdata.split(':')[0]
                conflag = 1
            elif eachdata.find('FFF6') >= 0:
                txfff6 = eachdata.split(':')[0]
                conflag = 1
            else:
                pass
        print(f'rxfff4,txfff6{rxfff4,txfff6}')
        if conflag == 1:
            testex_at['result'] = '合格'
            testex_at['real'] = str(rxfff4) + ':' + str(txfff6)
            trytimes = 0
            trytimesflag = 0
        else:
            trytimes += 1
            trytimes_max = int(testex_at['save'].split(':')[1])
            if trytimes <= trytimes_max:
                trytimesflag = 1
                if trytimes == trytimes_max:
                    win32api.MessageBox(0, "蓝牙连接异常，请检查", "警告", win32con.MB_OK)
                else:
                    pass
            else:
                trytimes = 0
                trytimesflag = 0
            testex_at['result'] = '不合格'
            testex_at['real'] = '蓝牙连接失败'
    elif testex_at['param'].find('AT+CHRX') >= 0 or testex_at['param'].find('AT+CHTX') >= 0:
        # if testex_at['real'].find('OK') >= 0:
        #     testex_at['result'] = '合格'
        # else:
        testex_at['result'] = '合格'
    elif testex_at['param'].find('AT+RENEW') >= 0 or testex_at['param'].find('AT+DISCON') >= 0:
        testex_at['result'] = '合格'
    else:pass

    return trytimesflag,trytimes
#计算自动化程序从准备工作到结束工作所用时间。
def beattimescal(logstarttime):
    logendtime = datetime.datetime.now()
    timelong = logendtime - logstarttime
    beattimes = timelong.seconds//60 - 1
    print(f'心跳次数：{beattimes}')
    return beattimes

#将 把本终端地址本次统计的登录和心跳信息，复制到自动化的log文件夹中。
def logcopy698(sourcepath,targetname) :
    source = sourcepath
    targetpath = os.getcwd() + "\log\\698主站log\\"
    mkdir(targetpath)
    target = os.getcwd() + "\log\\698主站log\\" + targetname[:-4] +'主站日志'+ datetime.datetime.now().strftime('%Y%m%d%H%M%S') +'.txt'

    # try:
    copyfile(source, target)
    # except:
    #     print('copy file faild')
    return None


def log698deal(logpath,addr_experiment,logrow,logandbeat,logstarttime):
    filenamelist = []
    logop = logpath.split("->")[0]
    logpath = logpath.split("->")[1] + "\Log\\"
    print(f'addr_experiment:{addr_experiment}')
    logfilename = reversebit(addr_experiment[:12])
    print(f'logfilename:{logfilename}')
    filenamelist = file_name(logpath)
    if logop == '清空LOG':
        logstarttime = datetime.datetime.now()
        try:
            if not os.path.exists(logpath):
                logrow['real'] = '清空LOG文件成功'
                logrow['result'] = '合格'
            else:
                for item in filenamelist:
                    if item.find(logfilename) >= 0:
                        lognamepath = logpath + item
                        if not os.path.exists( lognamepath):
                            pass
                        else:
                            os.remove(lognamepath)
                logrow['real'] = '清空LOG文件成功'
                logrow['result'] = '合格'
        except:
            logrow['real'] = '清空LOG文件出现异常'
            logrow['result'] = '不合格'

    elif logop == '统计登录和心跳':
        try:
            beattimes = 0
            logtimes = 0
            for item in filenamelist:
                if item.find(logfilename) >= 0:
                    lognamepath = logpath + item
                    logcopy698(lognamepath, item)
                    f = open(lognamepath, "r",encoding='UTF-8')  # 设置文件对象
                    filecontent = f.read()  # 直接将文件中按行读到list里，效果与方法2一样
                    beattimes += filecontent.count('心跳')
                    logtimes += filecontent.count('登录')
                    f.close()
            print(beattimes,logtimes)
            logging.info('心跳次数'+str(beattimes))
            logging.info('登录次数' + str(logtimes))
            logrow['real'] = '登录：'+str(logtimes)+ '次；' + '心跳：'+str(beattimes)+ '次'
            beatexpecttimes = beattimescal(logstarttime)
            if (logtimes > 1) or (beattimes < beatexpecttimes) or (logtimes == 0) or (beattimes == 0):
                logandbeat = ['4G有异常', logrow['real']]
                logrow['result'] = '不合格'
            else:
                logandbeat = ['4G无异常', logrow['real']]
                logrow['result'] = '合格'
        except:
            logrow['real'] = '统计登录和心跳次数出现异常'
            logrow['result'] = '不合格'
            logandbeat = ['4G有异常', logrow['real']]
        # os.system("taskkill /F /IM 面向对象协议调试软件.exe")


    return logandbeat,logstarttime

def registertotal(online_off_time, online_off_name,registersavepath,online_off_time_temp, online_off_name_temp,reportfloderpath):
    #4G和以太网登录信息统计
    if registersavepath == ''  or registersavepath[-17:-9] !=  datetime.datetime.now().strftime('%Y%m%d%H%M')[:8] :
        registersavepath = reportfloderpath + '4G和以太网登录情况统计' + datetime.datetime.now().strftime('%Y%m%d%H%M') + ".xlsx"
    else:pass
    #主通道
    write_excel_xlsx_register(registersavepath, [online_off_name, online_off_time], '4G登录情况')
    online_off_time = []
    online_off_name = []
    #临时通道
    write_excel_xlsx_register(registersavepath, [online_off_name_temp, online_off_time_temp], '以太网登录情况')
    online_off_time_temp = []
    online_off_name_temp = []
    return online_off_time_temp, online_off_name_temp,online_off_time, online_off_name,registersavepath


def visave(data,path_experiment,name):
    vidatapath = os.getcwd() + r"\report\\" + f'{path_experiment}' + '\\' + f'{path_experiment}' + '终端及模块版本信息.xlsx'
    print('pathversion:', vidatapath)
    write_excel_xlsx_register(vidatapath, [[name], ['"' + data + '"']], '终端及模块版本信息')
    return None
#参数类用例比较时，将 NULL 和[]做统一处理。无论结果回NULL 还是[]，都判合格
def nulltoempty(ldata):
    if ldata['expect'] == '对象不存在' or ldata['expect'] == '[]':
        if ldata['real'] == '对象不存在' or ldata['real'] == '[]':
            return True
        else:
            return False
    else:
        return False

def minorhour(testname,minandhourlist):
    findflag = 0
    for item in minandhourlist:
        if testname.find(item) >= 0:
            findflag = 1
        else:pass
    if findflag == 1:
        return  True
    else:
        return False

#带标点符号的文字，筛选为只剩数字。例如：'0.94：怎么办；出现异常；出现异常。'-----0.94
def find_unchinese(file):
    pattern = re.compile(r'[\u4e00-\u9fa5]')
    punctuation = re.compile(r'[\u3002\uff1b\uff0c\uff1a\u201c\u201d\uff08\uff09\u3001\uff1f\u300a\u300b]')
    chinese = re.sub(pattern, '', file)
    num = re.sub(punctuation, '', chinese)
    return num

#通过终端类型和地区（及特殊型号），给出正确的最终终端类型列表。
def rtutypeget(rtutypeini,regionlist):
    rtutypelist = []
    if (rtutypeini == ['能源控制器公变']) and (regionlist == ['无']):
        rtutypelist = ['通用','能源通用','能源控制器公变','计量']
    elif (rtutypeini == ['能源控制器专变']) and (regionlist == ['无']):
        rtutypelist = ['通用','能源通用','能源控制器专变','计量']
    elif (rtutypeini == ['集中器']) and (regionlist == ['I型']):
        rtutypelist = ['通用','698通用','集中器','集中器I型']
    elif (rtutypeini == ['集中器']) and (regionlist == ['I型_安徽']):
        rtutypelist = ['通用', '698通用', '集中器', '集中器I型', '集中器I型_安徽']
    elif (rtutypeini == ['集中器']) and (regionlist == ['I型_湖北']):
        rtutypelist = ['通用','698通用','集中器','集中器I型','集中器I型_湖北']
    elif (rtutypeini == ['集中器']) and (regionlist == ['I型_山西']):
        rtutypelist = ['通用', '698通用', '集中器', '集中器I型', '集中器I型_山西']
    elif (rtutypeini == ['集中器']) and (regionlist == ['II型']):
        rtutypelist = ['通用','698通用','集中器','集中器II型']
    elif (rtutypeini == ['专变']) and (regionlist == ['I型']):
        rtutypelist = ['通用','698通用','专变','专变I型']
    elif (rtutypeini == ['专变']) and (regionlist == ['I型A4']):
        rtutypelist = ['通用','698通用','专变','专变I型','专变I型A4']
    elif (rtutypeini == ['专变']) and (regionlist == ['I型B8']):
        rtutypelist = ['通用', '698通用', '专变', '专变I型', '专变I型B8']
    elif (rtutypeini == ['专变']) and (regionlist == ['II型']):
        rtutypelist = ['通用','698通用','专变','专变II型']
    elif (rtutypeini == ['专变']) and (regionlist == ['II型B2']):
        rtutypelist = ['通用','698通用','专变','专变II型','专变II型B2']
    elif (rtutypeini == ['专变']) and (regionlist == ['II型B4']):
        rtutypelist = ['通用','698通用','专变','专变II型','专变II型B4']
    elif (rtutypeini == ['专变']) and (regionlist == ['III型']):
        rtutypelist = ['通用','698通用','专变','专变III型']
    elif (rtutypeini == ['专变']) and (regionlist == ['III型_安徽']):
        rtutypelist = ['通用','698通用','专变','专变III型','安徽']
        #以下为硬件自动化使用
    elif (rtutypeini == ['专变']) and (regionlist == ['698主版本']):
        rtutypelist = ['通用','698通用','专变']
    elif (rtutypeini == ['集中器']) and (regionlist == ['698主版本']):
        rtutypelist = ['通用','698通用','公变']
    elif (rtutypeini == ['I型专变']) and (regionlist == ['698主版本']):
        rtutypelist = ['通用','698通用','专变','I型专变']
    elif (rtutypeini == ['III型专变']) and (regionlist == ['698主版本']):
        rtutypelist = ['通用', '698通用', '专变', 'III型专变']
    elif (rtutypeini == ['专变']) and (regionlist == ['能源控制器']):
        rtutypelist = ['通用', '能源通用', '专变']
    elif (rtutypeini == ['公变']) and (regionlist == ['能源控制器']):
        rtutypelist = ['通用', '能源通用', '公变']
    elif (rtutypeini == ['控制扩展模块']) and (regionlist == ['698主版本']):
        rtutypelist = ['通用', '698通用', '专变','控制扩展模块','I型专变']
    return rtutypelist

#处理一些结果有特殊要求的标识
#ipiid行号,exceltasklist 原测试表格
def inresult(ldata,ipiid,exceltasklist):
    if ldata['oad_omd'] == '81070400' and (ldata['save'].find('E_A_OR_T') < 0):
        if ldata['expect'].find(ldata['real']) >= 0:
            ldata['result'] = u'合格'
        else:
            ldata['result'] = u'不合格'
    elif (ldata['oad_omd'][:4] == 'F205') and (ldata['save'].find('E_A_OR_T') < 0):
        if len(strtolist(ldata['expect'])) == 2 and ldata['real'].find(ldata['expect'].strip('[]')) >= 0:
            ldata['result'] = u'合格'
        else:
            ldata['result'] = u'不合格'
    #购电控流程，验证剩余电量等于跳闸门限，执行告警逻辑或者跳闸逻辑都可以（不能无动作）
    elif ldata['save'].find('E_A_OR_T') >= 0:
        rowneed = 0
        if (ldata['oad_omd'] == '81070500') or (ldata['oad_omd'] == '81080500'):
            rowneed = 1
        elif ldata['oad_omd'] == 'F2060200':
            rowneed = 2
        elif (ldata['oad_omd'] == '81070400') or (ldata['oad_omd'] == '81080400'):
            rowneed = 3
        elif ldata['oad_omd'] == 'F2050201':
            rowneed = 4
        elif ldata['oad_omd'] == 'F2050202':
            rowneed = 5
        if ldata['oad_omd'] == '23011100':
            if ldata['real'] in ldata['expect'].split('/'):
                ldata['result'] = u'合格'
            else:
                ldata['result'] = u'不合格'
        elif ldata['oad_omd'] in['F2060200','81070500','81070400','81080500','81080400'] :
            if (ldata['real'] == (ldata['expect'].split('/'))[0]) and ((strtolist(exceltasklist[ipiid - rowneed]['real'])[4] == '03' ) or (strtolist(exceltasklist[ipiid - rowneed]['real'])[3] == '03' )) :
                ldata['result'] = u'合格'
            elif (ldata['real'] == (ldata['expect'].split('/'))[1]) and ((strtolist(exceltasklist[ipiid - rowneed]['real'])[6] == '02' ) or (strtolist(exceltasklist[ipiid - rowneed]['real'])[6] == '01' ) ) :
                ldata['result'] = u'合格'
            else:
                ldata['result'] = u'不合格'
        elif (ldata['oad_omd'][:4] == 'F205') :
            if (ldata['real'].find(ldata['expect'].split('/')[0].strip('[]')) >= 0) and ((strtolist(exceltasklist[ipiid - rowneed]['real'])[4] == '03' ) or (strtolist(exceltasklist[ipiid - rowneed]['real'])[3] == '03' )):
                ldata['result'] = u'合格'
            elif (ldata['real'].find(ldata['expect'].split('/')[1].strip('[]')) >= 0) and ((strtolist(exceltasklist[ipiid - rowneed]['real'])[6] == '02' ) or (strtolist(exceltasklist[ipiid - rowneed]['real'])[6] == '01' ) ):
                ldata['result'] = u'合格'
            else:
                ldata['result'] = u'不合格'
    elif ldata['param'].find('DFC_EXPECT') >= 0:
        if ldata['real'].find(ldata['expect']) >= 0:
            ldata['result'] = u'合格'
        else:
            ldata['result'] = u'不合格'
    else:
        ldata['result'] = u'不合格'




    return None

#a_dict：原始字典；要插入的key; 想要在pos_key前插入；value要插入的值
def insert_key_value(a_dict, key, pos_key, value):
    new_dict = {}
    for k, v in a_dict.items():
        if k==pos_key:
            new_dict[key] = value  # insert new key
        new_dict[k] = v
    return new_dict

def equalreal(realrow,testexcel):
    for itemrow in  testexcel:
        if testexcel[itemrow]['save'] == realrow['expect']:
            if testexcel[itemrow]['real'] == realrow['real']:
                realrow['result'] = '合格'
            else:
                realrow['result'] = '不合格'
    return None

def realtimeexpect(tasklist,rowtest,rtu_current_time,rtu_time_list,testtime,low_ele):
    if (rowtest['param'].find(rtu_time_list) >= 0):
        rowtest['real']= rowtest['real']+'\n终端当前时间:'+ str(bcdtodatetime(rtu_current_time))
    if (rowtest['save'].find('_EXPECT') >= 0):
        #剩余电量，real值为一个数值
        # if rowtest['real'].find('[') < 0 and rowtest['real'].find("'") < 0:
        #     if float(rowtest['real']) != 0:
        #         rowtest['result'] = '合格'
        #     else:
        #         rowtest['result'] = '不合格'
        # #跌落前，预期值判断23010900电量，例如：[1.37,0.0,0.0,1.37,0.0]
        if (rowtest['real'].find('[') >= 0) and  (rowtest['oad_omd'] == '23010900') and rowtest['param'].find('REGULAR_V') < 0:
            if float(strtolist(rowtest['real'])[0]) == float(strtolist(rowtest['expect'])[0]):
                rowtest['result'] = '合格'
            else:
                rowtest['result'] = '不合格'
        #需量：[[4.2794,'20220311111200'],[0.0,'20000101000000'],[0.0,'20000101000000'],[4.2794,'20220311111200'],[0.0,'20000101000000']]
        if rowtest['real'].find('[[') >= 0:
            if float(strtolist(rowtest['real'])[0][0]) != 0:
                rowtest['result'] = '合格'
            else:
                rowtest['result'] = '不合格'
        elif rowtest['oad_omd'] in['45000500','F2090201','F20D0200']:
            if (rowtest['real'] != '') and (rowtest['real'] != []) and (rowtest['real'].find('异常') >= 0):
                rowtest['result'] = '不合格'
            else:
                rowtest['result'] = '合格'
        if rowtest['param'].find('REGULAR_V') >= 0:
            if rowtest['oad_omd'] == '00100200':
                if float(strtolist(rowtest['real'])[0]) > low_ele[0]:
                    rowtest['result'] = '合格'
                    low_ele[0] = float(strtolist(rowtest['real'])[0])
                else:
                    rowtest['result'] = '不合格'
            elif rowtest['oad_omd'] == '00300200':
                if float(strtolist(rowtest['real'])[0]) > low_ele[1]:
                    rowtest['result'] = '合格'
                    low_ele[1] = float(strtolist(rowtest['real'])[0])
                else:
                    rowtest['result'] = '不合格'
            elif rowtest['oad_omd'] == '23010B00':
                rowtest['expect'] = str(low_ele[2])
                if float(rowtest['real']) == low_ele[2]:
                    rowtest['result'] = '合格'
                else:
                    rowtest['result'] = '不合格'
            elif rowtest['oad_omd'] == '23010900':
                rowtest['expect'] = str(low_ele[3])
                if float(strtolist(rowtest['real'])[0]) == low_ele[3]:
                    rowtest['result'] = '合格'
                else:
                    rowtest['result'] = '不合格'
     #预期值为上面行的实际值的结果处理流程
    if rowtest['expect'].find('_EXPECT') >= 0:
        equalreal(rowtest, tasklist)
    print('质量底线预期值（00100200,00300200）low_ele:', low_ele)
    logging.info('质量底线预期值（00100200,00300200）low_ele:' + str(low_ele))
    return low_ele

def loweleget(dataparam,low_ele,testname,testtime):
    data = dataparam.split('；')
    num = data[1].split('：')[1]
    low_ele[2] = low_ele[2] - int(num)
    if (testname == '硬件复位') and (testtime != 0) :
        low_ele[3] = 250
    low_ele[3] = low_ele[3] + int(num)
    print('质量底线预期值（剩余电量和总加组月有功）low_ele:',low_ele)
    logging.info('质量底线预期值（剩余电量和总加组月有功）low_ele:'+ str(low_ele))
    return low_ele

def tryagain(comresult,comtrytimes,tryflag):
    if (comresult == '通讯异常') and (tryflag == 0):
        comtrytimes =1
    return comtrytimes





def repeattimeget(testrow,repeatnow,repeattimes,repeatresult,paramold,paramoldlist,test_plan,paramfreezelist,nrowcir,testrownum,tasklist,testrownow,palannumlist):
    #因为参数处理流程比较复杂，如果涉及到重复执行，就先恢复一下原来的参数格式
    if (repeatnow == 0) and (testrow['save'].find('CIR') < 0) :
        paramold = testrow['param']
    elif (repeatnow >= 1) and (testrow['save'].find('CIR') < 0):
        testrow['param'] = paramold
    #循环次数处理
    if ((repeattimes == 0) and (testrow['save'].find('REPEAT_EXECUTION') >= 0))and (testrow['param'].find('召测参数列表') <0 ):
        repeattimes = Placefunction.repeattimecal(testrow)
    elif testrow['param'].find('召测参数列表') >= 0:
        if  paramfreezelist == []:
            paramfreezelist= excelchange.typetimes(test_plan,testrow['param'])
            repeattimes,paramfreezelist,palannumlist = Placefunction.calltime(paramfreezelist,testrow['param'],palannumlist)
        print('repeattimes:',repeattimes)
        testrow['param'] = Placefunction.param6012deal(paramfreezelist[repeatnow])
        testrow['expect'] = palannumlist[repeatnow]
    elif testrow['save'].find('CIR') >= 0:
        if nrowcir == 0:
            repeattimes,nrowcir,paramoldlist = Placefunction.repeattrow(testrow,testrownum,tasklist,nrowcir,repeattimes,paramoldlist)
        if repeatnow > 0:
            testrow['param'] = paramoldlist[testrownow]
        testrownow += 1

    #循环结果处理
    if testrow['save'].find('REPEAT_EXECUTION') >= 0:
        repeatnow += 1
        #多次重复执行，有一次不合格，就判结果为不合格。
        if (testrow['result'] != '合格') and (testrow['result'] != ''):
            repeatresult = 1
    return repeattimes,repeatnow,repeatresult,paramold,paramoldlist,paramfreezelist,nrowcir,testrownow,palannumlist
# 等待
def wait(ldata):
    try:
        if isinstance(ldata, float):
            idelay = ldata
        else:
            idelay = float(ldata.encode('utf-8'))
        time.sleep(idelay)
    except ValueError:
        return False
    return True

def delaywork(tasklistrow):
    #延时列要起作用的一些行
    if tasklistrow['op'] == u'SSH操作' or tasklistrow['op'] == u'TELNET操作' or tasklistrow['op'] == u'蓝牙':
        if isinstance(tasklistrow['delay'], float):
            wait(tasklistrow['delay'])
        elif isinstance(tasklistrow['delay'], str):
            wait(float(tasklistrow['delay']))
        else:
            logging.info('excel delay type must check!')
    elif str(tasklistrow['delay']).find('JUSTSTOP') >= 0:
        waittime = float(tasklistrow['delay'].split(":")[1])
        wait(waittime)
    else:
        pass







if __name__ == '__main__':
    s = b'h0\x00\x01\x05\x03\x00\x00\x00\x00\x00\x00\xe9\xee\x81\x04\x01\x07\xe4\x01\n\x05\t\x07\r\x00\x00\x07\xe4\x01' \
        b'\n\x04\t\x06,\xea\xbf\x07\xe4\x01\n\x04\t\x06,\xea\xbf?\x05\x16'
    print(hexShow(s))
    sh = '687A00830503000000000000496B880202013106020006002022020000201E020000202002000020240200003300020000330902060' \
         '10106000000061C07E4010A1136161C07E4010A11370716010104020251450000001100020251451000001100020251F20002011100020251F20C020111000408000000C5AF16'
    print(hexascii(sh))
